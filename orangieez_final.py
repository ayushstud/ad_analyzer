"""
AdLens — V2 UPGRADED EDITION
==============================
PHASE 1  Real feature columns with REALISTIC correlations in dataset
PHASE 2  Auto-feedback loop — /feedback endpoint (SQLite)
PHASE 3  Advanced CV: sharpness, dominant color (KMeans), symmetry,
         color psychology (blue=trust, red=urgency, orange=CTA),
         face detection, text area %, brightness/contrast
PHASE 4  Florence-2 replaces BLIP (falls back gracefully)
         SigLIP replaces CLIP for semantic scoring (falls back to CLIP)
PHASE 5  Multi-Model Ensemble: 0.45×XGBoost + 0.35×Vision + 0.20×Copy
PHASE 6  Industry Benchmarks — percentile ranks from dataset
PHASE 7  A/B Variant Generator — AI suggests optimized variants
PHASE 8  Rule-based Psychology Engine (universal ad principles)
PHASE 9  Ranking model: "better than X% of ads"
"""
import ssl
ssl._create_default_https_context = ssl._create_unverified_context
import re, os, io, json, base64, time, sqlite3, datetime
import numpy as np
import openpyxl
import torch
device = "cuda" if torch.cuda.is_available() else "cpu"
print(f"[AdLens] Using device: {device}")
from PIL import Image, ImageStat, ImageFilter, ImageEnhance
from flask import Flask, request, render_template_string, jsonify
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, r2_score
import urllib.request

# ── Optional deps — graceful fallback ───────────────────────────────────────
try:
    import shap
    SHAP_AVAILABLE = True
except ImportError:
    SHAP_AVAILABLE = False
    print("[AdLens] shap not installed — pip install shap")

try:
    import easyocr
    _ocr_reader = easyocr.Reader(['en'], gpu=False)
    OCR_BACKEND = "easyocr"
except ImportError:
    try:
        import pytesseract
        OCR_BACKEND = "pytesseract"
    except ImportError:
        OCR_BACKEND = None

try:
    from sklearn.cluster import KMeans
    KMEANS_AVAILABLE = True
except ImportError:
    KMEANS_AVAILABLE = False

# ── PHASE 4: Best available regressor ───────────────────────────────────────
try:
    from xgboost import XGBRegressor
    REGRESSOR_NAME = "XGBoost"
    def make_regressor():
        return XGBRegressor(
            n_estimators=300, max_depth=6, learning_rate=0.05,
            subsample=0.85, colsample_bytree=0.85,
            min_child_weight=3, gamma=0.1,
            random_state=42, verbosity=0
        )
except ImportError:
    try:
        import lightgbm as lgb
        REGRESSOR_NAME = "LightGBM"
        def make_regressor():
            return lgb.LGBMRegressor(n_estimators=300, max_depth=6,
                                     learning_rate=0.05, random_state=42, verbose=-1)
    except ImportError:
        from sklearn.ensemble import GradientBoostingRegressor
        REGRESSOR_NAME = "GradientBoosting"
        def make_regressor():
            return GradientBoostingRegressor(n_estimators=200, max_depth=5,
                                             learning_rate=0.06, random_state=42)

print(f"[AdLens] Using regressor: {REGRESSOR_NAME}")

# ── PHASE 4: Florence-2 (primary) → SigLIP → CLIP (fallback chain) ──────────
# Florence-2 disabled — known bug with current transformers on Windows CPU
# ('Florence2LanguageConfig' has no attribute 'forced_bos_token_id')
# Using SigLIP directly instead — equal quality for ad semantic scoring.
FLORENCE_AVAILABLE = False
SIGLIP_AVAILABLE   = False
CLIP_AVAILABLE     = False

florence_model = florence_processor = None
siglip_model   = siglip_processor   = None
clip_model     = clip_processor     = None

# Load SigLIP (primary)
try:
    from transformers import AutoProcessor, AutoModel
    print("[AdLens] Loading SigLIP...")
    siglip_processor = AutoProcessor.from_pretrained("google/siglip-base-patch16-224")
    siglip_model     = AutoModel.from_pretrained("google/siglip-base-patch16-224").to(device)
    SIGLIP_AVAILABLE = True
    print("[AdLens] SigLIP loaded successfully")
except Exception as e:
    print(f"[AdLens] SigLIP not available ({e}) — trying CLIP")

# Load CLIP (fallback if SigLIP fails)
if not SIGLIP_AVAILABLE:
    try:
        from transformers import CLIPProcessor, CLIPModel
        print("[AdLens] Loading CLIP (fallback)...")
        clip_processor = CLIPProcessor.from_pretrained("openai/clip-vit-base-patch32")
        clip_model     = CLIPModel.from_pretrained("openai/clip-vit-base-patch32").to(device)
        CLIP_AVAILABLE = True
        print("[AdLens] CLIP loaded successfully")
    except Exception as e:
        print(f"[AdLens] CLIP not available ({e})")

# What VLM is active
if FLORENCE_AVAILABLE:
    VLM_NAME = "Florence-2"
elif SIGLIP_AVAILABLE:
    VLM_NAME = "SigLIP"
elif CLIP_AVAILABLE:
    VLM_NAME = "CLIP"
else:
    VLM_NAME = "None"

# Keep BLIP as last-resort caption fallback
BLIP_AVAILABLE = False
blip_model = blip_processor_inst = None
try:
    from transformers import BlipProcessor, BlipForConditionalGeneration
    blip_processor_inst = BlipProcessor.from_pretrained("Salesforce/blip-image-captioning-base")
    blip_model = BlipForConditionalGeneration.from_pretrained(
        "Salesforce/blip-image-captioning-base").to(device)
    BLIP_AVAILABLE = True
    print("[AdLens] BLIP loaded as caption fallback")
except Exception as e:
    print(f"[AdLens] BLIP not available ({e})")

# ── OpenCV face + sharpness ──────────────────────────────────────────────────
CV2_AVAILABLE = False
face_cascade  = None
try:
    import cv2
    _cascade_path = cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
    face_cascade  = cv2.CascadeClassifier(_cascade_path)
    CV2_AVAILABLE = True
    print("[AdLens] OpenCV ready (face + sharpness)")
except ImportError:
    print("[AdLens] cv2 not installed — pip install opencv-python")


app = Flask(__name__)
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
CLAUDE_MODEL = os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-20250514")

# ─────────────────────────────────────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "adlens_history.db")

def init_db():
    con = sqlite3.connect(DB_PATH)
    con.execute("""
        CREATE TABLE IF NOT EXISTS analyses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts TEXT, ad_copy TEXT, ad_type INTEGER, audience INTEGER,
            platform INTEGER, score INTEGER, ml_score INTEGER, vlm_score INTEGER,
            copy_score INTEGER, feedback TEXT, insights TEXT, suggestions TEXT,
            ctr_estimate REAL, face_count INTEGER, text_area_pct REAL,
            brightness REAL, sharpness REAL, symmetry_score REAL,
            dominant_color TEXT, color_psychology TEXT,
            siglip_luxury REAL, siglip_trust REAL, siglip_energy REAL,
            percentile_rank REAL, ensemble_breakdown TEXT,
            user_rating INTEGER, actual_ctr REAL, actual_conversion REAL,
            prediction_accurate INTEGER
        )""")
    con.execute("""
        CREATE TABLE IF NOT EXISTS feedback (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            analysis_id INTEGER NOT NULL,
            ts TEXT NOT NULL, rating INTEGER NOT NULL, comment TEXT,
            FOREIGN KEY (analysis_id) REFERENCES analyses(id)
        )""")
    con.commit(); con.close()

def _add_col(con, table, col, dtype):
    try: con.execute(f"ALTER TABLE {table} ADD COLUMN {col} {dtype}")
    except sqlite3.OperationalError: pass

def save_analysis(data: dict):
    con = sqlite3.connect(DB_PATH)
    cols = ", ".join(data.keys())
    placeholders = ", ".join(["?"] * len(data))
    cur = con.execute(
        f"INSERT INTO analyses ({cols}) VALUES ({placeholders})", list(data.values()))
    aid = cur.lastrowid
    con.commit(); con.close()
    return aid

def save_feedback(analysis_id, rating, comment=""):
    con = sqlite3.connect(DB_PATH)
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    con.execute("INSERT INTO feedback (analysis_id, ts, rating, comment) VALUES (?,?,?,?)",
                (analysis_id, ts, rating, comment))
    # prediction_accurate is a boolean: 1 if user says prediction was accurate (rating=1), else 0
    prediction_accurate = 1 if rating == 1 else 0
    con.execute("UPDATE analyses SET user_rating=?, prediction_accurate=? WHERE id=?",
                (rating, prediction_accurate, analysis_id))
    con.commit(); con.close()

def load_history(limit=50):
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    rows = con.execute("SELECT * FROM analyses ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
    con.close()
    return [dict(r) for r in rows]

def get_feedback_stats():
    con = sqlite3.connect(DB_PATH)
    total    = con.execute("SELECT COUNT(*) FROM feedback").fetchone()[0]
    positive = con.execute("SELECT COUNT(*) FROM feedback WHERE rating=1").fetchone()[0]
    con.close()
    if total == 0: return {"total": 0, "accuracy_rate": None}
    return {"total": total, "accuracy_rate": round(positive / total * 100, 1)}

init_db()

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 1: IMPROVED DATASET TRAINING — realistic correlations
# ─────────────────────────────────────────────────────────────────────────────
DATASET_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ad_dataset_100k.xlsx")

# Extended feature set — uses MORE real dataset columns
FEATURE_NAMES = [
    "Headline Len", "Ad Type", "Color Score", "Audience", "Platform",
    "Hour", "Day", "Audience×Platform Match", "Peak Hour Flag",
    "Visual Sentiment Boost", "CTR (raw)", "Brightness", "Emotional Impact",
    "Text Density", "Contrast", "Budget (USD)", "Engagement Rate %",
    "Sentiment", "Campaign Days",
]

FEATURE_DISPLAY = [
    "Headline Length", "Ad Type", "Color Score", "Audience", "Platform",
    "Hour", "Day", "Audience×Platform Match", "Peak Hour Flag",
    "Visual Sentiment Boost", "CTR Signal", "Brightness", "Emotional Impact",
    "Text Density", "Contrast", "Budget", "Engagement Rate",
    "Sentiment", "Campaign Duration",
]

FEATURE_DEFAULTS = [8, 1, 5, 3, 1, 18, 3, 0, 0, 0, 0.03, 128, 6, 5, 30, 5000, 3, 1, 30]

# Store dataset scores for percentile ranking
DATASET_SCORES = []

def load_and_train():
    global DATASET_SCORES
    wb = openpyxl.load_workbook(DATASET_PATH)
    ws = wb["Sheet1"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    X, y = [], []
    scores_raw = []

    for r in range(2, ws.max_row + 1):
        row = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        try:
            # PHASE 1: realistic correlations — use ALL available columns
            sentiment_raw = row.get("Sentiment") or 1
            # Normalize sentiment: can be numeric (1-5) or label
            if isinstance(sentiment_raw, str):
                sentiment_map = {"Positive": 3, "Neutral": 2, "Negative": 1}
                sentiment_val = sentiment_map.get(sentiment_raw, 2)
            else:
                sentiment_val = float(sentiment_raw)

            feats = [
                float(row.get("Headline Len")              or 8),
                float(row.get("Ad Type")                   or 1),
                float(row.get("Color Score")               or 5),
                float(row.get("Audience")                  or 3),
                float(row.get("Platform")                  or 1),
                float(row.get("Hour")                      or 18),
                float(row.get("Day")                       or 3),
                float(row.get("Audience\u00d7Platform Match") or 0),
                float(row.get("Peak Hour Flag")            or 0),
                float(row.get("Visual Sentiment Boost")    or 0),
                float(row.get("CTR (raw)")                 or 0.03),
                float(row.get("Brightness")                or 128),
                float(row.get("Emotional Impact")          or 6),
                # NEW columns for better accuracy
                float(row.get("Text Density")              or 5),
                float(row.get("Contrast")                  or 30),
                float(row.get("Budget (USD)")              or 5000),
                float(row.get("Engagement Rate %")         or 3),
                float(sentiment_val),
                float(row.get("Campaign Days")             or 30),
            ]
            target = float(row.get("Final Score") or 50)
            X.append(feats)
            y.append(target)
            scores_raw.append(target)
        except (TypeError, ValueError):
            continue

    DATASET_SCORES = sorted(scores_raw)
    X, y = np.array(X), np.array(y)
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    scaler = StandardScaler()
    X_train_s = scaler.fit_transform(X_train)
    X_test_s  = scaler.transform(X_test)

    clf = make_regressor()
    clf.fit(X_train_s, y_train)

    preds = clf.predict(X_test_s)
    mae = round(mean_absolute_error(y_test, preds), 2)
    r2  = round(r2_score(y_test, preds), 3)

    explainer = None
    if SHAP_AVAILABLE:
        try: explainer = shap.TreeExplainer(clf)
        except Exception as e: print(f"[AdLens] SHAP init failed: {e}")

    print(f"[AdLens] {REGRESSOR_NAME} trained | MAE={mae} | R²={r2} | "
          f"Train={len(X_train)} | Test={len(X_test)}")
    return clf, scaler, explainer, mae, r2, len(X_train), len(X_test)

import joblib

joblib.dump(clf, "model.pkl")
joblib.dump(scaler, "scaler.pkl")

print("[AdLens] Loading dataset and training model...")
try:
    model, scaler, SHAP_EXPLAINER, MODEL_MAE, MODEL_R2, TRAIN_SIZE, TEST_SIZE = load_and_train()
except FileNotFoundError:
    print(f"[AdLens] WARNING: Dataset not found at {DATASET_PATH} — using synthetic fallback model.")
    from sklearn.ensemble import GradientBoostingRegressor as _FallbackReg
    import numpy as _np_fb
    _X_fb = _np_fb.random.rand(500, len(FEATURE_NAMES)) * 10
    _y_fb = _np_fb.clip(_X_fb.mean(axis=1) * 10, 5, 100)
    scaler = StandardScaler()
    _X_fb_s = scaler.fit_transform(_X_fb)
    model = _FallbackReg(n_estimators=50, random_state=42)
    model.fit(_X_fb_s, _y_fb)
    SHAP_EXPLAINER = None
    MODEL_MAE = "N/A"
    MODEL_R2 = "N/A"
    TRAIN_SIZE = 0
    TEST_SIZE = 0
    DATASET_SCORES = list(range(0, 101))
except Exception as _e:
    print(f"[AdLens] ERROR loading dataset: {_e}")
    raise

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 9: PERCENTILE RANKING
# ─────────────────────────────────────────────────────────────────────────────
def get_percentile_rank(score):
    """Returns how much % of dataset ads this score beats."""
    if not DATASET_SCORES:
        return None
    below = sum(1 for s in DATASET_SCORES if s < score)
    return round(below / len(DATASET_SCORES) * 100, 1)

# ─────────────────────────────────────────────────────────────────────────────
# ENGINEERED FEATURES
# ─────────────────────────────────────────────────────────────────────────────
HIGH_SYNERGY = {(4,2),(1,1),(5,1),(1,4),(3,3),(2,4),(4,4),(5,3),(6,2)}

def get_engineered_features(ad_type, audience, platform, hour, text_density, emotional_impact):
    apm = 1 if (platform, audience) in HIGH_SYNERGY else 0
    phf = 1 if 18 <= hour <= 21 else 0
    vsb = round(emotional_impact * 1.2 - text_density * 0.5, 2)
    return apm, phf, vsb

# ─────────────────────────────────────────────────────────────────────────────
# SHAP
# ─────────────────────────────────────────────────────────────────────────────
def get_shap_insights(features_scaled, features_raw):
    if not SHAP_AVAILABLE or SHAP_EXPLAINER is None:
        return []

    try:
        sv = SHAP_EXPLAINER.shap_values(features_scaled)

        if isinstance(sv, list):
            sv = sv[0]

        sv = sv[0] if sv.ndim > 1 else sv

        contributions = sorted(
            zip(FEATURE_DISPLAY, sv),
            key=lambda x: abs(x[1]),
            reverse=True
        )

        insights = []

        for name, val in contributions[:3]:

            impact = abs(round(val, 1))

            if impact < 1:
                continue

            if name == "Emotional Impact" and val > 0:
                insights.append("Strong emotional appeal detected")

            elif name == "CTR Signal" and val < 0:
                insights.append("Current layout may reduce engagement")

            elif name == "Contrast" and val > 0:
                insights.append("Good contrast improves visibility")

            elif name == "Text Density" and val < 0:
                insights.append("Too much text may reduce performance")

            elif name == "Visual Sentiment Boost" and val > 0:
                insights.append("Visual style is engaging and attention-grabbing")

            elif val > 0:
                insights.append(f"{name} positively impacts performance")

            else:
                insights.append(f"{name} may reduce ad effectiveness")

        return insights

    except Exception as e:
        print(f"[AdLens] SHAP error: {e}")
        return []

# ─────────────────────────────────────────────────────────────────────────────
# AD COPY ANALYZER
# ─────────────────────────────────────────────────────────────────────────────
POSITIVE_WORDS = {'amazing','best','free','new','save','win','exclusive','limited','offer',
                  'guaranteed','proven','trusted','top','premium','instant','easy','fast',
                  'boost','grow','improve','powerful','smart','innovative','leading'}
NEGATIVE_WORDS = {'bad','worst','fail','lose','costly','expensive','difficult','hard',
                  'problem','issues','trouble','risk','danger','avoid','scam','fake'}
CTA_PHRASES    = ['shop now','buy now','get started','sign up','learn more','try free',
                  'download','subscribe','order now','claim','book','join','get offer',
                  'start free','explore','see more','apply now','contact us']
URGENCY_WORDS  = ['today','now','limited','hurry','expires','last chance','only',
                  'urgent','flash','ending','soon','24 hours','hours left']

def analyze_ad_copy(text):
    if not text: return None
    words = text.lower().split()
    pos   = sum(1 for w in words if w in POSITIVE_WORDS)
    neg   = sum(1 for w in words if w in NEGATIVE_WORDS)
    has_cta     = any(p in text.lower() for p in CTA_PHRASES)
    has_urgency = any(u in text.lower() for u in URGENCY_WORDS)
    sentiment_label = "Positive" if pos > neg else ("Negative" if neg > pos else "Neutral")
    avg_word_len    = sum(len(w) for w in words) / max(len(words), 1)
    readability     = round(max(0.0, 10 - avg_word_len), 1)
    # Spam: too many power words
    spam_ratio = pos / max(len(words), 1)
    is_spam    = spam_ratio > 0.4

    copy_score = 50
    copy_score += 15 if has_cta else 0
    copy_score += 10 if has_urgency else 0
    copy_score += min(15, pos * 3)
    copy_score -= min(20, neg * 5)
    copy_score += min(10, readability)
    copy_score -= 10 if is_spam else 0
    copy_score -= 10 if len(words) > 25 else 0
    copy_score -= 5  if len(words) < 3 else 0
    copy_score = max(0, min(100, copy_score))

    return {
        'word_count': len(words), 'has_cta': has_cta, 'has_urgency': has_urgency,
        'sentiment_label': sentiment_label, 'readability': readability,
        'copy_score': copy_score, 'is_spam_like': is_spam,
        'spam_ratio': round(spam_ratio, 2),
    }

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 3: ADVANCED CV FEATURES
# ─────────────────────────────────────────────────────────────────────────────

# Color psychology mapping (dominant hue → meaning)
COLOR_PSYCHOLOGY = {
    "red":    {"meaning": "Urgency / Excitement",  "ctr_boost": 0.8,  "emoji": "🔴"},
    "orange": {"meaning": "CTA Energy / Warmth",   "ctr_boost": 0.9,  "emoji": "🟠"},
    "blue":   {"meaning": "Trust / Reliability",   "ctr_boost": 0.6,  "emoji": "🔵"},
    "green":  {"meaning": "Growth / Health",        "ctr_boost": 0.5,  "emoji": "🟢"},
    "yellow": {"meaning": "Optimism / Attention",  "ctr_boost": 0.7,  "emoji": "🟡"},
    "purple": {"meaning": "Luxury / Creativity",   "ctr_boost": 0.6,  "emoji": "🟣"},
    "black":  {"meaning": "Premium / Authority",   "ctr_boost": 0.3,  "emoji": "⚫"},
    "white":  {"meaning": "Clean / Minimal",       "ctr_boost": 0.4,  "emoji": "⚪"},
    "pink":   {"meaning": "Playful / Youthful",    "ctr_boost": 0.5,  "emoji": "🩷"},
    "neutral":{"meaning": "Balanced / Versatile",  "ctr_boost": 0.3,  "emoji": "🔘"},
}

def rgb_to_color_name(r, g, b):
    """Map RGB centroid to a named color using heuristics."""
    if r > 180 and g < 80 and b < 80:   return "red"
    if r > 200 and 100 < g < 170 and b < 80: return "orange"
    if b > 150 and r < 120 and g < 150: return "blue"
    if g > 140 and r < 120 and b < 120: return "green"
    if r > 200 and g > 200 and b < 100: return "yellow"
    if r > 150 and b > 150 and g < 100: return "purple"
    if r > 180 and g > 150 and b > 160: return "pink"
    if r < 60 and g < 60 and b < 60:   return "black"
    if r > 200 and g > 200 and b > 200: return "white"
    return "neutral"

def extract_dominant_colors(pil_image, n_colors=3):
    """PHASE 3: KMeans dominant color extraction — color psychology."""
    if not KMEANS_AVAILABLE:
        return "neutral", []
    try:
        img = pil_image.convert("RGB").resize((100, 100))
        pixels = np.array(img).reshape(-1, 3).astype(float)
        km = KMeans(n_clusters=n_colors, random_state=42, n_init=5)
        km.fit(pixels)
        centroids = km.cluster_centers_
        counts    = np.bincount(km.labels_)
        # Sort by frequency
        sorted_idx = np.argsort(counts)[::-1]
        dominant_rgb = centroids[sorted_idx[0]]
        color_names  = [rgb_to_color_name(*centroids[i].astype(int)) for i in sorted_idx]
        return color_names[0], color_names
    except Exception as e:
        print(f"[AdLens] KMeans color error: {e}")
        return "neutral", []

def compute_sharpness(pil_image):
    """PHASE 3: Image sharpness via Laplacian variance. Higher = sharper."""
    try:
        if CV2_AVAILABLE:
            import cv2
            img_np = np.array(pil_image.convert("L"))
            lap_var = cv2.Laplacian(img_np, cv2.CV_64F).var()
            return round(float(lap_var), 1)
        else:
            # PIL fallback
            gray = pil_image.convert("L")
            edges = gray.filter(ImageFilter.FIND_EDGES)
            stat = ImageStat.Stat(edges)
            return round(stat.var[0], 1)
    except Exception:
        return 50.0

def compute_symmetry(pil_image):
    """PHASE 3: Left/right brightness symmetry — 0-10, higher = more balanced."""
    try:
        gray = pil_image.convert("L").resize((100, 100))
        arr  = np.array(gray)
        left  = arr[:, :50].mean()
        right = arr[:, 50:].mean()
        diff  = abs(left - right)
        score = round(10 - min(10, diff / 10), 1)
        return score
    except Exception:
        return 5.0

def detect_faces(pil_image):
    if not CV2_AVAILABLE or face_cascade is None: return 0
    try:
        import cv2
        img_np = np.array(pil_image.convert("RGB"))
        gray   = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)
        faces  = face_cascade.detectMultiScale(gray, 1.1, 5, minSize=(30,30))
        return len(faces) if isinstance(faces, np.ndarray) else 0
    except Exception:
        return 0

def estimate_text_area_pct(pil_image):
    try:
        gray   = pil_image.convert("L")
        pixels = list(gray.getdata())
        high_contrast = sum(1 for p in pixels if p < 60 or p > 200)
        return round(high_contrast / len(pixels) * 100, 1)
    except Exception:
        return 0.0

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 4: FLORENCE-2 / SIGLIP / CLIP scoring
# ─────────────────────────────────────────────────────────────────────────────
VLM_CONCEPTS = {
    "luxury":     ["luxury product", "premium quality", "high-end brand", "elegant design"],
    "trust":      ["trustworthy brand", "professional service", "reliable", "credible business"],
    "energy":     ["exciting vibrant energy", "dynamic action", "high CTR ad", "bold colors"],
    "minimalist": ["clean minimal design", "white space layout", "simple clean ad"],
    "crowded":    ["cluttered design", "too much text", "busy overwhelming ad"],
    "modern":     ["modern trendy design", "contemporary aesthetic", "Gen Z style"],
    "emotional":  ["emotional story", "human connection", "touching inspiring"],
}

def florence_analyze(pil_image):
    """Florence-2: detailed caption + region understanding."""
    if not FLORENCE_AVAILABLE: return None
    try:
        inputs = florence_processor(
            text="<DETAILED_CAPTION>", images=pil_image, return_tensors="pt"
        ).to(device)
        with torch.no_grad():
            out = florence_model.generate(
                **inputs, max_new_tokens=120,
                do_sample=False, num_beams=3
            )
        result = florence_processor.batch_decode(out, skip_special_tokens=False)[0]
        # Extract caption
        caption = result.replace("<DETAILED_CAPTION>", "").replace("</s>", "").strip()
        return caption
    except Exception as e:
        print(f"[AdLens] Florence-2 error: {e}")
        return None

def siglip_score_concepts(pil_image):
    """SigLIP: semantic concept scoring using softmax normalization across concepts."""
    if not SIGLIP_AVAILABLE: return {}
    try:
        # Score every concept's best prompt against the image in one batch
        all_concepts = list(VLM_CONCEPTS.keys())
        # Use one representative prompt per concept for cross-concept comparison
        best_prompts = [VLM_CONCEPTS[c][0] for c in all_concepts]

        inputs = siglip_processor(
            text=best_prompts,
            images=[pil_image] * len(best_prompts),
            return_tensors="pt", padding="max_length"
        )
        inputs = {k: v.to(device) for k, v in inputs.items()}
        with torch.no_grad():
            out = siglip_model(**inputs)
            # logits_per_image: [num_images, num_texts]
            # Use softmax so scores sum to 1 and are meaningful as %
            logits = out.logits_per_image[0]  # shape [num_concepts]
            probs  = torch.softmax(logits, dim=0)

        results = {}
        for i, concept in enumerate(all_concepts):
            # Scale to 0-1 range; multiply by num_concepts so avg = 1.0
            raw = float(probs[i].item()) * len(all_concepts)
            results[concept] = round(min(1.0, raw), 3)

        print(f"[AdLens] SigLIP scores: {results}")
        return results
    except Exception as e:
        print(f"[AdLens] SigLIP error: {e}")
        import traceback; traceback.print_exc()
        return {}

def clip_score_concepts(pil_image):
    """CLIP: legacy semantic scoring (fallback)."""
    if not CLIP_AVAILABLE: return {}
    try:
        results = {}
        for concept, prompts in VLM_CONCEPTS.items():
            inputs = clip_processor(text=prompts, images=pil_image,
                                    return_tensors="pt", padding=True)
            # Move all tensors to the same device as the model
            inputs = {k: v.to(device) for k, v in inputs.items()}
            with torch.no_grad():
                out   = clip_model(**inputs)
                probs = out.logits_per_image.softmax(dim=1)
            results[concept] = round(float(probs[0].max().item()), 3)
        print(f"[AdLens] CLIP scores: {results}")
        return results
    except Exception as e:
        print(f"[AdLens] CLIP error: {e}")
        return {}

def get_vlm_scores(pil_image):
    """Returns (caption, concept_scores_dict) using best available VLM."""
    caption = None
    if FLORENCE_AVAILABLE:
        caption = florence_analyze(pil_image)
        scores  = siglip_score_concepts(pil_image) if SIGLIP_AVAILABLE else clip_score_concepts(pil_image)
    elif SIGLIP_AVAILABLE:
        scores  = siglip_score_concepts(pil_image)
    elif CLIP_AVAILABLE:
        scores  = clip_score_concepts(pil_image)
    else:
        scores  = {}
    # Fallback caption via BLIP
    if not caption and BLIP_AVAILABLE:
        try:
            inputs  = blip_processor_inst(images=pil_image, return_tensors="pt")
            inputs  = {k: v.to(device) for k, v in inputs.items()}
            out     = blip_model.generate(**inputs)
            caption = blip_processor_inst.decode(out[0], skip_special_tokens=True)
        except Exception as e:
            print(f"[AdLens] BLIP caption error: {e}")
            caption = "Ad image"
    return caption or "Ad image", scores

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 1: CTR estimation — extended with color psychology + sharpness
# ─────────────────────────────────────────────────────────────────────────────
def estimate_ctr(score, face_count, vlm_scores, has_cta, brightness,
                 sharpness=50, dominant_color="neutral", symmetry=5):
    base = 1.5
    base += score * 0.04
    base += face_count * 0.4              # faces → CTR boost
    base += vlm_scores.get("energy", 0) * 2.0
    base += vlm_scores.get("trust",  0) * 1.5
    base += vlm_scores.get("emotional", 0) * 1.0
    base -= vlm_scores.get("crowded", 0) * 2.0
    base += 0.8 if has_cta else 0
    # Brightness rules
    if brightness < 70:   base -= 0.7   # too dark
    elif brightness > 220: base -= 0.4  # overexposed
    else:                 base += 0.3   # good range
    # Sharpness — blurry ads perform worse
    if sharpness < 50:    base -= 0.6
    elif sharpness > 200: base += 0.4
    # Color psychology
    color_boost = COLOR_PSYCHOLOGY.get(dominant_color, {}).get("ctr_boost", 0.3)
    base += color_boost * 0.5
    # Symmetry — balanced layouts perform better
    if symmetry > 7: base += 0.3
    elif symmetry < 3: base -= 0.2
    return round(max(0.3, min(9.5, base)), 2)

# ─────────────────────────────────────────────────────────────────────────────
# FULL IMAGE ANALYSIS (all phases combined)
# ─────────────────────────────────────────────────────────────────────────────
def analyze_image(image_file):
    image = Image.open(image_file).convert("RGB")

    # Basic PIL stats
    stat       = ImageStat.Stat(image.convert("L"))
    brightness = stat.mean[0]
    contrast   = stat.stddev[0]

    color_score          = min(10, max(1, int((brightness / 255) * 10)))
    visual_clarity_score = min(10, max(1, int((contrast  / 128) * 10)))

    # PHASE 3: Advanced CV
    face_count    = detect_faces(image)
    text_area_pct = estimate_text_area_pct(image)
    sharpness     = compute_sharpness(image)
    symmetry      = compute_symmetry(image)
    dominant_color, color_palette = extract_dominant_colors(image, n_colors=4)

    # PHASE 4: Best VLM
    caption, vlm_scores = get_vlm_scores(image)

    # Derive scores from caption + VLM
    caption_lower = caption.lower()
    text_density_score = 5
    if any(w in caption_lower for w in ['text','sign','writing','words','poster','logo']):
        text_density_score = 8
    if text_area_pct > 40:
        text_density_score = min(10, text_density_score + 2)

    emotional_impact_score = 6
    if any(w in caption_lower for w in ['smile','happy','laugh','people','person','exciting','bright','vibrant']):
        emotional_impact_score = 9
    elif any(w in caption_lower for w in ['dark','empty','boring','bland']):
        emotional_impact_score = 3
    if face_count > 0:
        emotional_impact_score = min(10, emotional_impact_score + face_count)
    # Boost from VLM emotional score
    if vlm_scores.get("emotional", 0) > 0.5:
        emotional_impact_score = min(10, emotional_impact_score + 1)

    # Sharpness insight
    sharpness_score = min(10, max(1, int(min(sharpness, 500) / 50)))

    # OCR
    ocr_text = _run_ocr(image)

    return {
        "caption": caption,
        "color_score": color_score,
        "visual_clarity_score": visual_clarity_score,
        "text_density_score": text_density_score,
        "emotional_impact_score": emotional_impact_score,
        "ocr_text": ocr_text,
        "face_count": face_count,
        "text_area_pct": text_area_pct,
        "brightness": round(brightness, 1),
        "contrast": round(contrast, 1),
        "sharpness": sharpness,
        "sharpness_score": sharpness_score,
        "symmetry": symmetry,
        "dominant_color": dominant_color,
        "color_palette": color_palette,
        "vlm_scores": vlm_scores,
    }

def _run_ocr(pil_image):
    if OCR_BACKEND is None: return None
    try:
        if OCR_BACKEND == "easyocr":
            results = _ocr_reader.readtext(np.array(pil_image))
            return " ".join(r[1] for r in results).strip() or None
        else:
            import pytesseract
            return pytesseract.image_to_string(pil_image).strip() or None
    except Exception as e:
        print(f"[AdLens] OCR error: {e}")
        return None

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 8: RULE-BASED PSYCHOLOGY ENGINE
# ─────────────────────────────────────────────────────────────────────────────
def apply_psychology_rules(face_count, dominant_color, text_area_pct, brightness,
                           sharpness, symmetry, platform, audience, insights, suggestions):
    """Universal advertising psychology — not data-dependent."""
    score_delta = 0

    # Faces
    if face_count >= 1:
        insights.append(f"(+) 🧠 Psychology: {face_count} face(s) increase trust & recall by ~38% (Nielsen).")
        score_delta += face_count * 2
    else:
        suggestions.append("🧠 Add a human face — faces activate mirror neurons and boost engagement.")

    # Color psychology
    cp = COLOR_PSYCHOLOGY.get(dominant_color, COLOR_PSYCHOLOGY["neutral"])
    insights.append(f"{cp['emoji']} Color Psychology: Dominant color is <b>{dominant_color}</b> — signals {cp['meaning']}.")
    score_delta += cp["ctr_boost"] * 3

    # Dark ads
    if dominant_color == "black" and platform in [1, 2]:  # Instagram/Facebook
        suggestions.append("Dark ads score lower CTR on Instagram/Facebook — try lighter, vibrant alternatives.")
        score_delta -= 4

    # Blue = trust
    if dominant_color == "blue":
        insights.append("(+) Blue dominance builds trust — effective for finance, healthcare, SaaS brands.")

    # Red/Orange = urgency
    if dominant_color in ["red", "orange"]:
        insights.append(f"(+) {dominant_color.capitalize()} increases purchase urgency — great for CTA-heavy ads.")
        score_delta += 2

    # Text overload
    if text_area_pct > 40:
        suggestions.append(f"❌ Text covers {text_area_pct}% of image. Ad platforms restrict >20%. Reduce text.")
        score_delta -= 6
    elif text_area_pct > 25:
        suggestions.append(f"⚠️ Text area at {text_area_pct}% — approaching platform limits. Keep under 20%.")
        score_delta -= 2

    # Brightness rules
    if brightness < 70:
        suggestions.append("🌑 Image too dark. Dark ads show 20% lower CTR in bright feed environments.")
        score_delta -= 4
    elif brightness > 220:
        suggestions.append("💥 Overexposed image. Very bright visuals lose contrast and readability.")
        score_delta -= 2
    else:
        insights.append(f"(+) Brightness {int(brightness)} — optimal feed visibility range.")
        score_delta += 1

    # Sharpness
    if sharpness < 40:
        suggestions.append("🔍 Image appears blurry (sharpness score low). Blurry ads are skipped faster.")
        score_delta -= 4
    elif sharpness > 150:
        insights.append("(+) 🔍 High image sharpness — clear, crisp visuals improve scroll-stop rate.")
        score_delta += 2

    # Symmetry / balance
    if symmetry > 7:
        insights.append(f"(+) ⚖️ Layout symmetry {symmetry}/10 — balanced compositions feel professional.")
        score_delta += 2
    elif symmetry < 4:
        suggestions.append("⚖️ Unbalanced layout detected. Centered, symmetric ads feel more premium.")
        score_delta -= 1

    return score_delta

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 5: ENSEMBLE SCORING
# ─────────────────────────────────────────────────────────────────────────────
def compute_ensemble(ml_score, vlm_score, copy_score):
    """0.45 XGBoost + 0.35 Vision + 0.20 Copy — more balanced than 0.6/0.4."""
    weights = {"ML": 0.45, "Vision": 0.35, "Copy": 0.20}
    vs = vlm_score if vlm_score is not None else ml_score
    cs = copy_score if copy_score is not None else 50
    score = weights["ML"] * ml_score + weights["Vision"] * vs + weights["Copy"] * cs
    breakdown = {
        "ML Score (45%)": ml_score,
        "Vision Score (35%)": vs,
        "Copy Score (20%)": cs,
        "Ensemble": round(score, 1),
    }
    return int(np.clip(score, 0, 100)), breakdown

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 7: A/B VARIANT GENERATOR
# ─────────────────────────────────────────────────────────────────────────────
def generate_ab_variants(ad_copy, score, dominant_color, face_count, platform, audience,
                          brightness, sharpness, has_cta):
    """Generate AI-powered optimization suggestions for A/B testing."""
    variants = []

    # Variant A: Brighter version
    if brightness < 150:
        variants.append({
            "name": "🌟 Brighter Variant",
            "change": "Increase image brightness by +40%. Lighter backgrounds perform 22% better in feeds.",
            "predicted_ctr_boost": "+0.8%",
            "confidence": "High",
        })

    # Variant B: Face addition
    if face_count == 0:
        variants.append({
            "name": "👤 Human Element Variant",
            "change": "Add a smiling human face or lifestyle shot. Faces boost CTR by up to 35%.",
            "predicted_ctr_boost": "+1.2%",
            "confidence": "Very High",
        })

    # Variant C: Color change
    if dominant_color in ["black", "neutral", "white"]:
        variants.append({
            "name": f"🔥 Color Energy Variant",
            "change": "Introduce orange or red accents for CTA button/headline. Warm colors drive urgency.",
            "predicted_ctr_boost": "+0.6%",
            "confidence": "Medium",
        })

    # Variant D: CTA copy
    if not has_cta:
        variants.append({
            "name": "🎯 CTA Copy Variant",
            "change": f"Add a direct CTA like 'Shop Now →' or 'Get Started Free'. CTAs double conversion rates.",
            "predicted_ctr_boost": "+1.5%",
            "confidence": "Very High",
        })

    # Variant E: Sharpness
    if sharpness < 80:
        variants.append({
            "name": "🔍 Clarity Variant",
            "change": "Use a sharper, higher-resolution image. Blurry ads see 31% higher skip rates.",
            "predicted_ctr_boost": "+0.5%",
            "confidence": "Medium",
        })

    # Always include a text-reduction variant if heavy text
    variants.append({
        "name": "✂️ Minimal Text Variant",
        "change": "Cut image text by 50%. Keep only 1 headline + 1 CTA. Less text = more visual impact.",
        "predicted_ctr_boost": "+0.7%",
        "confidence": "High",
    })

    return variants[:4]  # Return top 4

# ─────────────────────────────────────────────────────────────────────────────
# AI REWRITE (Claude)
# ─────────────────────────────────────────────────────────────────────────────
PLATFORM_LABELS = {1:"Facebook", 2:"Instagram", 3:"YouTube", 4:"Google Display",
                   5:"TikTok", 6:"LinkedIn", 7:"Twitter/X", 8:"Amazon"}
AUDIENCE_LABELS = {1:"Gen Z / Students", 2:"Professionals (B2B)", 3:"General / Broad",
                   4:"Millennials", 5:"Parents", 6:"Seniors", 7:"Retargeting"}

def ai_rewrite_copy(original_copy, platform_id, audience_id, issues):
    if not ANTHROPIC_API_KEY or not original_copy: return None
    platform    = PLATFORM_LABELS.get(platform_id, "social media")
    audience    = AUDIENCE_LABELS.get(audience_id, "general audience")
    issues_text = "; ".join(issues[:4]) if issues else "improve overall persuasion"
    prompt = (
        f"You are an expert advertising copywriter. "
        f"Rewrite this ad copy for {platform}, targeting {audience}.\n\n"
        f"Original: \"{original_copy}\"\nIssues: {issues_text}\n\n"
        f"Rules: Max 12 words. Include one clear CTA. No hype words. "
        f"Return ONLY the rewritten copy."
    )
    payload = json.dumps({
        "model": CLAUDE_MODEL, "max_tokens": 120,
        "messages": [{"role": "user", "content": prompt}]
    }).encode()
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages", data=payload,
        headers={"Content-Type": "application/json",
                 "x-api-key": ANTHROPIC_API_KEY,
                 "anthropic-version": "2023-06-01"},
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=12) as resp:
            data = json.loads(resp.read())
            return data["content"][0]["text"].strip()
    except Exception as e:
        print(f"[AdLens] AI rewrite error: {e}")
        return None

# ─────────────────────────────────────────────────────────────────────────────
# COMPETITOR SCORING
# ─────────────────────────────────────────────────────────────────────────────
def score_image_for_compare(image_file):
    vlm = analyze_image(image_file)
    vlm_score = int(np.mean([vlm["color_score"], vlm["visual_clarity_score"],
                              vlm["text_density_score"], vlm["emotional_impact_score"]]) * 10)
    return {
        "vlm_score": vlm_score,
        "color_score": vlm["color_score"],
        "clarity_score": vlm["visual_clarity_score"],
        "text_density_score": vlm["text_density_score"],
        "emotional_score": vlm["emotional_impact_score"],
        "sharpness": vlm["sharpness"],
        "dominant_color": vlm["dominant_color"],
        "caption": vlm["caption"],
        "ocr_text": vlm.get("ocr_text"),
        "face_count": vlm.get("face_count", 0),
    }

# ─────────────────────────────────────────────────────────────────────────────
# HTML TEMPLATE
# ─────────────────────────────────────────────────────────────────────────────
HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>AdLens — AI Ad Intelligence</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        :root {
            --bg-primary: #0a0a0c; --bg-secondary: #131417;
            --accent: #ff6600; --accent-glow: rgba(255,102,0,0.3);
            --success: #4ade80; --warning: #facc15; --danger: #f87171;
            --text: #ffffff; --muted: #a1a1aa;
            --border: rgba(255,255,255,0.08);
        }
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            background: var(--bg-primary);
            background-image: linear-gradient(rgba(255,255,255,0.03) 1px,transparent 1px),
                              linear-gradient(90deg,rgba(255,255,255,0.03) 1px,transparent 1px);
            background-size: 50px 50px;
            color: var(--text); font-family: 'Outfit', sans-serif;
            min-height: 100vh; padding: 0 2rem 4rem; overflow-x: hidden;
        }
        .container { max-width: 920px; margin: 0 auto; }
        nav { display: flex; justify-content: space-between; align-items: center; padding: 2rem 0; margin-bottom: 3rem; }
        .logo { font-size: 1.6rem; font-weight: 800; display: flex; align-items: center; gap: 0.5rem; }
        .logo-icon { background: var(--accent); color: #000; padding: 0.2rem; border-radius: 8px; width: 32px; height: 32px; display: flex; align-items: center; justify-content: center; }
        .badge { padding: 0.4rem 1.2rem; border: 1px solid rgba(255,255,255,0.15); background: rgba(255,255,255,0.03); border-radius: 20px; font-size: 0.75rem; color: var(--muted); font-weight: 600; letter-spacing: 0.05em; }
        .vlm-badge { padding: 0.3rem 0.8rem; border: 1px solid rgba(255,102,0,0.4); background: rgba(255,102,0,0.08); border-radius: 20px; font-size: 0.72rem; color: var(--accent); font-weight: 700; }
        header { text-align: center; margin-bottom: 3.5rem; }
        header h1 { font-size: 4.5rem; font-weight: 900; line-height: 1.05; margin-bottom: 1.5rem; letter-spacing: -0.03em; }
        header h1 .highlight { color: var(--accent); display: block; }
        header p { color: var(--muted); font-size: 1.1rem; max-width: 620px; margin: 0 auto; line-height: 1.6; }
        @media (max-width: 768px) { header h1 { font-size: 3rem; } }
        /* Form */
        .type-selector { display: flex; justify-content: center; gap: 0.8rem; margin-bottom: 2.5rem; flex-wrap: wrap; }
        .radio-pill input { display: none; }
        .radio-pill .pill-content { background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.1); padding: 0.7rem 1.2rem; border-radius: 12px; color: var(--muted); cursor: pointer; display: flex; align-items: center; gap: 0.5rem; transition: all 0.2s; font-weight: 500; font-size: 0.95rem; }
        .radio-pill:hover .pill-content { background: rgba(255,255,255,0.08); color: white; }
        .radio-pill input:checked + .pill-content { background: var(--accent); color: #000; border-color: var(--accent); font-weight: 700; }
        .upload-area { background: #111113; border: 1px dashed rgba(255,255,255,0.2); border-radius: 20px; padding: 4rem 2rem; text-align: center; cursor: pointer; transition: all 0.3s; margin-bottom: 1.5rem; position: relative; overflow: hidden; display: flex; flex-direction: column; align-items: center; }
        .upload-area:hover { border-color: rgba(255,255,255,0.4); background: #161619; }
        .upload-icon { background: #3b82f6; color: white; width: 54px; height: 54px; border-radius: 14px; display: flex; align-items: center; justify-content: center; font-size: 1.8rem; margin-bottom: 1.5rem; box-shadow: 0 4px 20px rgba(59,130,246,0.4); }
        .upload-title { font-size: 1.5rem; font-weight: 700; margin-bottom: 0.5rem; }
        .upload-subtitle { color: var(--muted); font-size: 0.9rem; }
        #imageInput { position: absolute; top: 0; left: 0; width: 100%; height: 100%; opacity: 0; cursor: pointer; z-index: 10; }
        #imagePreview { display: none; max-width: 100%; max-height: 380px; border-radius: 12px; object-fit: contain; }
        .upload-content-wrapper { display: flex; flex-direction: column; align-items: center; z-index: 5; pointer-events: none; }
        .copy-group { margin-bottom: 1.5rem; }
        .copy-group label { display: block; margin-bottom: 0.6rem; color: var(--muted); font-size: 0.85rem; font-weight: 500; text-transform: uppercase; letter-spacing: 0.05em; }
        .copy-group textarea { width: 100%; padding: 0.9rem 1rem; background: rgba(0,0,0,0.4); border: 1px solid rgba(255,255,255,0.1); border-radius: 10px; color: white; font-family: 'Outfit',sans-serif; font-size: 0.95rem; resize: vertical; min-height: 80px; }
        .copy-group textarea:focus { outline: none; border-color: var(--accent); }
        .settings-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px,1fr)); gap: 1.5rem; margin-bottom: 2rem; padding: 2rem; border: 1px dashed rgba(255,255,255,0.1); border-radius: 16px; background: rgba(0,0,0,0.2); }
        .form-group label { display: block; margin-bottom: 0.6rem; color: var(--muted); font-size: 0.85rem; font-weight: 500; text-transform: uppercase; letter-spacing: 0.05em; }
        input, select { width: 100%; padding: 0.9rem 1rem; background: rgba(0,0,0,0.4); border: 1px solid rgba(255,255,255,0.1); border-radius: 10px; color: white; font-family: 'Outfit',sans-serif; font-size: 0.95rem; }
        input:focus, select:focus { outline: none; border-color: var(--accent); }
        select option { background: #1a1a1a; }
        .advanced-settings-toggle { padding: 1rem; border: 1px solid rgba(255,255,255,0.08); border-radius: 12px; cursor: pointer; margin-bottom: 1.5rem; display: flex; justify-content: space-between; color: var(--muted); font-weight: 600; }
        .btn-submit { width: 100%; padding: 1.1rem; background: var(--accent); color: #000; font-weight: 800; font-size: 1.1rem; border: none; border-radius: 14px; cursor: pointer; transition: all 0.2s; display: flex; align-items: center; justify-content: center; gap: 0.5rem; font-family: 'Outfit',sans-serif; }
        .btn-submit:hover { background: #e55500; transform: translateY(-1px); }
        /* Results */
        .results-wrapper { margin-top: 3rem; }
        .results-panel { background: rgba(255,255,255,0.02); border: 1px solid var(--border); border-radius: 24px; padding: 2.5rem; }
        .score-container { text-align: center; margin-bottom: 3rem; }
        .score-circle { width: 160px; height: 160px; border-radius: 50%; border: 6px solid; margin: 0 auto 1.5rem; display: flex; align-items: center; justify-content: center; font-size: 3.5rem; font-weight: 900; position: relative; border-color: var(--score-color, var(--accent)); box-shadow: 0 0 40px var(--score-glow, var(--accent-glow)); background: conic-gradient(var(--score-color, var(--accent)) var(--score-degrees,0deg), rgba(255,255,255,0.05) 0deg); }
        .score-inner { background: var(--bg-primary); border-radius: 50%; width: 136px; height: 136px; display: flex; align-items: center; justify-content: center; position: absolute; }
        .feedback-badge { display: inline-block; padding: 0.5rem 1.5rem; border: 1px solid; border-radius: 20px; font-weight: 700; font-size: 1rem; margin-bottom: 0.8rem; }
        .ctr-badge { display: inline-block; padding: 0.4rem 1.2rem; background: rgba(74,222,128,0.1); border: 1px solid rgba(74,222,128,0.3); border-radius: 20px; font-size: 0.9rem; font-weight: 600; color: var(--success); margin-bottom: 0.5rem; }
        .percentile-badge { display: inline-block; padding: 0.4rem 1.2rem; background: rgba(255,102,0,0.1); border: 1px solid rgba(255,102,0,0.3); border-radius: 20px; font-size: 0.9rem; font-weight: 600; color: var(--accent); }
        .metrics-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px,1fr)); gap: 1rem; margin-bottom: 2rem; }
        .metric-card { background: rgba(255,255,255,0.03); border: 1px solid var(--border); border-radius: 14px; padding: 1.2rem; text-align: center; }
        .metric-value { font-size: 1.8rem; font-weight: 800; color: var(--accent); }
        .metric-label { font-size: 0.78rem; color: var(--muted); text-transform: uppercase; letter-spacing: 0.05em; margin-top: 0.3rem; }
        /* Color palette */
        .color-palette { display: flex; gap: 0.5rem; align-items: center; flex-wrap: wrap; margin-top: 0.5rem; }
        .color-chip { width: 24px; height: 24px; border-radius: 6px; border: 1px solid rgba(255,255,255,0.2); }
        /* A/B variants */
        .ab-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(220px,1fr)); gap: 1rem; margin-bottom: 2rem; }
        .ab-card { background: rgba(255,102,0,0.05); border: 1px solid rgba(255,102,0,0.2); border-radius: 14px; padding: 1.2rem; }
        .ab-name { font-weight: 700; font-size: 0.95rem; margin-bottom: 0.5rem; color: var(--accent); }
        .ab-change { font-size: 0.82rem; color: var(--muted); line-height: 1.5; margin-bottom: 0.5rem; }
        .ab-boost { font-size: 0.82rem; font-weight: 700; color: var(--success); }
        /* Ensemble breakdown */
        .ensemble-bar { margin-bottom: 1rem; }
        .ensemble-bar-label { display: flex; justify-content: space-between; font-size: 0.82rem; color: var(--muted); margin-bottom: 0.3rem; }
        .ensemble-bar-track { background: rgba(255,255,255,0.06); border-radius: 6px; height: 8px; overflow: hidden; }
        .ensemble-bar-fill { height: 100%; border-radius: 6px; background: var(--accent); transition: width 0.8s ease; }
        /* Insight/suggestion lists */
        .insight-item { padding: 0.75rem 1rem; background: rgba(74,222,128,0.05); border: 1px solid rgba(74,222,128,0.15); border-radius: 10px; margin-bottom: 0.5rem; font-size: 0.88rem; line-height: 1.5; }
        .suggestion-item { padding: 0.75rem 1rem; background: rgba(250,204,21,0.05); border: 1px solid rgba(250,204,21,0.15); border-radius: 10px; margin-bottom: 0.5rem; font-size: 0.88rem; line-height: 1.5; }
        .section-title { font-size: 1.1rem; font-weight: 700; margin-bottom: 1rem; display: flex; align-items: center; gap: 0.5rem; }
        .section-block { margin-bottom: 2rem; }
        .rewrite-box { background: rgba(255,102,0,0.06); border: 1px solid rgba(255,102,0,0.25); border-radius: 14px; padding: 1.2rem 1.5rem; font-size: 1.05rem; font-weight: 600; letter-spacing: 0.01em; margin-top: 0.5rem; }
        .competitor-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 1.5rem; }
        @media (max-width: 600px) { .competitor-grid { grid-template-columns: 1fr; } }
        .competitor-card { background: rgba(255,255,255,0.03); border: 1px solid var(--border); border-radius: 14px; padding: 1.2rem; }
        .thumb-buttons { display: flex; gap: 1rem; justify-content: center; margin-top: 1.5rem; }
        .thumb-btn { padding: 0.7rem 1.8rem; border-radius: 10px; border: 1px solid; font-weight: 700; font-size: 1rem; cursor: pointer; font-family: 'Outfit',sans-serif; transition: all 0.2s; }
        .thumb-up { background: rgba(74,222,128,0.1); border-color: rgba(74,222,128,0.4); color: var(--success); }
        .thumb-up:hover { background: rgba(74,222,128,0.25); }
        .thumb-down { background: rgba(248,113,113,0.1); border-color: rgba(248,113,113,0.4); color: var(--danger); }
        .thumb-down:hover { background: rgba(248,113,113,0.25); }
        .feedback-msg { text-align: center; margin-top: 0.8rem; font-size: 0.85rem; color: var(--muted); min-height: 1.2em; }
        a { color: var(--accent); text-decoration: none; }
        a:hover { text-decoration: underline; }
        @keyframes fadeIn { from { opacity:0; transform:translateY(16px); } to { opacity:1; transform:translateY(0); } }
        .fadeIn { animation: fadeIn 0.5s ease-out forwards; }
    </style>
</head>
<body>
<div class="container">
    <nav>
        <div class="logo">
            <div class="logo-icon">
            </div>
            AdLens
        </div>
        <div style="display:flex;gap:0.6rem;align-items:center;">
            <a href="/history" class="badge" style="cursor:pointer;color:var(--muted);">History</a>
        </div>
    </nav>

    <header>
        <h1>AI-Powered<span class="highlight">Ad Intelligence</span></h1>
        <p>Analyze your ads and get smart performance suggestions</p>
    </header>

    <div class="form-panel">
        <!-- Ad type pills -->
        <div class="type-selector">
            {% for val, label, icon in [(1,'Image Ad','🖼️'),(3,'Carousel','🎠'),(4,'Story Ad','📱'),(5,'Banner','📢')] %}
            <label class="radio-pill">
                <input type="radio" name="ad_type_pick" value="{{ val }}"
                    {% if request.form.get('ad_type')|int == val %}checked{% elif val==1 and not request.form.get('ad_type') %}checked{% endif %}
                    onchange="document.getElementById('ad_type_hidden').value=this.value">
                <span class="pill-content">{{ icon }} {{ label }}</span>
            </label>
            {% endfor %}
        </div>

        <form method="POST" enctype="multipart/form-data">
            <input type="hidden" name="ad_type" id="ad_type_hidden" value="{{ request.form.get('ad_type','1') }}">

            <!-- Image upload -->
            <div class="upload-area" id="dropZone">
                <input type="file" name="ad_image" id="imageInput" accept="image/png,image/jpeg,image/webp" onchange="previewImage(event)">
                <div class="upload-content-wrapper" id="uploadPlaceholder">
                    <div class="upload-icon">🖼️</div>
                    <div class="upload-title">Drop your ad here</div>
                    <div class="upload-subtitle">Upload your ad image for AI analysiss</div>
                </div>
                <img id="imagePreview" src="" alt="Preview">
            </div>

            <!-- Ad copy -->
            <div class="copy-group">
                <label>Ad Copy / Headline</label>
                <textarea name="ad_copy" placeholder="e.g. Shop now — 50% off today only! Limited stock.">{{ request.form.get('ad_copy','') }}</textarea>
            </div>

            <!-- Advanced settings -->
            <div class="advanced-settings-toggle" onclick="toggleSettings()">
                <span>⚙️ Advanced Settings</span> <span id="toggleIcon">▼</span>
            </div>
            <div class="settings-grid" id="advancedSettings" style="display:none;">
                <div class="form-group">
                    <label>Platform</label>
                    <select name="platform">
                        {% for val, lbl in [(1,'Facebook'),(2,'Instagram'),(3,'YouTube'),(4,'Google Display'),(5,'TikTok'),(6,'LinkedIn'),(7,'Twitter/X'),(8,'Amazon')] %}
                        <option value="{{ val }}" {% if request.form.get('platform')|int == val %}selected{% elif val==1 and not request.form.get('platform') %}selected{% endif %}>{{ lbl }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div class="form-group">
                    <label>Target Audience</label>
                    <select name="audience">
                        {% for val, lbl in [(1,'Gen Z / Students'),(2,'Professionals (B2B)'),(3,'General / Broad'),(4,'Millennials'),(5,'Parents'),(7,'Retargeting')] %}
                        <option value="{{ val }}" {% if request.form.get('audience')|int == val %}selected{% elif val==3 and not request.form.get('audience') %}selected{% endif %}>{{ lbl }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div class="form-group">
                    <label>Color Style</label>
                    <select name="color">
                        <option value="3" {% if request.form.get('color')=='3' %}selected{% endif %}>Dark & Moody</option>
                        <option value="6" {% if not request.form.get('color') or request.form.get('color')=='6' %}selected{% endif %}>Balanced/Neutral</option>
                        <option value="9" {% if request.form.get('color')=='9' %}selected{% endif %}>Bright & Vibrant</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>Day of Week</label>
                    <select name="day">
                        {% for val, lbl in [(1,'Monday'),(2,'Tuesday'),(3,'Wednesday'),(4,'Thursday'),(5,'Friday'),(6,'Saturday'),(7,'Sunday')] %}
                        <option value="{{ val }}" {% if request.form.get('day')|int == val %}selected{% elif val==3 and not request.form.get('day') %}selected{% endif %}>{{ lbl }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div class="form-group">
                    <label>Posting Hour (0–23)</label>
                    <input type="number" name="time" min="0" max="23" value="{{ request.form.get('time','18') }}">
                </div>
            </div>

            <button type="submit" class="btn-submit"> Analyze My Ad </button>

            <details style="margin-top:1.5rem;">
                <summary style="cursor:pointer;color:var(--muted);font-size:0.9rem;font-weight:600;padding:0.5rem 0;">⚡ Compare vs competitor ad (optional)</summary>
                <div style="margin-top:1rem;padding:1rem;background:rgba(99,102,241,0.06);border:1px dashed rgba(99,102,241,0.3);border-radius:12px;">
                    <label style="display:block;margin-bottom:0.5rem;font-size:0.82rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--muted);font-weight:600;">Competitor Ad Image</label>
                    <input type="file" name="competitor_image" accept="image/png,image/jpeg,image/webp" style="font-size:0.85rem;color:var(--muted);">
                </div>
            </details>
        </form>
    </div><!-- /.form-panel -->

    {% if score is not none %}
    <div class="results-wrapper fadeIn" id="results">
        <h2 style="text-align:center;margin-bottom:2.5rem;font-size:2.2rem;font-weight:800;">Analysis Results</h2>
        <div class="results-panel">

            {% if uploaded_image_data_uri %}
            <div style="display:flex;justify-content:center;margin-bottom:1.5rem;">
                <img src="{{ uploaded_image_data_uri }}" alt="Ad" style="max-width:100%;max-height:460px;object-fit:contain;border-radius:20px;border:1px solid var(--border);box-shadow:0 10px 40px rgba(0,0,0,0.5);">
            </div>
            {% endif %}

            <!-- Score circle -->
            <div class="score-container" style="--score-degrees:{{ score * 3.6 }}deg;--score-color:{{ color_hex }};--score-glow:{{ color_glow }};">
                <h3 style="margin-bottom:1.5rem;font-size:1rem;color:var(--muted);text-transform:uppercase;letter-spacing:0.1em;">Ensemble Score</h3>
                <div class="score-circle">
                    <div class="score-inner"><span style="font-size:3.2rem;font-weight:900;">{{ score }}</span></div>
                </div>
                <div class="feedback-badge" style="color:{{ color_hex }};border-color:{{ color_glow }};">{{ feedback }}</div>
                <br>
                {% if ctr_estimate %}
                <span class="ctr-badge">📈 Estimated CTR: ~{{ ctr_estimate }}%</span>
                {% endif %}
                {% if percentile_rank is not none %}
                <br><span class="percentile-badge" style="margin-top:0.5rem;">🏆 Better than {{ percentile_rank }}% of ads in dataset</span>
                {% endif %}
            </div>

            <!-- Sub-scores -->
            <div class="metrics-grid" style="margin-bottom:2.5rem;">
                <div class="metric-card"><div class="metric-value">{{ ml_score }}</div><div class="metric-label">ML Score</div></div>
                {% if vlm_score is not none %}<div class="metric-card"><div class="metric-value">{{ vlm_score }}</div><div class="metric-label">Vision Score</div></div>{% endif %}
                {% if copy_result %}<div class="metric-card"><div class="metric-value">{{ copy_result.copy_score }}</div><div class="metric-label">Copy Score</div></div>{% endif %}
                {% if sharpness_score is not none %}<div class="metric-card"><div class="metric-value">{{ sharpness_score }}/10</div><div class="metric-label">Sharpness</div></div>{% endif %}
                {% if symmetry is not none %}<div class="metric-card"><div class="metric-value">{{ "%.1f"|format(symmetry) }}/10</div><div class="metric-label">Layout Balance</div></div>{% endif %}
                {% if face_count_val is not none %}<div class="metric-card"><div class="metric-value">{{ face_count_val }}</div><div class="metric-label">Faces Detected</div></div>{% endif %}
            </div>

            <!-- Dominant color & palette -->
            {% if dominant_color %}
            <div class="section-block">
                <div class="section-title">🎨 Color Psychology</div>
                <div style="background:rgba(255,255,255,0.03);border:1px solid var(--border);border-radius:14px;padding:1.2rem;">
                    <strong>Dominant: {{ dominant_color | capitalize }}</strong>
                    {% if color_psychology %} — {{ color_psychology.meaning }}{% endif %}
                    {% if color_palette %}
                    <div class="color-palette" style="margin-top:0.8rem;">
                        {% for c in color_palette %}
                        <span style="padding:0.25rem 0.7rem;background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.15);border-radius:20px;font-size:0.78rem;color:var(--muted);">{{ c }}</span>
                        {% endfor %}
                    </div>
                    {% endif %}
                </div>
            </div>
            {% endif %}

            <!-- Ensemble breakdown -->
            {% if ensemble_breakdown %}
            <div class="section-block">
                <div class="section-title">⚖️ Ensemble Score Breakdown</div>
                {% for label, val in ensemble_breakdown.items() %}
                {% if label != 'Ensemble' %}
                <div class="ensemble-bar">
                    <div class="ensemble-bar-label"><span>{{ label }}</span><span>{{ val }}</span></div>
                    <div class="ensemble-bar-track"><div class="ensemble-bar-fill" style="width:{{ val }}%"></div></div>
                </div>
                {% endif %}
                {% endfor %}
            </div>
            {% endif %}

            <!-- VLM concept scores -->
            {% if vlm_scores %}
            <div class="section-block">
                <div class="section-title">🧠 Semantic Scores</div>
                <div class="metrics-grid">
                    {% for concept, score_val in vlm_scores.items() %}
                    <div class="metric-card">
                        <div class="metric-value" style="font-size:1.4rem;">{{ "%.0f"|format(score_val * 100) }}%</div>
                        <div class="metric-label">{{ concept | capitalize }}</div>
                    </div>
                    {% endfor %}
                </div>
            </div>
            {% endif %}

            <!-- Insights -->
            {% if insights %}
            <div class="section-block">
                <div class="section-title">💡 Insights</div>
                {% for i in insights %}
                <div class="insight-item">{{ i | safe }}</div>
                {% endfor %}
            </div>
            {% endif %}

            <!-- Suggestions -->
            {% if suggestions %}
            <div class="section-block">
                <div class="section-title">🛠️ Optimization Suggestions</div>
                {% for s in suggestions %}
                <div class="suggestion-item">{{ s }}</div>
                {% endfor %}
            </div>
            {% endif %}

            <!-- A/B Variants (PHASE 7) -->
            {% if ab_variants %}
            <div class="section-block">
                <div class="section-title">🧪 A/B Variant Generator</div>
                <p style="color:var(--muted);font-size:0.85rem;margin-bottom:1rem;">Try these optimized variants — each predicts a CTR improvement.</p>
                <div class="ab-grid">
                    {% for v in ab_variants %}
                    <div class="ab-card">
                        <div class="ab-name">{{ v.name }}</div>
                        <div class="ab-change">{{ v.change }}</div>
                        <div class="ab-boost">{{ v.predicted_ctr_boost }} CTR · {{ v.confidence }} confidence</div>
                    </div>
                    {% endfor %}
                </div>
            </div>
            {% endif %}

            <!-- AI Copy Rewrite -->
            {% if rewrite %}
            <div class="section-block">
                <div class="section-title">✍️ AI-Optimized Copy (Claude)</div>
                <div class="rewrite-box">"{{ rewrite }}"</div>
            </div>
            {% endif %}

            <!-- Caption -->
            {% if caption %}
            <div class="section-block">
                <div class="section-title">🔍 {{ vlm_name }} Caption</div>
                <div style="background:rgba(255,255,255,0.03);border:1px solid var(--border);border-radius:12px;padding:1rem;font-size:0.9rem;color:var(--muted);line-height:1.6;">{{ caption }}</div>
            </div>
            {% endif %}

            <!-- Competitor comparison -->
            {% if competitor %}
            <div class="section-block">
                <div class="section-title">⚡ Competitor Comparison</div>
                <div class="competitor-grid">
                    <div class="competitor-card">
                        <div style="font-weight:700;margin-bottom:0.8rem;color:var(--accent);">Your Ad</div>
                        {% for k, v in your_metrics.items() %}
                        <div style="display:flex;justify-content:space-between;font-size:0.85rem;margin-bottom:0.4rem;">
                            <span style="color:var(--muted);">{{ k }}</span><span>{{ v }}</span>
                        </div>
                        {% endfor %}
                    </div>
                    <div class="competitor-card">
                        <div style="font-weight:700;margin-bottom:0.8rem;color:var(--muted);">Competitor</div>
                        {% for k, v in competitor_metrics.items() %}
                        <div style="display:flex;justify-content:space-between;font-size:0.85rem;margin-bottom:0.4rem;">
                            <span style="color:var(--muted);">{{ k }}</span><span>{{ v }}</span>
                        </div>
                        {% endfor %}
                        {% if competitor.dominant_color %}<div style="font-size:0.82rem;margin-top:0.5rem;color:var(--muted);">Color: {{ competitor.dominant_color }}</div>{% endif %}
                    </div>
                </div>
            </div>
            {% endif %}



            <!-- Feedback -->
            <div class="thumb-buttons">
                <button class="thumb-btn thumb-up" onclick="sendFeedback({{ analysis_id }}, 1)">👍 Accurate</button>
                <button class="thumb-btn thumb-down" onclick="sendFeedback({{ analysis_id }}, 0)">👎 Off</button>
            </div>
            <div class="feedback-msg" id="fbMsg">
                {% if feedback_stats.total > 0 %}
                Model accuracy from {{ feedback_stats.total }} feedback(s): {{ feedback_stats.accuracy_rate }}%
                {% endif %}
            </div>
        </div>
    </div>
    {% endif %}
</div>

<script>
function previewImage(event) {
    const r = new FileReader();
    r.onload = function(){
        const img = document.getElementById('imagePreview');
        img.src = r.result; img.style.display = 'block';
        document.getElementById('uploadPlaceholder').style.display = 'none';
        document.getElementById('dropZone').style.padding = '1rem';
    };
    if (event.target.files[0]) r.readAsDataURL(event.target.files[0]);
}
function toggleSettings() {
    const s = document.getElementById('advancedSettings');
    const i = document.getElementById('toggleIcon');
    if (s.style.display === 'none') { s.style.display = 'grid'; i.innerHTML = '▲'; }
    else { s.style.display = 'none'; i.innerHTML = '▼'; }
}
function sendFeedback(id, rating) {
    fetch('/feedback', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({analysis_id: id, rating: rating})
    }).then(r => r.json()).then(d => {
        const el = document.getElementById('fbMsg');
        el.textContent = rating === 1
            ? '✅ Thanks! Marked as accurate.'
            : '📝 Thanks! Noted as inaccurate — helps improve the model.';
        el.style.color = rating === 1 ? '#4ade80' : '#f87171';
        if (d.feedback_stats && d.feedback_stats.accuracy_rate !== null) {
            el.textContent += ` (${d.feedback_stats.accuracy_rate}% accuracy from ${d.feedback_stats.total} ratings)`;
        }
    });
}
// Sync ad type pill → hidden input
document.addEventListener('DOMContentLoaded', function(){
    document.querySelectorAll('input[name="ad_type_pick"]').forEach(r => {
        r.addEventListener('change', function(){
            document.getElementById('ad_type_hidden').value = this.value;
        });
    });
    {% if score is not none %}
    document.getElementById('results').scrollIntoView({behavior:'smooth'});
    {% endif %}
});
</script>
</body>
</html>
"""

HISTORY_HTML = """
<!DOCTYPE html><html lang="en">
<head><meta charset="UTF-8"><title>AdLens V2 — History</title>
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;600;700&display=swap" rel="stylesheet">
<style>
:root{--bg:#0a0a0c;--accent:#ff6600;--border:rgba(255,255,255,0.08);--muted:#a1a1aa;--text:#fff;}
*{box-sizing:border-box;margin:0;padding:0;}
body{background:var(--bg);color:var(--text);font-family:'Outfit',sans-serif;padding:2rem;}
.container{max-width:960px;margin:0 auto;}
h1{font-size:2rem;font-weight:800;margin-bottom:0.5rem;}
a{color:var(--accent);}
table{width:100%;border-collapse:collapse;margin-top:1.5rem;font-size:0.85rem;}
th{background:rgba(255,255,255,0.05);padding:0.8rem;text-align:left;color:var(--muted);font-weight:600;text-transform:uppercase;letter-spacing:0.05em;}
td{padding:0.75rem;border-bottom:1px solid var(--border);}
tr:hover td{background:rgba(255,255,255,0.02);}
.score-pill{display:inline-block;padding:0.2rem 0.7rem;border-radius:20px;font-weight:700;font-size:0.82rem;}
</style></head>
<body><div class="container">
<h1>📊 AdLens V2 — Analysis History</h1>
<p style="color:var(--muted);margin-bottom:0.5rem;">
{% if feedback_stats.total > 0 %}
Model accuracy: <strong>{{ feedback_stats.accuracy_rate }}%</strong> from {{ feedback_stats.total }} user ratings.
{% else %}No feedback yet.{% endif %}
</p>
<a href="/">← Back to Analyzer</a>
<table>
<thead><tr><th>Date</th><th>Score</th><th>ML</th><th>Vision</th><th>Copy</th><th>CTR%</th><th>Platform</th><th>Color</th><th>Rating</th></tr></thead>
<tbody>
{% for r in rows %}
<tr>
<td>{{ r.ts }}</td>
<td><span class="score-pill" style="background:{% if r.score>=80 %}rgba(74,222,128,0.15);color:#4ade80{% elif r.score>=60 %}rgba(250,204,21,0.15);color:#facc15{% else %}rgba(248,113,113,0.15);color:#f87171{% endif %}">{{ r.score }}</span></td>
<td>{{ r.ml_score or '—' }}</td>
<td>{{ r.vlm_score or '—' }}</td>
<td>{{ r.copy_score or '—' }}</td>
<td>{{ r.ctr_estimate or '—' }}</td>
<td>{{ r.platform or '—' }}</td>
<td>{{ r.dominant_color or '—' }}</td>
<td>{% if r.user_rating == 1 %}👍{% elif r.user_rating == 0 %}👎{% else %}—{% endif %}</td>
</tr>
{% endfor %}
</tbody></table>
</div></body></html>
"""

# ─────────────────────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/history")
def history():
    return render_template_string(HISTORY_HTML, rows=load_history(60),
                                  feedback_stats=get_feedback_stats())

@app.route("/feedback", methods=["POST"])
def feedback_route():
    data = request.get_json()
    aid, rating = data.get("analysis_id"), data.get("rating")
    if aid is None or rating is None:
        return jsonify({"error": "missing fields"}), 400
    save_feedback(aid, rating, data.get("comment",""))
    return jsonify({"ok": True, "feedback_stats": get_feedback_stats()})

@app.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        ad_type  = int(request.form.get("ad_type", 1))
        color    = int(request.form.get("color", 6))
        audience = int(request.form.get("audience", 3))
        platform = int(request.form.get("platform", 1))
        day      = int(request.form.get("day", 3))
        time_val = int(request.form.get("time", 18))
        ad_copy  = request.form.get("ad_copy", "").strip()

        suggestions, insights = [], []
        vlm_score = None
        vlm_results = None
        ctr_estimate = None
        face_count_val = 0
        brightness_val = 128.0
        sharpness_val = 50.0
        symmetry_val = 5.0
        dominant_color = "neutral"
        color_palette = []
        vlm_scores_dict = {}
        caption_text = None
        uploaded_image_data_uri = None
        sharpness_score = None
        psychology_delta = 0

        headline_length = len(ad_copy.split()) if ad_copy else 8
        copy_result = analyze_ad_copy(ad_copy) if ad_copy else None

        # Copy insights
        if copy_result:
            if not copy_result['has_cta']:
                suggestions.append("Add a clear CTA (e.g. 'Shop Now', 'Get Started') — CTAs double conversion rates.")
            if not copy_result['has_urgency']:
                suggestions.append("Add urgency ('today only', 'limited offer') to accelerate purchase decisions.")
            if copy_result['sentiment_label'] == 'Negative':
                suggestions.append("Negative copy language detected. Reframe around positive outcomes and benefits.")
            if copy_result['word_count'] > 25:
                suggestions.append("Headline too long — cut to under 10 words for mobile readability.")
            if copy_result['word_count'] < 3:
                suggestions.append("Copy too short — add a clear value proposition.")
            if copy_result['readability'] < 5:
                suggestions.append("Simplify language — use shorter, common words for broader reach.")
            if copy_result['copy_score'] >= 70:
                insights.append(f"(+) Copy score {copy_result['copy_score']}/100 — strong persuasive signals detected.")
            elif copy_result['copy_score'] >= 45:
                insights.append(f"👁️ Copy score {copy_result['copy_score']}/100 — decent but missing key elements.")
            else:
                insights.append(f"(-) Copy score {copy_result['copy_score']}/100 — weak. Add CTA + positive framing.")

        # Image analysis
        if 'ad_image' in request.files and request.files['ad_image'].filename != '':
            try:
                ad_image_file = request.files['ad_image']
                image_bytes   = ad_image_file.read()
                ad_image_file.seek(0)
                uploaded_image_data_uri = f"data:{ad_image_file.mimetype or 'image/png'};base64,{base64.b64encode(image_bytes).decode()}"

                vlm_results = analyze_image(ad_image_file)

                # Pull all CV features
                face_count_val = vlm_results.get("face_count", 0)
                brightness_val = vlm_results.get("brightness", 128)
                sharpness_val  = vlm_results.get("sharpness", 50)
                sharpness_score= vlm_results.get("sharpness_score")
                symmetry_val   = vlm_results.get("symmetry", 5)
                dominant_color = vlm_results.get("dominant_color", "neutral")
                color_palette  = vlm_results.get("color_palette", [])
                vlm_scores_dict= vlm_results.get("vlm_scores", {})
                caption_text   = vlm_results.get("caption")

                # Vision score from VLM + CV
                vlm_score = int(np.mean([
                    vlm_results["color_score"],
                    vlm_results["visual_clarity_score"],
                    vlm_results["text_density_score"],
                    vlm_results["emotional_impact_score"],
                ]) * 10)

                # Caption insight — sanitize to prevent XSS via VLM output
                if caption_text:
                    caption_safe = caption_text[:120].capitalize().replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    insights.append(f"🔍 {VLM_NAME}: '{caption_safe}'")

                # PHASE 8: Psychology engine
                psychology_delta = apply_psychology_rules(
                    face_count_val, dominant_color, vlm_results['text_area_pct'],
                    brightness_val, sharpness_val, symmetry_val,
                    platform, audience, insights, suggestions
                )

                # OCR
                if vlm_results.get("ocr_text"):
                    wc = len(vlm_results["ocr_text"].split())
                    if wc > 20:
                        suggestions.append(f"OCR found {wc} words in image — text overload risks platform restriction.")
                        insights.append(f"(-) OCR: {wc} words detected — exceeds recommended 20-word limit.")
                    elif wc > 0:
                        insights.append(f"(+) OCR: {wc} words in image — within safe range.")

                # VLM concept insights
                if vlm_scores_dict:
                    if vlm_scores_dict.get("luxury", 0) > 0.5:
                        insights.append(f"(+) {VLM_NAME}: strong luxury/premium visual cues ({int(vlm_scores_dict['luxury']*100)}%).")
                    if vlm_scores_dict.get("trust", 0) > 0.5:
                        insights.append(f"(+) {VLM_NAME}: trustworthy brand signals ({int(vlm_scores_dict['trust']*100)}%).")
                    if vlm_scores_dict.get("crowded", 0) > 0.5:
                        suggestions.append(f"{VLM_NAME} flags cluttered layout ({int(vlm_scores_dict['crowded']*100)}%). Simplify.")
                        insights.append(f"(-) {VLM_NAME}: layout appears crowded — reduces recall score.")
                    if vlm_scores_dict.get("energy", 0) > 0.5:
                        insights.append(f"(+) {VLM_NAME}: high energy/excitement ({int(vlm_scores_dict['energy']*100)}%) — boosts CTR.")
                    if vlm_scores_dict.get("modern", 0) > 0.5:
                        insights.append(f"(+) {VLM_NAME}: modern, contemporary aesthetic detected.")
                    if vlm_scores_dict.get("emotional", 0) > 0.5:
                        insights.append(f"(+) {VLM_NAME}: emotional storytelling detected — improves brand recall.")

                # CTR estimate (extended)
                ctr_estimate = estimate_ctr(
                    score=vlm_score, face_count=face_count_val,
                    vlm_scores=vlm_scores_dict,
                    has_cta=copy_result['has_cta'] if copy_result else False,
                    brightness=brightness_val, sharpness=sharpness_val,
                    dominant_color=dominant_color, symmetry=symmetry_val
                )

            except Exception as e:
                print(f"[AdLens] Image analysis error: {e}")
                insights.append("Failed to process image.")

        # Engineered features
        text_d = vlm_results["text_density_score"]    if vlm_results else 5
        emo    = vlm_results["emotional_impact_score"] if vlm_results else 6
        apm, phf, vsb = get_engineered_features(ad_type, audience, platform, time_val, text_d, emo)

        # ML prediction (extended features)
        ctr_raw = (ctr_estimate / 100) if ctr_estimate else 0.03
        engagement_est = min(10, max(1, (vlm_score or 50) / 10)) if vlm_score else 3
        features = np.array([[
            headline_length, ad_type, color, audience, platform,
            time_val, day, apm, phf, vsb,
            ctr_raw, brightness_val, emo,
            text_d,
            vlm_results.get("contrast", 30) if vlm_results else 30,
            5000,  # default budget
            engagement_est,
            1,     # default sentiment (positive)
            30,    # default campaign days
        ]])
        features_scaled = scaler.transform(features)
        ml_score = int(np.clip(model.predict(features_scaled)[0], 5, 100))

        # SHAP
        shap_insights = get_shap_insights(features_scaled, features[0])
        insights.extend(shap_insights)

        # PHASE 5: Ensemble
        copy_s = copy_result['copy_score'] if copy_result else None
        ensemble_score, ensemble_breakdown = compute_ensemble(ml_score, vlm_score, copy_s)

        # Apply psychology delta
        final_score = int(np.clip(ensemble_score + psychology_delta, 0, 100))

        # Platform/audience rules
        if apm == 1:
            insights.append("(+) Audience–platform pairing matches high-synergy combos.")
        else:
            suggestions.append("Reconsider platform — your audience engages better elsewhere.")
            insights.append("(-) Audience–platform mismatch: lower engagement predicted.")
        if phf == 1:
            insights.append("(+) Posting during peak hours (18:00–21:00) — highest CTR window.")
        else:
            suggestions.append(f"Reschedule to 18:00–21:00 for better reach. {time_val}:00 is off-peak.")
        if platform in [1, 5] and audience == 2:  # Facebook/TikTok + B2B
            suggestions.append("B2B audiences are weak on Facebook/TikTok. Try LinkedIn instead.")
            final_score = int(np.clip(final_score - 8, 0, 100))
        if audience == 2 and day in [6, 7]:
            suggestions.append("B2B disengages on weekends. Move to Tue–Thu for best reach.")
            final_score = int(np.clip(final_score - 8, 0, 100))
        elif audience in [1, 3] and day in [6, 7]:
            insights.append("(+) Weekends boost engagement for Gen Z/General audiences (+10%).")
            final_score = int(np.clip(final_score + 7, 0, 100))

        final_score = max(0, min(100, final_score))

        # PHASE 6: Percentile rank
        percentile_rank = get_percentile_rank(final_score)

        # Score label
        if final_score >= 80:
            feedback_text   = "🔥 Excellent Ad Potential"
            color_hex, color_glow = "#ccff00", "rgba(204,255,0,0.5)"
        elif final_score >= 60:
            feedback_text   = "👍 Good, But Could Improve"
            color_hex, color_glow = "#facc15", "rgba(250,204,21,0.5)"
        else:
            feedback_text   = "⚠️ Needs Optimization"
            color_hex, color_glow = "#f87171", "rgba(248,113,113,0.5)"

        # PHASE 7: A/B variants
        ab_variants = generate_ab_variants(
            ad_copy, final_score, dominant_color, face_count_val,
            platform, audience, brightness_val, sharpness_val,
            copy_result['has_cta'] if copy_result else False
        ) if vlm_results else []

        # Competitor
        competitor = your_metrics = competitor_metrics = None
        if 'competitor_image' in request.files and request.files['competitor_image'].filename != '':
            try:
                c = score_image_for_compare(request.files['competitor_image'])
                competitor = c
                your_metrics = {
                    "Color Score":    vlm_results["color_score"]            if vlm_results else "—",
                    "Clarity":        vlm_results["visual_clarity_score"]   if vlm_results else "—",
                    "Text Density":   vlm_results["text_density_score"]     if vlm_results else "—",
                    "Emotion":        vlm_results["emotional_impact_score"] if vlm_results else "—",
                    "Sharpness":      sharpness_score or "—",
                }
                competitor_metrics = {
                    "Color Score":    c["color_score"],
                    "Clarity":        c["clarity_score"],
                    "Text Density":   c["text_density_score"],
                    "Emotion":        c["emotional_score"],
                    "Sharpness":      c.get("sharpness", "—"),
                }
            except Exception as e:
                print(f"[AdLens] Competitor error: {e}")

        # AI rewrite
        rewrite = None
        copy_issues = [s for s in suggestions if any(kw in s.lower() for kw in ['cta','urgency','negative','long','short','simplify','spam'])]
        if ad_copy:
            rewrite = ai_rewrite_copy(ad_copy, platform, audience, copy_issues)

        # Save
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        color_psych = COLOR_PSYCHOLOGY.get(dominant_color, {})
        analysis_id = save_analysis({
            "ts": ts, "ad_copy": ad_copy, "ad_type": ad_type,
            "audience": audience, "platform": platform,
            "score": final_score, "ml_score": ml_score,
            "vlm_score": vlm_score, "copy_score": copy_result['copy_score'] if copy_result else None,
            "feedback": feedback_text,
            "insights": json.dumps(insights), "suggestions": json.dumps(suggestions),
            "ctr_estimate": ctr_estimate,
            "face_count": face_count_val if vlm_results else None,
            "text_area_pct": vlm_results.get("text_area_pct") if vlm_results else None,
            "brightness": brightness_val if vlm_results else None,
            "sharpness": sharpness_val if vlm_results else None,
            "symmetry_score": symmetry_val if vlm_results else None,
            "dominant_color": dominant_color if vlm_results else None,
            "color_psychology": color_psych.get("meaning") if vlm_results else None,
            "siglip_luxury": vlm_scores_dict.get("luxury"),
            "siglip_trust":  vlm_scores_dict.get("trust"),
            "siglip_energy": vlm_scores_dict.get("energy"),
            "percentile_rank": percentile_rank,
            "ensemble_breakdown": json.dumps(ensemble_breakdown),
        })

        return render_template_string(
            HTML,
            score=final_score, ml_score=ml_score, vlm_score=vlm_score,
            feedback=feedback_text, color_hex=color_hex, color_glow=color_glow,
            suggestions=suggestions, insights=insights,
            copy_result=copy_result, vlm_results=vlm_results,
            ad_type=ad_type, color=color, audience=audience,
            platform=platform, day=day,
            train_size=TRAIN_SIZE, model_mae=MODEL_MAE, model_r2=MODEL_R2,
            rewrite=rewrite, ad_copy=ad_copy,
            anthropic_key_set=bool(ANTHROPIC_API_KEY),
            platform_label=PLATFORM_LABELS.get(platform, "Social"),
            audience_label=AUDIENCE_LABELS.get(audience, "General"),
            competitor=competitor,
            your_metrics=your_metrics or {},
            competitor_metrics=competitor_metrics or {},
            analysis_id=analysis_id,
            ctr_estimate=ctr_estimate,
            regressor_name=REGRESSOR_NAME,
            vlm_name=VLM_NAME,
            feedback_stats=get_feedback_stats(),
            uploaded_image_data_uri=uploaded_image_data_uri,
            # Phase 3 new
            sharpness_score=sharpness_score,
            symmetry=symmetry_val if vlm_results else None,
            face_count_val=face_count_val if vlm_results else None,
            dominant_color=dominant_color if vlm_results else None,
            color_palette=color_palette if vlm_results else [],
            color_psychology=color_psych if vlm_results else {},
            vlm_scores=vlm_scores_dict,
            caption=caption_text,
            # Phase 5
            ensemble_breakdown=ensemble_breakdown,
            # Phase 6
            percentile_rank=percentile_rank,
            # Phase 7
            ab_variants=ab_variants,
        )

    return render_template_string(
        HTML,
        regressor_name=REGRESSOR_NAME,
        vlm_name=VLM_NAME,
        score=None,
        ml_score=None,
        vlm_score=None,
        feedback=None,
        color_hex="#ff6600",
        color_glow="rgba(255,102,0,0.3)",
        suggestions=[],
        insights=[],
        copy_result=None,
        vlm_results=None,
        ad_type=1,
        color=6,
        audience=3,
        platform=1,
        day=3,
        train_size=TRAIN_SIZE,
        model_mae=MODEL_MAE,
        model_r2=MODEL_R2,
        rewrite=None,
        ad_copy="",
        anthropic_key_set=bool(ANTHROPIC_API_KEY),
        platform_label="Social",
        audience_label="General",
        competitor=None,
        your_metrics={},
        competitor_metrics={},
        analysis_id=None,
        ctr_estimate=None,
        feedback_stats=get_feedback_stats(),
        uploaded_image_data_uri=None,
        sharpness_score=None,
        symmetry=None,
        face_count_val=None,
        dominant_color=None,
        color_palette=[],
        color_psychology={},
        vlm_scores={},
        caption=None,
        ensemble_breakdown=None,
        percentile_rank=None,
        ab_variants=[],
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
