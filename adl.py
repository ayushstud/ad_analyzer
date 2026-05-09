"""
AdLens — Upgraded Edition
=========================
PHASE 1  Real feature columns (CTR, conversion, shares, brightness, face count, text area %)
PHASE 2  Auto-feedback loop  — /feedback endpoint stores user thumbs up/down in SQLite
          and tags every analysis with actual_ctr / actual_conversion fields
PHASE 3  Computer-vision intelligence — face detection (OpenCV Haar cascade),
          color harmony, text area %, brightness/contrast from PIL
          CLIP semantic similarity replaces pure-BLIP captioning when available
PHASE 4  XGBoost replaces GradientBoostingRegressor;
          LightGBM and CatBoost auto-selected when libraries are present
"""

import re
import os
import io
import json
import time
import sqlite3
import datetime
import numpy as np
import openpyxl
import torch
from PIL import Image, ImageStat
from transformers import BlipProcessor, BlipForConditionalGeneration
from flask import Flask, request, render_template_string, jsonify
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, r2_score

# ── Optional heavy deps — graceful fallback everywhere ──────────────────────

try:
    import shap
    SHAP_AVAILABLE = True
except ImportError:
    SHAP_AVAILABLE = False
    print("[AdLens] shap not installed — pip install shap")

try:
    import easyocr
    _ocr_reader = easyocr.Reader(['en'], gpu=False)
    OCR_BACKEND = "easyocr"
except ImportError:
    try:
        import pytesseract
        OCR_BACKEND = "pytesseract"
    except ImportError:
        OCR_BACKEND = None

# ── PHASE 4: Best available regressor ───────────────────────────────────────
try:
    from xgboost import XGBRegressor
    REGRESSOR_NAME = "XGBoost"
    def make_regressor():
        return XGBRegressor(
            n_estimators=200, max_depth=5, learning_rate=0.07,
            subsample=0.85, colsample_bytree=0.85,
            random_state=42, verbosity=0
        )
except ImportError:
    try:
        import lightgbm as lgb
        REGRESSOR_NAME = "LightGBM"
        def make_regressor():
            return lgb.LGBMRegressor(
                n_estimators=200, max_depth=5, learning_rate=0.07,
                random_state=42, verbose=-1
            )
    except ImportError:
        from sklearn.ensemble import GradientBoostingRegressor
        REGRESSOR_NAME = "GradientBoosting"
        def make_regressor():
            return GradientBoostingRegressor(
                n_estimators=150, max_depth=4, learning_rate=0.08, random_state=42
            )

print(f"[AdLens] Using regressor: {REGRESSOR_NAME}")

# ── PHASE 3: CLIP semantic similarity ───────────────────────────────────────
CLIP_AVAILABLE = False
clip_model = clip_processor = None
try:
    from transformers import CLIPProcessor, CLIPModel
    clip_processor = CLIPProcessor.from_pretrained("openai/clip-vit-base-patch32")
    clip_model     = CLIPModel.from_pretrained("openai/clip-vit-base-patch32")
    CLIP_AVAILABLE = True
    print("[AdLens] CLIP loaded successfully")
except Exception as e:
    print(f"[AdLens] CLIP not available ({e}) — using BLIP only")

# ── PHASE 3: OpenCV face detection ──────────────────────────────────────────
CV2_AVAILABLE = False
face_cascade = None
try:
    import cv2
    _cascade_path = cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
    face_cascade  = cv2.CascadeClassifier(_cascade_path)
    CV2_AVAILABLE = True
    print("[AdLens] OpenCV face detection ready")
except ImportError:
    print("[AdLens] cv2 not installed — pip install opencv-python")

# ── MediaPipe emotion (optional) ─────────────────────────────────────────────
MEDIAPIPE_AVAILABLE = False
try:
    import mediapipe as mp
    MEDIAPIPE_AVAILABLE = True
except ImportError:
    pass

app = Flask(__name__)

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

# ─────────────────────────────────────────────────────────────────────────────
# DATABASE  (extended schema — backward compat with existing rows)
# ─────────────────────────────────────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "adlens_history.db")

def init_db():
    con = sqlite3.connect(DB_PATH)

    # Original table
    con.execute("""
        CREATE TABLE IF NOT EXISTS analyses (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            ts              TEXT    NOT NULL,
            ad_copy         TEXT,
            ad_type         INTEGER,
            audience        INTEGER,
            platform        INTEGER,
            score           INTEGER,
            ml_score        INTEGER,
            vlm_score       INTEGER,
            copy_score      INTEGER,
            feedback        TEXT,
            insights        TEXT,
            suggestions     TEXT
        )
    """)

    # PHASE 1: new real-world metric columns (ALTER TABLE if upgrading)
    _add_col(con, "analyses", "ctr_estimate",        "REAL")
    _add_col(con, "analyses", "face_count",          "INTEGER")
    _add_col(con, "analyses", "text_area_pct",       "REAL")
    _add_col(con, "analyses", "brightness",          "REAL")
    _add_col(con, "analyses", "clip_luxury_score",   "REAL")
    _add_col(con, "analyses", "clip_trust_score",    "REAL")
    _add_col(con, "analyses", "clip_energy_score",   "REAL")

    # PHASE 2: feedback loop columns
    _add_col(con, "analyses", "user_rating",         "INTEGER")   # 1=thumbs up, 0=thumbs down
    _add_col(con, "analyses", "actual_ctr",          "REAL")      # user can fill in later
    _add_col(con, "analyses", "actual_conversion",   "REAL")
    _add_col(con, "analyses", "prediction_accurate", "INTEGER")   # 1/0

    # PHASE 2: feedback table (keeps individual feedback events)
    con.execute("""
        CREATE TABLE IF NOT EXISTS feedback (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            analysis_id INTEGER NOT NULL,
            ts          TEXT    NOT NULL,
            rating      INTEGER NOT NULL,
            comment     TEXT,
            FOREIGN KEY (analysis_id) REFERENCES analyses(id)
        )
    """)

    con.commit()
    con.close()

def _add_col(con, table, col, dtype):
    try:
        con.execute(f"ALTER TABLE {table} ADD COLUMN {col} {dtype}")
    except sqlite3.OperationalError:
        pass  # column already exists

def save_analysis(ts, ad_copy, ad_type, audience, platform, score,
                  ml_score, vlm_score, copy_score, feedback_text, insights,
                  suggestions, ctr_estimate=None, face_count=None,
                  text_area_pct=None, brightness=None,
                  clip_luxury=None, clip_trust=None, clip_energy=None):
    con = sqlite3.connect(DB_PATH)
    cur = con.execute("""
        INSERT INTO analyses
            (ts, ad_copy, ad_type, audience, platform, score, ml_score,
             vlm_score, copy_score, feedback, insights, suggestions,
             ctr_estimate, face_count, text_area_pct, brightness,
             clip_luxury_score, clip_trust_score, clip_energy_score)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (ts, ad_copy, ad_type, audience, platform, score, ml_score,
          vlm_score, copy_score, feedback_text,
          json.dumps(insights), json.dumps(suggestions),
          ctr_estimate, face_count, text_area_pct, brightness,
          clip_luxury, clip_trust, clip_energy))
    analysis_id = cur.lastrowid
    con.commit()
    con.close()
    return analysis_id

def save_feedback(analysis_id, rating, comment=""):
    """PHASE 2: store per-analysis user rating."""
    con = sqlite3.connect(DB_PATH)
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    con.execute(
        "INSERT INTO feedback (analysis_id, ts, rating, comment) VALUES (?,?,?,?)",
        (analysis_id, ts, rating, comment)
    )
    con.execute(
        "UPDATE analyses SET user_rating=?, prediction_accurate=? WHERE id=?",
        (rating, rating, analysis_id)
    )
    con.commit()
    con.close()

def load_history(limit=50):
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    rows = con.execute(
        "SELECT * FROM analyses ORDER BY id DESC LIMIT ?", (limit,)
    ).fetchall()
    con.close()
    return [dict(r) for r in rows]

def get_feedback_stats():
    """Returns accuracy rate from user feedback — real-world learning signal."""
    con = sqlite3.connect(DB_PATH)
    total = con.execute("SELECT COUNT(*) FROM feedback").fetchone()[0]
    positive = con.execute("SELECT COUNT(*) FROM feedback WHERE rating=1").fetchone()[0]
    con.close()
    if total == 0:
        return {"total": 0, "accuracy_rate": None}
    return {"total": total, "accuracy_rate": round(positive / total * 100, 1)}

init_db()

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 4: IMPROVED MODEL TRAINING
# ─────────────────────────────────────────────────────────────────────────────
DATASET_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "adlens_hybrid_datatset.xlsx")

# PHASE 1: Extended feature set — maps dataset column names to defaults
FEATURE_NAMES = [
    "Headline Len", "Ad Type", "Color Score", "Audience", "Platform",
    "Hour", "Day", "Audience×Platform Match", "Peak Hour Flag",
    "Visual Sentiment Boost",
    # PHASE 1 additions
    "CTR (raw)", "Brightness", "Emotional Impact",
]

FEATURE_DISPLAY = [
    "Headline Length", "Ad Type", "Color Score", "Audience", "Platform",
    "Hour", "Day", "Audience×Platform Match", "Peak Hour Flag",
    "Visual Sentiment Boost", "CTR Signal", "Brightness", "Emotional Impact",
]

FEATURE_DEFAULTS = [8, 1, 5, 3, 1, 18, 3, 0, 0, 0, 0.03, 128, 6]

def load_and_train():
    wb = openpyxl.load_workbook(DATASET_PATH)
    ws = wb["Hybrid Dataset"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    X, y = [], []
    for r in range(2, ws.max_row + 1):
        row = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        try:
            feats = [
                float(row.get("Headline Len")              or 8),
                float(row.get("Ad Type")                   or 1),
                float(row.get("Color Score")               or 5),
                float(row.get("Audience")                  or 3),
                float(row.get("Platform")                  or 1),
                float(row.get("Hour")                      or 18),
                float(row.get("Day")                       or 3),
                float(row.get("Audience\u00d7Platform Match") or 0),
                float(row.get("Peak Hour Flag")            or 0),
                float(row.get("Visual Sentiment Boost")    or 0),
                # PHASE 1: real engagement signals
                float(row.get("CTR (raw)")                 or 0.03),
                float(row.get("Brightness")                or 128),
                float(row.get("Emotional Impact")          or 6),
            ]
            target = float(row.get("Final Score") or 50)
            X.append(feats)
            y.append(target)
        except (TypeError, ValueError):
            continue

    X, y = np.array(X), np.array(y)
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    scaler  = StandardScaler()
    X_train_s = scaler.fit_transform(X_train)
    X_test_s  = scaler.transform(X_test)

    clf = make_regressor()
    clf.fit(X_train_s, y_train)

    preds = clf.predict(X_test_s)
    mae   = round(mean_absolute_error(y_test, preds), 2)
    r2    = round(r2_score(y_test, preds), 3)

    explainer = None
    if SHAP_AVAILABLE:
        try:
            explainer = shap.TreeExplainer(clf)
        except Exception as e:
            print(f"[AdLens] SHAP init failed: {e}")

    print(f"[AdLens] {REGRESSOR_NAME} trained on {len(X_train)} samples | "
          f"Tested on {len(X_test)} | MAE={mae} | R²={r2}")
    return clf, scaler, explainer, mae, r2, len(X_train), len(X_test)

print("[AdLens] Loading dataset and training model...")
model, scaler, SHAP_EXPLAINER, MODEL_MAE, MODEL_R2, TRAIN_SIZE, TEST_SIZE = load_and_train()

# ─────────────────────────────────────────────────────────────────────────────
# ENGINEERED FEATURES
# ─────────────────────────────────────────────────────────────────────────────
HIGH_SYNERGY = {(4,2),(1,1),(5,1),(1,4),(3,3),(2,4),(4,4)}

def get_engineered_features(ad_type, audience, platform, hour, text_density, emotional_impact):
    apm = 1 if (platform, audience) in HIGH_SYNERGY else 0
    phf = 1 if 18 <= hour <= 21 else 0
    vsb = round(emotional_impact * 1.2 - text_density * 0.5, 2)
    return apm, phf, vsb

# ─────────────────────────────────────────────────────────────────────────────
# SHAP EXPLAINABILITY
# ─────────────────────────────────────────────────────────────────────────────
def get_shap_insights(features_scaled, features_raw):
    if not SHAP_AVAILABLE or SHAP_EXPLAINER is None:
        return []
    try:
        shap_values = SHAP_EXPLAINER.shap_values(features_scaled)
        # XGBoost / LightGBM return array directly; GBR returns list
        if isinstance(shap_values, list):
            sv = shap_values[0]
        else:
            sv = shap_values[0] if shap_values.ndim > 1 else shap_values
        contributions = list(zip(FEATURE_DISPLAY, sv))
        contributions.sort(key=lambda x: abs(x[1]), reverse=True)
        insights = []
        for name, val in contributions[:5]:
            direction = "(+)" if val > 0 else "(-)"
            impact    = abs(round(val, 1))
            if impact < 0.5:
                continue
            insights.append(
                f"🧠 SHAP {direction} <b>{name}</b> contributed "
                f"{'+' if val>0 else '-'}{impact} pts to this prediction."
            )
        return insights
    except Exception as e:
        print(f"[AdLens] SHAP inference error: {e}")
        return []

# ─────────────────────────────────────────────────────────────────────────────
# AD COPY ANALYZER
# ─────────────────────────────────────────────────────────────────────────────
POSITIVE_WORDS = {'amazing','best','free','new','save','win','exclusive','limited','offer',
                  'guaranteed','proven','trusted','top','premium','instant','easy','fast',
                  'boost','grow','improve','powerful','smart','innovative','leading','award',
                  'special','unique','perfect','great','excellent','outstanding','incredible'}
NEGATIVE_WORDS = {'bad','worst','fail','lose','costly','expensive','difficult','hard',
                  'slow','weak','outdated','broken','cheap','risky','uncertain','avoid','worst'}
CTA_WORDS      = {'buy','shop','get','try','start','join','sign','click','learn','discover',
                  'order','subscribe','download','register','book','claim','grab','explore',
                  'call','visit','watch','read','find','see','check','use','apply'}
URGENCY_WORDS  = {'now','today','limited','hurry','last','expires','urgent','soon','only',
                  'left','deadline','ending','final','instant','immediately','quickly'}

def analyze_ad_copy(text):
    if not text or not text.strip():
        return None
    text      = text.strip()
    words     = text.split()
    word_count= len(words)
    text_lower= text.lower()
    word_set  = set(re.findall(r'\b\w+\b', text_lower))
    pos_hits     = word_set & POSITIVE_WORDS
    neg_hits     = word_set & NEGATIVE_WORDS
    cta_hits     = word_set & CTA_WORDS
    urgency_hits = word_set & URGENCY_WORDS
    raw_sentiment   = (len(pos_hits) - len(neg_hits)) / max(word_count, 1)
    sentiment_score = round(min(1.0, max(-1.0, raw_sentiment * 10)), 3)
    has_cta          = len(cta_hits) > 0
    has_urgency      = len(urgency_hits) > 0
    has_question     = '?' in text
    exclamation_count= text.count('!')
    avg_word_len = sum(len(w) for w in words) / max(word_count, 1)
    readability  = round(min(10, max(1, 10 - avg_word_len + 2)), 1)
    copy_score = 50
    copy_score += sentiment_score * 20
    copy_score += 10 if has_cta     else -10
    copy_score += 8  if has_urgency else 0
    copy_score += 5  if has_question else 0
    copy_score += min(5, exclamation_count * 2)
    copy_score += (readability - 5) * 2
    if word_count < 3:  copy_score -= 15
    if word_count > 25: copy_score -= 10
    spam_ratio   = len(pos_hits) / max(len(word_set), 1)
    is_spam_like = spam_ratio > 0.4 and word_count < 15
    if is_spam_like:
        copy_score -= 20
    copy_score = int(min(100, max(0, copy_score)))
    if sentiment_score > 0.3:    sentiment_label = "Positive"
    elif sentiment_score < -0.3: sentiment_label = "Negative"
    else:                         sentiment_label = "Neutral"
    return {
        'word_count': word_count, 'sentiment_score': sentiment_score,
        'sentiment_label': sentiment_label, 'has_cta': has_cta,
        'cta_words_found': list(cta_hits)[:3], 'has_urgency': has_urgency,
        'urgency_words': list(urgency_hits)[:3], 'readability': readability,
        'has_question': has_question, 'copy_score': copy_score,
        'is_spam_like': is_spam_like, 'spam_ratio': round(spam_ratio, 2),
    }

# ─────────────────────────────────────────────────────────────────────────────
# BLIP + PHASE 3 COMPUTER VISION
# ─────────────────────────────────────────────────────────────────────────────
device    = "cuda" if torch.cuda.is_available() else "cpu"
processor = BlipProcessor.from_pretrained("Salesforce/blip-image-captioning-base")
blip_model= BlipForConditionalGeneration.from_pretrained(
    "Salesforce/blip-image-captioning-base").to(device)

# PHASE 3: CLIP concept probes
CLIP_CONCEPTS = {
    "luxury":     ["luxury product", "premium quality", "high end brand", "elegant design"],
    "trust":      ["trustworthy brand", "professional service", "reliable product", "credible"],
    "energy":     ["exciting energy", "vibrant colors", "dynamic action", "high CTR ad style"],
    "minimalist": ["clean minimal design", "white space", "simple layout"],
    "crowded":    ["cluttered design", "too much text", "busy layout", "overwhelming ad"],
}

def clip_score_concepts(pil_image):
    """PHASE 3: Use CLIP to score semantic ad qualities. Returns dict of concept→score."""
    if not CLIP_AVAILABLE:
        return {}
    try:
        results = {}
        for concept, prompts in CLIP_CONCEPTS.items():
            inputs = clip_processor(
                text=prompts, images=pil_image,
                return_tensors="pt", padding=True
            )
            with torch.no_grad():
                outputs = clip_model(**inputs)
            probs = outputs.logits_per_image.softmax(dim=1)
            results[concept] = round(float(probs.max()), 3)
        return results
    except Exception as e:
        print(f"[AdLens] CLIP scoring error: {e}")
        return {}

def detect_faces(pil_image):
    """PHASE 3: Count faces using OpenCV Haar cascade."""
    if not CV2_AVAILABLE or face_cascade is None:
        return 0
    try:
        import cv2
        img_np = np.array(pil_image.convert("RGB"))
        gray   = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)
        faces  = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))
        return len(faces) if isinstance(faces, np.ndarray) else 0
    except Exception as e:
        print(f"[AdLens] Face detection error: {e}")
        return 0

def estimate_text_area_pct(pil_image):
    """
    PHASE 3: Rough estimate of % image area covered by text/high-contrast edges.
    Uses PIL without CV2. Returns 0-100.
    """
    try:
        # Convert to grayscale and check edge density (proxy for text area)
        gray  = pil_image.convert("L")
        w, h  = gray.size
        # Simple threshold: pixels < 80 or > 180 (high contrast = likely text or logo)
        pixels = list(gray.getdata())
        high_contrast = sum(1 for p in pixels if p < 60 or p > 200)
        return round(high_contrast / len(pixels) * 100, 1)
    except Exception:
        return 0.0

def estimate_ctr(score, face_count, clip_scores, has_cta, brightness):
    """
    PHASE 1: Synthetic CTR estimate based on ad quality signals.
    Mimics real CTR ranges (0.5%–8%) for display ads.
    """
    base = 1.5
    base += score * 0.04          # better score → higher CTR
    base += face_count * 0.3      # human faces boost CTR
    base += clip_scores.get("energy", 0) * 2.0
    base += clip_scores.get("trust",  0) * 1.5
    base -= clip_scores.get("crowded", 0) * 2.0
    base += 0.8 if has_cta else 0
    if brightness < 80:   base -= 0.5   # too dark
    if brightness > 220:  base -= 0.3   # overexposed
    return round(max(0.3, min(9.0, base)), 2)

def analyze_image(image_file):
    """Full image analysis: BLIP + PHASE 3 CV + CLIP."""
    image = Image.open(image_file).convert('RGB')

    # ── Basic stats ──────────────────────────────────────────────────────────
    stat       = ImageStat.Stat(image.convert("L"))
    brightness = stat.mean[0]
    contrast   = stat.stddev[0]

    color_score          = min(10, max(1, int((brightness / 255) * 10)))
    visual_clarity_score = min(10, max(1, int((contrast  / 128) * 10)))

    # ── BLIP caption ─────────────────────────────────────────────────────────
    inputs  = processor(images=image, return_tensors="pt").to(device)
    out     = blip_model.generate(**inputs)
    caption = processor.decode(out[0], skip_special_tokens=True)

    caption_lower = caption.lower()
    text_density_score = 5
    if any(w in caption_lower for w in ['text','sign','writing','words','poster','logo']):
        text_density_score = 8

    emotional_impact_score = 6
    if any(w in caption_lower for w in ['smile','happy','laugh','people','person','exciting','bright','vibrant']):
        emotional_impact_score = 9
    elif any(w in caption_lower for w in ['dark','empty','boring','bland']):
        emotional_impact_score = 3

    # ── PHASE 3: Computer Vision extras ─────────────────────────────────────
    face_count    = detect_faces(image)
    text_area_pct = estimate_text_area_pct(image)
    clip_scores   = clip_score_concepts(image)

    # Refine emotional_impact with face count
    if face_count > 0:
        emotional_impact_score = min(10, emotional_impact_score + face_count)

    # Refine text_density with actual estimate
    if text_area_pct > 40:
        text_density_score = min(10, text_density_score + 2)

    # ── OCR ──────────────────────────────────────────────────────────────────
    ocr_text = _run_ocr(image)

    return {
        "caption":               caption,
        "color_score":           color_score,
        "visual_clarity_score":  visual_clarity_score,
        "text_density_score":    text_density_score,
        "emotional_impact_score":emotional_impact_score,
        "ocr_text":              ocr_text,
        # PHASE 3 additions
        "face_count":            face_count,
        "text_area_pct":         text_area_pct,
        "brightness":            round(brightness, 1),
        "contrast":              round(contrast, 1),
        "clip_scores":           clip_scores,
    }

def _run_ocr(pil_image):
    if OCR_BACKEND is None:
        return None
    try:
        if OCR_BACKEND == "easyocr":
            results = _ocr_reader.readtext(np.array(pil_image))
            return " ".join(r[1] for r in results).strip() or None
        else:
            import pytesseract
            text = pytesseract.image_to_string(pil_image).strip()
            return text or None
    except Exception as e:
        print(f"[AdLens] OCR error: {e}")
        return None

# ─────────────────────────────────────────────────────────────────────────────
# AI REWRITE  (Claude via urllib)
# ─────────────────────────────────────────────────────────────────────────────
import urllib.request

PLATFORM_LABELS = {1: "Instagram", 2: "YouTube", 3: "Facebook"}
AUDIENCE_LABELS = {1: "Gen Z / Students", 2: "Professionals (B2B)", 3: "General / Broad"}

def ai_rewrite_copy(original_copy, platform_id, audience_id, issues):
    if not ANTHROPIC_API_KEY or not original_copy:
        return None
    platform    = PLATFORM_LABELS.get(platform_id, "social media")
    audience    = AUDIENCE_LABELS.get(audience_id, "general audience")
    issues_text = "; ".join(issues[:4]) if issues else "improve overall persuasion"
    prompt = (
        f"You are an expert advertising copywriter. "
        f"Rewrite the following ad copy for {platform}, targeting {audience}.\n\n"
        f"Original copy: \"{original_copy}\"\n\n"
        f"Issues to fix: {issues_text}\n\n"
        f"Rules:\n"
        f"- Maximum 12 words\n"
        f"- Include one clear CTA\n"
        f"- Match the platform tone (Instagram = casual/visual, YouTube = direct, Facebook = benefits-led)\n"
        f"- No hype words like 'amazing', 'best ever', 'incredible'\n"
        f"- Return ONLY the rewritten copy, nothing else, no quotes."
    )
    payload = json.dumps({
        "model":      "claude-sonnet-4-20250514",
        "max_tokens": 120,
        "messages":   [{"role": "user", "content": prompt}]
    }).encode()
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages", data=payload,
        headers={"Content-Type": "application/json",
                 "x-api-key":    ANTHROPIC_API_KEY,
                 "anthropic-version": "2023-06-01"},
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
            return data["content"][0]["text"].strip()
    except Exception as e:
        print(f"[AdLens] AI rewrite error: {e}")
        return None

# ─────────────────────────────────────────────────────────────────────────────
# COMPETITOR SCORING
# ─────────────────────────────────────────────────────────────────────────────
def score_image_for_compare(image_file):
    vlm = analyze_image(image_file)
    vlm_score = int(np.mean([
        vlm["color_score"], vlm["visual_clarity_score"],
        vlm["text_density_score"], vlm["emotional_impact_score"],
    ]) * 10)
    return {
        "vlm_score":          vlm_score,
        "color_score":        vlm["color_score"],
        "clarity_score":      vlm["visual_clarity_score"],
        "text_density_score": vlm["text_density_score"],
        "emotional_score":    vlm["emotional_impact_score"],
        "caption":            vlm["caption"],
        "ocr_text":           vlm.get("ocr_text"),
        "face_count":         vlm.get("face_count", 0),
        "clip_scores":        vlm.get("clip_scores", {}),
    }

# ─────────────────────────────────────────────────────────────────────────────
# HTML TEMPLATES
# ─────────────────────────────────────────────────────────────────────────────
HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>AdLens - Analyze your ad</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        :root {
            --bg-primary: #0a0a0c; --bg-secondary: #131417;
            --accent-primary: #ccff00; --accent-glow: rgba(204, 255, 0, 0.3);
            --success: #ccff00; --warning: #facc15; --danger: #f87171;
            --text-main: #ffffff; --text-muted: #a1a1aa;
            --glass-border: rgba(255, 255, 255, 0.08); --btn-text: #000;
        }
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            background-color: var(--bg-primary);
            background-image: linear-gradient(rgba(255,255,255,0.03) 1px,transparent 1px),linear-gradient(90deg,rgba(255,255,255,0.03) 1px,transparent 1px);
            background-size: 50px 50px; color: var(--text-main); font-family: 'Outfit', sans-serif;
            min-height: 100vh; display: flex; flex-direction: column; padding: 0 2rem 3rem 2rem; overflow-x: hidden;
        }
        .container { max-width: 900px; width: 100%; margin: 0 auto; display: flex; flex-direction: column; }
        nav { display: flex; justify-content: space-between; align-items: center; padding: 2rem 0; margin-bottom: 3rem; }
        .logo { font-size: 1.6rem; font-weight: 800; display: flex; align-items: center; gap: 0.5rem; letter-spacing: -0.02em; }
        .logo-icon { background: var(--accent-primary); color: #000; padding: 0.2rem; border-radius: 8px; font-size: 1.2rem; display: flex; align-items: center; justify-content: center; width: 32px; height: 32px; }
        .badge { padding: 0.4rem 1.2rem; border: 1px solid rgba(255,255,255,0.15); background: rgba(255,255,255,0.03); border-radius: 20px; font-size: 0.75rem; color: var(--text-muted); font-weight: 600; letter-spacing: 0.05em; }
        header { text-align: center; margin-bottom: 3.5rem; }
        header h1 { font-size: 5rem; font-weight: 900; line-height: 1.05; margin-bottom: 1.5rem; letter-spacing: -0.03em; text-transform: capitalize; }
        header h1 .highlight { color: var(--accent-primary); display: block; }
        header p { color: var(--text-muted); font-size: 1.15rem; max-width: 650px; margin: 0 auto; line-height: 1.6; font-weight: 300; }
        @media (max-width: 768px) { header h1 { font-size: 3.5rem; } }
        .form-panel { background: rgba(255,255,255,0.01); border: 1px solid transparent; border-radius: 24px; padding: 0; }
        .type-selector { display: flex; justify-content: center; gap: 0.8rem; margin-bottom: 2.5rem; flex-wrap: wrap; }
        .radio-pill input { display: none; }
        .radio-pill .pill-content { background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.1); padding: 0.7rem 1.2rem; border-radius: 12px; color: var(--text-muted); cursor: pointer; display: flex; align-items: center; gap: 0.5rem; transition: all 0.2s; font-weight: 500; font-size: 0.95rem; }
        .radio-pill:hover .pill-content { background: rgba(255,255,255,0.08); color: white; }
        .radio-pill input:checked + .pill-content { background: var(--accent-primary); color: var(--btn-text); border-color: var(--accent-primary); font-weight: 600; }
        .upload-area { background: #111113; border: 1px dashed rgba(255,255,255,0.2); border-radius: 20px; padding: 4.5rem 2rem; text-align: center; cursor: pointer; transition: all 0.3s; margin-bottom: 1.5rem; position: relative; overflow: hidden; display: flex; flex-direction: column; align-items: center; justify-content: center; }
        .upload-area:hover { border-color: rgba(255,255,255,0.4); background: #161619; }
        .upload-icon { background: #3b82f6; color: white; width: 54px; height: 54px; border-radius: 14px; display: flex; align-items: center; justify-content: center; font-size: 1.8rem; margin-bottom: 1.5rem; box-shadow: 0 4px 20px rgba(59,130,246,0.4); }
        .upload-title { font-size: 1.6rem; font-weight: 700; margin-bottom: 0.5rem; color: white; }
        .upload-subtitle { color: var(--text-muted); font-size: 0.95rem; }
        #imageInput { position: absolute; top: 0; left: 0; width: 100%; height: 100%; opacity: 0; cursor: pointer; z-index: 10; }
        #imagePreview { display: none; max-width: 100%; max-height: 400px; border-radius: 12px; object-fit: contain; }
        .upload-content-wrapper { display: flex; flex-direction: column; align-items: center; z-index: 5; pointer-events: none; }
        .copy-group { margin-bottom: 1.5rem; }
        .copy-group label { display: block; margin-bottom: 0.6rem; color: var(--text-muted); font-size: 0.85rem; font-weight: 500; text-transform: uppercase; letter-spacing: 0.05em; }
        .copy-group textarea { width: 100%; padding: 0.9rem 1rem; background: rgba(0,0,0,0.4); border: 1px solid rgba(255,255,255,0.1); border-radius: 10px; color: white; font-family: 'Outfit', sans-serif; font-size: 0.95rem; transition: all 0.3s; resize: vertical; min-height: 80px; }
        .copy-group textarea:focus { outline: none; border-color: var(--accent-primary); }
        .settings-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 1.5rem; margin-bottom: 2rem; padding: 2rem; border: 1px dashed rgba(255,255,255,0.1); border-radius: 16px; background: rgba(0,0,0,0.2); }
        .form-group label { display: block; margin-bottom: 0.6rem; color: var(--text-muted); font-size: 0.85rem; font-weight: 500; text-transform: uppercase; letter-spacing: 0.05em; }
        input, select { width: 100%; padding: 0.9rem 1rem; background: rgba(0,0,0,0.4); border: 1px solid rgba(255,255,255,0.1); border-radius: 10px; color: white; font-family: 'Outfit', sans-serif; font-size: 0.95rem; transition: all 0.3s; }
        input:focus, select:focus { outline: none; border-color: var(--accent-primary); }
        .btn-submit { width: 100%; padding: 1.25rem; background: linear-gradient(180deg,#788d37 0%,#516027 100%); color: #fff; border: 1px solid rgba(255,255,255,0.05); border-radius: 16px; font-size: 1.15rem; font-weight: 700; font-family: 'Outfit',sans-serif; cursor: pointer; transition: all 0.3s; box-shadow: inset 0 1px 0 rgba(255,255,255,0.15); display: flex; align-items: center; justify-content: center; gap: 0.8rem; }
        .btn-submit:hover { background: linear-gradient(180deg,#879f3e 0%,#5b6d2c 100%); transform: translateY(-1px); }
        .advanced-settings-toggle { text-align: center; color: var(--text-muted); font-size: 0.9rem; cursor: pointer; margin-bottom: 1.5rem; display: flex; align-items: center; justify-content: center; gap: 0.5rem; transition: color 0.2s; padding: 0.5rem; }
        .advanced-settings-toggle:hover { color: #fff; }
        .results-wrapper { margin-top: 4rem; padding-top: 4rem; border-top: 1px solid rgba(255,255,255,0.05); }
        .results-panel { display: flex; flex-direction: column; gap: 1.5rem; }
        .score-container { display: flex; align-items: center; justify-content: center; flex-direction: column; padding: 4rem; background: rgba(255,255,255,0.01); border-radius: 24px; border: 1px solid var(--glass-border); position: relative; }
        .score-circle { width: 200px; height: 200px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 5rem; font-weight: 800; background: conic-gradient(var(--score-color) var(--score-degrees), rgba(255,255,255,0.03) 0); position: relative; box-shadow: 0 0 50px var(--score-glow); animation: fadeInScale 0.8s ease-out forwards; }
        .score-circle::before { content: ""; position: absolute; inset: 12px; background: var(--bg-primary); border-radius: 50%; }
        .score-value { position: relative; z-index: 10; color: #fff; letter-spacing: -0.05em; }
        .feedback-badge { margin-top: 2.5rem; padding: 0.8rem 2.5rem; border-radius: 30px; font-weight: 700; font-size: 1.25rem; background: rgba(0,0,0,0.4); border: 1px solid; animation: slideUp 0.5s ease-out 0.3s both; }
        .metrics-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 1.5rem; }
        .metric-card { background: rgba(255,255,255,0.02); border: 1px solid var(--glass-border); padding: 2.5rem; border-radius: 20px; animation: slideUp 0.5s ease-out both; }
        .list-items { list-style: none; margin-top: 1.5rem; }
        .list-items li { padding: 1.2rem 0; border-bottom: 1px solid rgba(255,255,255,0.03); display: flex; align-items: flex-start; gap: 1rem; font-size: 1.1rem; color: #d4d4d8; line-height: 1.6; }
        .list-items li:last-child { border-bottom: none; }
        .icon-good { color: var(--success); font-size: 1.4rem; }
        .icon-bad  { color: var(--danger);  font-size: 1.4rem; }
        .icon-warn { color: var(--warning); font-size: 1.4rem; }
        .copy-pill { display:inline-flex; align-items:center; gap:0.5rem; margin-top:0.8rem; padding:0.4rem 1rem; border-radius:20px; font-size:0.85rem; font-weight:600; border:1px solid rgba(255,255,255,0.1); background:rgba(255,255,255,0.04); color:var(--text-muted); }
        .copy-pill.good { border-color:rgba(204,255,0,0.3); color:var(--accent-primary); }
        .copy-pill.warn { border-color:rgba(250,204,21,0.3); color:var(--warning); }
        .copy-pill.bad  { border-color:rgba(248,113,113,0.3); color:var(--danger); }
        @keyframes fadeInScale { from { opacity:0; transform:scale(0.8); } to { opacity:1; transform:scale(1); } }
        @keyframes slideUp     { from { opacity:0; transform:translateY(20px); } to { opacity:1; transform:translateY(0); } }
        /* PHASE 2: Feedback widget */
        .feedback-widget { background: rgba(99,102,241,0.06); border: 1px solid rgba(99,102,241,0.2); border-radius: 16px; padding: 1.5rem 2rem; display: flex; align-items: center; gap: 1.5rem; flex-wrap: wrap; }
        .feedback-widget span { color: var(--text-muted); font-size: 0.95rem; font-weight: 500; }
        .thumb-btn { background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.1); border-radius: 10px; padding: 0.6rem 1.4rem; cursor: pointer; font-size: 1.3rem; transition: all 0.2s; color: white; font-family: 'Outfit', sans-serif; }
        .thumb-btn:hover { background: rgba(255,255,255,0.1); transform: scale(1.05); }
        .thumb-btn.selected-up   { background: rgba(204,255,0,0.15); border-color: rgba(204,255,0,0.4); }
        .thumb-btn.selected-down { background: rgba(248,113,113,0.15); border-color: rgba(248,113,113,0.4); }
        .feedback-thanks { color: var(--accent-primary); font-size: 0.9rem; font-weight: 600; display: none; }
        /* PHASE 3: CV badges */
        .cv-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 1rem; margin-top: 1rem; }
        .cv-badge { background: rgba(0,0,0,0.3); border: 1px solid rgba(255,255,255,0.08); border-radius: 12px; padding: 1rem; text-align: center; }
        .cv-badge .cv-val { font-size: 1.6rem; font-weight: 800; color: var(--accent-primary); }
        .cv-badge .cv-lbl { font-size: 0.75rem; color: var(--text-muted); margin-top: 0.3rem; text-transform: uppercase; letter-spacing: 0.05em; }
        /* CLIP bars */
        .clip-bar-wrap { margin-bottom: 0.6rem; }
        .clip-bar-label { display: flex; justify-content: space-between; font-size: 0.82rem; color: var(--text-muted); margin-bottom: 0.3rem; }
        .clip-bar-bg { background: rgba(255,255,255,0.06); border-radius: 4px; height:6px; }
        .clip-bar-fill { height:6px; border-radius:4px; background: var(--accent-primary); transition: width 0.6s ease; }
        /* History / compare / rewrite (same as original) */
        .compare-grid { display:grid; grid-template-columns:1fr 1fr; gap:1.5rem; }
        .compare-col  { background:rgba(255,255,255,0.02); border:1px solid var(--glass-border); border-radius:16px; padding:1.5rem; }
        .compare-col h4 { font-size:0.9rem; text-transform:uppercase; letter-spacing:0.06em; color:var(--text-muted); margin-bottom:1rem; font-weight:600; }
        .compare-bar-wrap { margin-bottom:0.6rem; }
        .compare-bar-label { display:flex; justify-content:space-between; font-size:0.82rem; color:var(--text-muted); margin-bottom:0.3rem; }
        .compare-bar-bg { background:rgba(255,255,255,0.06); border-radius:4px; height:6px; }
        .compare-bar-fill { height:6px; border-radius:4px; transition:width 0.6s ease; }
        .bar-you { background:var(--accent-primary); }
        .bar-them { background:#6366f1; }
        .rewrite-card { background:rgba(99,102,241,0.08); border:1px solid rgba(99,102,241,0.25); border-radius:16px; padding:1.5rem; margin-top:1rem; }
        .rewrite-copy { font-size:1.3rem; font-weight:700; color:#fff; line-height:1.4; margin:0.75rem 0; }
        .rewrite-note { font-size:0.82rem; color:var(--text-muted); }
        /* CTR estimate badge */
        .ctr-badge { display:inline-flex; align-items:center; gap:0.5rem; padding:0.5rem 1.2rem; border-radius:20px; font-weight:700; font-size:1rem; background:rgba(204,255,0,0.1); border:1px solid rgba(204,255,0,0.25); color:var(--accent-primary); margin-top:1rem; }
    </style>
</head>
<body>
<div class="container">
    <nav>
        <div class="logo">
            <span class="logo-icon">🔍</span>
            <span>Ad<span style="color:var(--accent-primary);">Lens</span></span>
        </div>
        <div class="badge">{{ regressor_name }} · CLIP · CV</div>
        <a href="/history" style="font-size:0.85rem;color:var(--text-muted);text-decoration:none;padding:0.4rem 1rem;border:1px solid rgba(255,255,255,0.1);border-radius:20px;transition:color 0.2s;" onmouseover="this.style.color='#fff'" onmouseout="this.style.color=''">📋 History</a>
    </nav>
    <header>
        <h1><div>Analyze your ad.</div><div class="highlight">Improve it instantly.</div></h1>
        <p>Upload any ad — poster, banner, or social screenshot — and get an AI audit with scores, CLIP semantic analysis, face detection, and real-world engagement estimates.</p>
    </header>
    <div class="dashboard">
        <div class="form-panel">
            <form id="adForm" method="POST" enctype="multipart/form-data">
                <div class="type-selector">
                    <label class="radio-pill"><input type="radio" name="ad_type" value="1" {% if request.method=='GET' or ad_type==1 %}checked{% endif %}><span class="pill-content">📱 Poster</span></label>
                    <label class="radio-pill"><input type="radio" name="ad_type" value="3" {% if request.method=='POST' and ad_type==3 %}checked{% endif %}><span class="pill-content">📐 Banner</span></label>
                    <label class="radio-pill"><input type="radio" name="ad_type" value="2" {% if request.method=='POST' and ad_type==2 %}checked{% endif %}><span class="pill-content">📱 Social Ad</span></label>
                    <label class="radio-pill"><input type="radio" name="ad_type" value="4" {% if request.method=='POST' and ad_type==4 %}checked{% endif %}><span class="pill-content">✉️ Email Ad</span></label>
                </div>
                <div class="copy-group">
                    <label>✍️ Ad Headline / Copy (optional)</label>
                    <textarea name="ad_copy" placeholder="e.g. Get 50% off today only — limited time offer!">{{ request.form.get('ad_copy','') }}</textarea>
                </div>
                <div class="upload-area" id="dropZone">
                    <input type="file" id="imageInput" name="ad_image" accept="image/png,image/jpeg,image/webp" onchange="previewImage(event)">
                    <div class="upload-content-wrapper" id="uploadPlaceholder">
                        <div class="upload-icon">⬆️</div>
                        <div class="upload-title">Drop your ad here</div>
                        <div class="upload-subtitle">PNG, JPG, WEBP — up to 10MB</div>
                    </div>
                    <img id="imagePreview" src="" alt="Ad Preview">
                </div>
                <div class="advanced-settings-toggle" onclick="toggleSettings()">
                    <span>⚙️ Advanced Settings</span> <span id="toggleIcon">▼</span>
                </div>
                <div class="settings-grid" id="advancedSettings" style="display:none;">
                    <div class="form-group">
                        <label>Color Style</label>
                        <select name="color">
                            <option value="3" {% if request.form.get('color')=='3' %}selected{% endif %}>Dark & Moody</option>
                            <option value="6" {% if not request.form.get('color') or request.form.get('color')=='6' %}selected{% endif %}>Balanced/Neutral</option>
                            <option value="9" {% if request.form.get('color')=='9' %}selected{% endif %}>Bright & Vibrant</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Target Audience</label>
                        <select name="audience">
                            <option value="1" {% if request.form.get('audience')=='1' %}selected{% endif %}>Students / Gen Z</option>
                            <option value="2" {% if request.form.get('audience')=='2' %}selected{% endif %}>Professionals (B2B)</option>
                            <option value="3" {% if not request.form.get('audience') or request.form.get('audience')=='3' %}selected{% endif %}>General / Broad</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Platform</label>
                        <select name="platform">
                            <option value="1" {% if not request.form.get('platform') or request.form.get('platform')=='1' %}selected{% endif %}>Instagram</option>
                            <option value="2" {% if request.form.get('platform')=='2' %}selected{% endif %}>YouTube</option>
                            <option value="3" {% if request.form.get('platform')=='3' %}selected{% endif %}>Facebook</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Day of Week</label>
                        <select name="day">
                            <option value="1" {% if request.form.get('day')=='1' %}selected{% endif %}>Monday</option>
                            <option value="2" {% if request.form.get('day')=='2' %}selected{% endif %}>Tuesday</option>
                            <option value="3" {% if not request.form.get('day') or request.form.get('day')=='3' %}selected{% endif %}>Wednesday</option>
                            <option value="4" {% if request.form.get('day')=='4' %}selected{% endif %}>Thursday</option>
                            <option value="5" {% if request.form.get('day')=='5' %}selected{% endif %}>Friday</option>
                            <option value="6" {% if request.form.get('day')=='6' %}selected{% endif %}>Saturday</option>
                            <option value="7" {% if request.form.get('day')=='7' %}selected{% endif %}>Sunday</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Posting Hour (0-23)</label>
                        <input type="number" name="time" min="0" max="23" value="{{ request.form.get('time','18') }}" required>
                    </div>
                </div>
                <button type="submit" class="btn-submit"><span>🔬</span> Analyze My Ad</button>
                <details style="margin-top:1.5rem;">
                    <summary style="cursor:pointer;color:var(--text-muted);font-size:0.9rem;font-weight:600;padding:0.5rem 0;list-style:none;display:flex;align-items:center;gap:0.5rem;">⚡ Compare against a competitor ad (optional)</summary>
                    <div style="margin-top:1rem;padding:1rem;background:rgba(99,102,241,0.06);border:1px dashed rgba(99,102,241,0.3);border-radius:12px;">
                        <label style="display:block;margin-bottom:0.5rem;font-size:0.82rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-muted);font-weight:600;">Competitor Ad Image</label>
                        <input type="file" name="competitor_image" accept="image/png,image/jpeg,image/webp" style="font-size:0.85rem;color:var(--text-muted);cursor:pointer;">
                    </div>
                </details>
            </form>
        </div>

        <script>
            function previewImage(event) {
                const reader = new FileReader();
                reader.onload = function(){
                    const output = document.getElementById('imagePreview');
                    const placeholder = document.getElementById('uploadPlaceholder');
                    const dropZone = document.getElementById('dropZone');
                    output.src = reader.result;
                    output.style.display = 'block';
                    placeholder.style.display = 'none';
                    dropZone.style.padding = '1rem';
                };
                if(event.target.files[0]) { reader.readAsDataURL(event.target.files[0]); }
            }
            function toggleSettings() {
                const s = document.getElementById('advancedSettings');
                const i = document.getElementById('toggleIcon');
                if(s.style.display==='none'){s.style.display='grid';i.innerHTML='▲';}
                else{s.style.display='none';i.innerHTML='▼';}
            }
        </script>

        {% if score is defined %}
        <div class="results-wrapper" id="results">
            <h2 style="text-align:center;margin-bottom:2.5rem;font-size:2.5rem;font-weight:800;">Analysis Results</h2>
            <div class="results-panel">

                <!-- Main score -->
                <div class="score-container" style="--score-degrees:{{ score * 3.6 }}deg;--score-color:{{ color_hex }};--score-glow:{{ color_glow }};">
                    <h3 style="margin-bottom:2rem;font-size:1.1rem;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.1em;font-weight:600;">Final Score</h3>
                    <div class="score-circle"><span class="score-value">{{ score }}</span></div>
                    <div class="feedback-badge" style="color:{{ color_hex }};border-color:{{ color_glow }};">{{ feedback }}</div>
                    {% if ctr_estimate %}
                    <div class="ctr-badge">📈 Estimated CTR: ~{{ ctr_estimate }}%</div>
                    {% endif %}
                </div>

                <!-- Score breakdown -->
                <div class="metrics-grid">
                    <div class="metric-card" style="text-align:center;">
                        <span style="display:block;font-size:1rem;color:var(--text-muted);margin-bottom:1rem;text-transform:uppercase;letter-spacing:0.05em;font-weight:600;">🤖 {{ regressor_name }} Score</span>
                        <span style="font-size:3.5rem;font-weight:800;color:#d4d4d8;">{{ ml_score }}</span>
                        <div style="font-size:0.78rem;color:var(--text-muted);margin-top:0.5rem;">{{ train_size }} training samples</div>
                        <div style="font-size:0.75rem;color:var(--text-muted);margin-top:0.2rem;">MAE {{ model_mae }} · R² {{ model_r2 }}</div>
                    </div>
                    {% if vlm_score %}
                    <div class="metric-card" style="text-align:center;">
                        <span style="display:block;font-size:1rem;color:var(--text-muted);margin-bottom:1rem;text-transform:uppercase;letter-spacing:0.05em;font-weight:600;">👁️ Visual Impact Score</span>
                        <span style="font-size:3.5rem;font-weight:800;color:var(--accent-primary);">{{ vlm_score }}</span>
                        <div style="font-size:0.78rem;color:var(--text-muted);margin-top:0.5rem;">BLIP{% if clip_available %} + CLIP{% endif %}</div>
                    </div>
                    {% else %}
                    <div class="metric-card" style="text-align:center;opacity:0.4;">
                        <span style="display:block;font-size:1rem;color:var(--text-muted);margin-bottom:1rem;text-transform:uppercase;letter-spacing:0.05em;font-weight:600;">👁️ Visual Impact Score</span>
                        <span style="font-size:1.5rem;font-weight:600;color:var(--text-muted);">No image</span>
                        <div style="font-size:0.78rem;color:var(--text-muted);margin-top:0.5rem;">Upload an ad image to enable</div>
                    </div>
                    {% endif %}
                </div>

                <!-- PHASE 3: Computer Vision Details -->
                {% if vlm_results %}
                <div class="metric-card">
                    <h3 style="color:#fff;font-size:1.4rem;font-weight:700;margin-bottom:1rem;">🔬 Computer Vision Analysis</h3>
                    <div class="cv-grid">
                        <div class="cv-badge">
                            <div class="cv-val">{{ vlm_results.face_count }}</div>
                            <div class="cv-lbl">Faces Detected</div>
                        </div>
                        <div class="cv-badge">
                            <div class="cv-val">{{ vlm_results.text_area_pct }}%</div>
                            <div class="cv-lbl">Text Area %</div>
                        </div>
                        <div class="cv-badge">
                            <div class="cv-val">{{ vlm_results.brightness|int }}</div>
                            <div class="cv-lbl">Brightness</div>
                        </div>
                        <div class="cv-badge">
                            <div class="cv-val">{{ vlm_results.contrast|int }}</div>
                            <div class="cv-lbl">Contrast</div>
                        </div>
                    </div>
                    {% if vlm_results.clip_scores %}
                    <h4 style="color:var(--text-muted);font-size:0.85rem;text-transform:uppercase;letter-spacing:0.05em;margin-top:1.5rem;margin-bottom:0.75rem;">🧠 CLIP Semantic Scores</h4>
                    {% for concept, val in vlm_results.clip_scores.items() %}
                    <div class="clip-bar-wrap">
                        <div class="clip-bar-label"><span style="text-transform:capitalize;">{{ concept }}</span><span>{{ "%.0f"|format(val*100) }}%</span></div>
                        <div class="clip-bar-bg"><div class="clip-bar-fill" style="width:{{ "%.0f"|format(val*100) }}%;{% if concept in ['crowded'] %}background:#f87171;{% endif %}"></div></div>
                    </div>
                    {% endfor %}
                    {% endif %}
                </div>
                {% endif %}

                <!-- Copy analysis -->
                {% if copy_result %}
                <div class="metric-card">
                    <h3 style="color:#fff;font-size:1.4rem;font-weight:700;margin-bottom:1rem;display:flex;align-items:center;gap:0.75rem;">✍️ Ad Copy Analysis</h3>
                    <div style="display:flex;flex-wrap:wrap;gap:0.6rem;margin-bottom:1rem;">
                        <span class="copy-pill {% if copy_result.copy_score >= 70 %}good{% elif copy_result.copy_score >= 45 %}warn{% else %}bad{% endif %}">Copy Score: {{ copy_result.copy_score }}/100</span>
                        <span class="copy-pill {% if copy_result.sentiment_label == 'Positive' %}good{% elif copy_result.sentiment_label == 'Negative' %}bad{% else %}warn{% endif %}">{{ copy_result.sentiment_label }} Tone</span>
                        {% if copy_result.has_cta %}<span class="copy-pill good">✓ CTA Detected</span>{% else %}<span class="copy-pill bad">✗ No CTA</span>{% endif %}
                        {% if copy_result.has_urgency %}<span class="copy-pill good">⚡ Urgency</span>{% endif %}
                        {% if copy_result.has_question %}<span class="copy-pill warn">? Engagement Hook</span>{% endif %}
                        {% if copy_result.is_spam_like %}<span class="copy-pill bad">🚨 Spam-like Copy</span>{% endif %}
                        <span class="copy-pill">{{ copy_result.word_count }} words</span>
                        <span class="copy-pill">Readability {{ copy_result.readability }}/10</span>
                    </div>
                    {% if copy_result.is_spam_like %}
                    <div style="background:rgba(248,113,113,0.1);border:1px solid rgba(248,113,113,0.3);border-radius:10px;padding:0.75rem 1rem;font-size:0.88rem;color:#f87171;margin-top:0.5rem;">
                        ⚠️ Keyword stuffing detected ({{ (copy_result.spam_ratio * 100)|int }}% power-word density). Rewrite with a specific value proposition instead.
                    </div>
                    {% endif %}
                </div>
                {% endif %}

                <!-- OCR -->
                {% if vlm_results and vlm_results.ocr_text %}
                <div class="metric-card">
                    <h3 style="color:#fff;font-size:1.4rem;font-weight:700;margin-bottom:1rem;">🔤 OCR — Text in Image</h3>
                    <div style="background:rgba(0,0,0,0.4);border:1px solid rgba(255,255,255,0.1);border-radius:10px;padding:0.75rem 1rem;font-size:0.9rem;color:#d4d4d8;font-family:monospace;white-space:pre-wrap;">{{ vlm_results.ocr_text }}</div>
                </div>
                {% endif %}

                <!-- Suggestions & Insights -->
                <div class="metrics-grid" style="grid-template-columns:1fr;">
                    <div class="metric-card">
                        <h3 style="color:#fff;font-size:1.4rem;font-weight:700;margin-bottom:1.5rem;">💡 Optimization Suggestions</h3>
                        <ul class="list-items">
                            {% for s in suggestions %}<li><span class="icon-warn">⚡</span> <span>{{ s }}</span></li>{% endfor %}
                            {% if not suggestions %}<li><span class="icon-good">✓</span> <span>Your ad looks great! No major optimizations needed.</span></li>{% endif %}
                        </ul>
                    </div>
                    <div class="metric-card">
                        <h3 style="color:#fff;font-size:1.4rem;font-weight:700;margin-bottom:1.5rem;">🔍 Explainable Insights</h3>
                        <ul class="list-items">
                            {% for i in insights %}
                            <li>
                                {% if "(+" in i %}<span class="icon-good">📈</span>
                                {% elif "(-" in i %}<span class="icon-bad">📉</span>
                                {% else %}<span>👁️</span>{% endif %}
                                <span>{{ i }}</span>
                            </li>
                            {% endfor %}
                        </ul>
                    </div>
                </div>

                <!-- AI Rewrite -->
                {% if rewrite %}
                <div class="metric-card rewrite-card" style="margin-top:0;">
                    <h3 style="color:#a5b4fc;font-size:1.1rem;font-weight:700;">✨ AI-Rewritten Copy</h3>
                    <p class="rewrite-note" style="margin-top:0.4rem;">Claude rewrote your copy based on the issues found:</p>
                    <div class="rewrite-copy">"{{ rewrite }}"</div>
                    <p class="rewrite-note">Targeting <b>{{ platform_label }}</b> · <b>{{ audience_label }}</b></p>
                </div>
                {% elif ad_copy and not anthropic_key_set %}
                <div class="metric-card" style="opacity:0.5;margin-top:0;">
                    <p style="font-size:0.9rem;color:var(--text-muted);">✨ <b>AI Copy Rewrite</b> — Set <code>ANTHROPIC_API_KEY</code> to enable Claude rewrites.</p>
                </div>
                {% endif %}

                <!-- PHASE 2: Feedback widget -->
                <div class="metric-card" style="padding:1.5rem 2rem;">
                    <h3 style="color:#fff;font-size:1.1rem;font-weight:700;margin-bottom:1rem;">💬 Was this prediction accurate?</h3>
                    <div class="feedback-widget" id="feedbackWidget">
                        <span>Help AdLens learn from real-world results:</span>
                        <button class="thumb-btn" id="thumbUp"   onclick="sendFeedback({{ analysis_id }}, 1)">👍 Yes, accurate</button>
                        <button class="thumb-btn" id="thumbDown" onclick="sendFeedback({{ analysis_id }}, 0)">👎 No, off-target</button>
                        <span class="feedback-thanks" id="feedbackThanks">✅ Feedback saved — thank you!</span>
                    </div>
                    <p style="font-size:0.78rem;color:var(--text-muted);margin-top:0.75rem;">
                        Your feedback trains the model. Stored locally in SQLite.
                        {% if feedback_stats.total > 0 %}
                        Current accuracy from {{ feedback_stats.total }} ratings: <b style="color:var(--accent-primary);">{{ feedback_stats.accuracy_rate }}%</b>
                        {% endif %}
                    </p>
                </div>

                <!-- Competitor -->
                {% if competitor %}
                <div class="metric-card" style="margin-top:0;">
                    <h3 style="color:#fff;font-size:1.4rem;font-weight:700;margin-bottom:1.5rem;">⚔️ Competitor Benchmarking</h3>
                    <div class="compare-grid">
                        <div class="compare-col">
                            <h4>🟢 Your Ad</h4>
                            {% for metric, val in your_metrics.items() %}
                            <div class="compare-bar-wrap">
                                <div class="compare-bar-label"><span>{{ metric }}</span><span>{{ val }}/10</span></div>
                                <div class="compare-bar-bg"><div class="compare-bar-fill bar-you" style="width:{{ val * 10 }}%"></div></div>
                            </div>
                            {% endfor %}
                            <div style="margin-top:1rem;font-size:1.5rem;font-weight:800;color:var(--accent-primary);">{{ vlm_score or '—' }}</div>
                            <div style="font-size:0.8rem;color:var(--text-muted);">Visual Score</div>
                        </div>
                        <div class="compare-col">
                            <h4>🟣 Competitor</h4>
                            {% for metric, val in competitor_metrics.items() %}
                            <div class="compare-bar-wrap">
                                <div class="compare-bar-label"><span>{{ metric }}</span><span>{{ val }}/10</span></div>
                                <div class="compare-bar-bg"><div class="compare-bar-fill bar-them" style="width:{{ val * 10 }}%"></div></div>
                            </div>
                            {% endfor %}
                            <div style="margin-top:1rem;font-size:1.5rem;font-weight:800;color:#6366f1;">{{ competitor.vlm_score }}</div>
                            <div style="font-size:0.8rem;color:var(--text-muted);">Visual Score</div>
                        </div>
                    </div>
                    <div style="margin-top:1.25rem;padding:1rem;background:rgba(0,0,0,0.3);border-radius:10px;font-size:0.9rem;color:#d4d4d8;">
                        {% if (vlm_score or 0) > competitor.vlm_score %}🏆 <b>Your ad scores higher</b> by {{ (vlm_score or 0) - competitor.vlm_score }} pts.
                        {% elif (vlm_score or 0) < competitor.vlm_score %}📉 <b>Competitor scores higher</b> by {{ competitor.vlm_score - (vlm_score or 0) }} pts.
                        {% else %}⚖️ <b>Visual scores are tied.</b> Differentiate through copy and targeting.{% endif %}
                        {% if competitor.caption %}<br><br>BLIP reads competitor as: <i>"{{ competitor.caption }}"</i>{% endif %}
                        {% if competitor.face_count is defined %}
                        <br>Competitor face count: <b>{{ competitor.face_count }}</b>
                        {% endif %}
                    </div>
                </div>
                {% endif %}

            </div>
        </div>

        <script>
        function sendFeedback(analysisId, rating) {
            fetch('/feedback', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({analysis_id: analysisId, rating: rating})
            }).then(r => r.json()).then(data => {
                document.getElementById('thumbUp').classList.toggle('selected-up', rating === 1);
                document.getElementById('thumbDown').classList.toggle('selected-down', rating === 0);
                document.getElementById('feedbackThanks').style.display = 'inline';
            });
        }
        window.onload = function() {
            setTimeout(() => { document.getElementById('results').scrollIntoView({ behavior: 'smooth', block: 'start' }); }, 100);
        }
        </script>
        {% endif %}
    </div>
</div>
</body>
</html>
"""

HISTORY_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8"><title>AdLens — History</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        :root { --bg-primary:#0a0a0c; --accent-primary:#ccff00; --text-main:#ffffff; --text-muted:#a1a1aa; --glass-border:rgba(255,255,255,0.08); }
        * { box-sizing:border-box; margin:0; padding:0; }
        body { background-color:var(--bg-primary); background-image:linear-gradient(rgba(255,255,255,0.03) 1px,transparent 1px),linear-gradient(90deg,rgba(255,255,255,0.03) 1px,transparent 1px); background-size:50px 50px; color:var(--text-main); font-family:'Outfit',sans-serif; min-height:100vh; padding:0 2rem 3rem; }
        .container { max-width:1100px; margin:0 auto; }
        nav { display:flex; justify-content:space-between; align-items:center; padding:2rem 0; margin-bottom:2rem; }
        .logo { font-size:1.4rem; font-weight:800; }
        h1 { font-size:2.5rem; font-weight:800; margin-bottom:0.5rem; }
        .sub { color:var(--text-muted); margin-bottom:2rem; }
        .card { background:rgba(255,255,255,0.02); border:1px solid var(--glass-border); border-radius:20px; padding:2rem; overflow-x:auto; }
        table { width:100%; border-collapse:collapse; }
        th { text-align:left; padding:0.75rem 1rem; font-size:0.78rem; text-transform:uppercase; letter-spacing:0.06em; color:var(--text-muted); border-bottom:1px solid rgba(255,255,255,0.06); }
        td { padding:0.9rem 1rem; font-size:0.9rem; color:#d4d4d8; border-bottom:1px solid rgba(255,255,255,0.04); vertical-align:top; max-width:180px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
        tr:hover td { background:rgba(255,255,255,0.02); }
        .chip { display:inline-block; padding:0.25rem 0.7rem; border-radius:20px; font-weight:700; font-size:0.82rem; }
        .high { background:rgba(204,255,0,0.15); color:#ccff00; }
        .mid  { background:rgba(250,204,21,0.15); color:#facc15; }
        .low  { background:rgba(248,113,113,0.15); color:#f87171; }
        .thumb-yes { color:#ccff00; } .thumb-no { color:#f87171; } .thumb-none { color:#a1a1aa; }
        .btn-back { display:inline-block; padding:0.5rem 1.25rem; border:1px solid rgba(255,255,255,0.12); border-radius:20px; color:var(--text-muted); text-decoration:none; font-size:0.85rem; font-weight:600; }
        .btn-back:hover { color:#fff; }
        .empty { text-align:center; padding:4rem; color:var(--text-muted); }
        .stats-bar { display:flex; gap:2rem; margin-bottom:2rem; }
        .stat-box { background:rgba(255,255,255,0.02); border:1px solid var(--glass-border); border-radius:12px; padding:1rem 1.5rem; }
        .stat-val { font-size:1.8rem; font-weight:800; color:var(--accent-primary); }
        .stat-lbl { font-size:0.78rem; color:var(--text-muted); text-transform:uppercase; letter-spacing:0.05em; margin-top:0.2rem; }
    </style>
</head>
<body>
<div class="container">
    <nav>
        <div class="logo">Ad<span style="color:var(--accent-primary);">Lens</span></div>
        <a href="/" class="btn-back">← Back to Analyzer</a>
    </nav>
    <h1>📋 Analysis History</h1>
    <p class="sub">Last {{ rows|length }} analyses with real feedback signals.</p>
    {% if feedback_stats.total > 0 %}
    <div class="stats-bar">
        <div class="stat-box"><div class="stat-val">{{ feedback_stats.total }}</div><div class="stat-lbl">User Ratings</div></div>
        <div class="stat-box"><div class="stat-val">{{ feedback_stats.accuracy_rate }}%</div><div class="stat-lbl">Prediction Accuracy</div></div>
    </div>
    {% endif %}
    <div class="card">
    {% if rows %}
    <table>
        <thead>
            <tr><th>#</th><th>Date</th><th>Ad Copy</th><th>Score</th><th>ML</th><th>Visual</th><th>Copy</th><th>CTR Est.</th><th>Faces</th><th>Platform</th><th>Rating</th></tr>
        </thead>
        <tbody>
        {% for r in rows %}
        <tr>
            <td style="color:var(--text-muted);">{{ r.id }}</td>
            <td style="white-space:nowrap;color:var(--text-muted);">{{ r.ts }}</td>
            <td title="{{ r.ad_copy }}">{{ (r.ad_copy or '—')[:35] }}{% if r.ad_copy and r.ad_copy|length > 35 %}…{% endif %}</td>
            <td><span class="chip {% if r.score >= 80 %}high{% elif r.score >= 60 %}mid{% else %}low{% endif %}">{{ r.score }}</span></td>
            <td>{{ r.ml_score or '—' }}</td>
            <td>{{ r.vlm_score or '—' }}</td>
            <td>{{ r.copy_score or '—' }}</td>
            <td>{% if r.ctr_estimate %}{{ r.ctr_estimate }}%{% else %}—{% endif %}</td>
            <td>{{ r.face_count if r.face_count is not none else '—' }}</td>
            <td>{{ {1:'Instagram',2:'YouTube',3:'Facebook'}.get(r.platform,'?') }}</td>
            <td>
                {% if r.user_rating == 1 %}<span class="thumb-yes">👍</span>
                {% elif r.user_rating == 0 %}<span class="thumb-no">👎</span>
                {% else %}<span class="thumb-none">—</span>{% endif %}
            </td>
        </tr>
        {% endfor %}
        </tbody>
    </table>
    {% else %}
    <div class="empty">No analyses yet. Run your first ad through the analyzer!</div>
    {% endif %}
    </div>
</div>
</body>
</html>
"""

# ─────────────────────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/history")
def history():
    rows = load_history(50)
    feedback_stats = get_feedback_stats()
    return render_template_string(HISTORY_HTML, rows=rows, feedback_stats=feedback_stats)

# PHASE 2: Feedback endpoint ──────────────────────────────────────────────────
@app.route("/feedback", methods=["POST"])
def feedback():
    data        = request.get_json()
    analysis_id = data.get("analysis_id")
    rating      = data.get("rating")   # 1 or 0
    comment     = data.get("comment", "")
    if analysis_id is None or rating is None:
        return jsonify({"error": "missing fields"}), 400
    save_feedback(analysis_id, rating, comment)
    stats = get_feedback_stats()
    return jsonify({"ok": True, "feedback_stats": stats})

@app.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        ad_type  = int(request.form.get("ad_type", 1))
        color    = int(request.form.get("color", 6))
        audience = int(request.form.get("audience", 3))
        platform = int(request.form.get("platform", 1))
        day      = int(request.form.get("day", 3))
        time_val = int(request.form.get("time", 18))
        ad_copy  = request.form.get("ad_copy", "").strip()

        suggestions, insights = [], []
        vlm_score, vlm_results = None, None
        ctr_estimate = None
        face_count_val = 0
        brightness_val = 128.0
        clip_l = clip_t = clip_e = None

        headline_length = len(ad_copy.split()) if ad_copy else 8

        # Copy analysis
        copy_result = analyze_ad_copy(ad_copy) if ad_copy else None
        if copy_result:
            if not copy_result['has_cta']:
                suggestions.append("Add a clear CTA (e.g. 'Shop Now', 'Get Started') — ads with CTAs get 2× higher CTR.")
            if not copy_result['has_urgency']:
                suggestions.append("Add urgency words ('today only', 'limited offer') to drive faster conversions.")
            if copy_result['sentiment_label'] == 'Negative':
                suggestions.append("Your copy has negative language. Reframe around benefits and positive outcomes.")
            if copy_result['word_count'] > 25:
                suggestions.append("Headline too long. Cut to under 10 words for mobile readability.")
            if copy_result['word_count'] < 3:
                suggestions.append("Copy too short — add a value proposition.")
            if copy_result['readability'] < 5:
                suggestions.append("Simplify language — use shorter, common words for broader reach.")
            if copy_result['copy_score'] >= 70:
                insights.append(f"(+) Copy score {copy_result['copy_score']}/100 — strong persuasive signals.")
            elif copy_result['copy_score'] >= 45:
                insights.append(f"👁️ Copy score {copy_result['copy_score']}/100 — decent but missing key elements.")
            else:
                insights.append(f"(-) Copy score {copy_result['copy_score']}/100 — weak structure. Add CTA + positive framing.")

        # Image analysis (BLIP + PHASE 3 CV)
        if 'ad_image' in request.files and request.files['ad_image'].filename != '':
            try:
                vlm_results = analyze_image(request.files['ad_image'])
                vlm_score   = int(np.mean([
                    vlm_results["color_score"], vlm_results["visual_clarity_score"],
                    vlm_results["text_density_score"], vlm_results["emotional_impact_score"]
                ]) * 10)

                face_count_val = vlm_results.get("face_count", 0)
                brightness_val = vlm_results.get("brightness", 128)
                clip_scores    = vlm_results.get("clip_scores", {})
                clip_l = clip_scores.get("luxury")
                clip_t = clip_scores.get("trust")
                clip_e = clip_scores.get("energy")

                insights.append(f"BLIP: '{vlm_results['caption'].capitalize()}'")
                insights.append(f"Visual Clarity: {vlm_results['visual_clarity_score']}/10 | Emotion: {vlm_results['emotional_impact_score']}/10")

                # PHASE 3 insights
                if face_count_val > 0:
                    insights.append(f"(+) {face_count_val} face(s) detected — human presence boosts CTR by ~30%.")
                else:
                    suggestions.append("No faces detected. Adding a human face or expression can boost engagement.")

                if vlm_results['text_area_pct'] > 35:
                    suggestions.append(f"Text covers ~{vlm_results['text_area_pct']}% of the image. Keep it under 20% to avoid platform flags.")
                    insights.append(f"(-) High text area ({vlm_results['text_area_pct']}%) — may get restricted on Facebook/Instagram.")

                if brightness_val < 80:
                    suggestions.append("Image is quite dark. Brighter ads perform 20% better in feed environments.")
                elif brightness_val > 220:
                    suggestions.append("Image may be overexposed. High brightness reduces contrast and readability.")
                else:
                    insights.append(f"(+) Brightness {int(brightness_val)} — good exposure range for feed ads.")

                # CLIP insights
                if clip_scores:
                    if clip_scores.get("luxury", 0) > 0.5:
                        insights.append(f"(+) CLIP detects strong luxury/premium visual cues ({int(clip_scores['luxury']*100)}%).")
                    if clip_scores.get("trust", 0) > 0.5:
                        insights.append(f"(+) CLIP detects trustworthy brand signals ({int(clip_scores['trust']*100)}%).")
                    if clip_scores.get("crowded", 0) > 0.5:
                        suggestions.append(f"CLIP flags cluttered/crowded layout ({int(clip_scores['crowded']*100)}%). Simplify for better recall.")
                        insights.append(f"(-) CLIP: layout appears crowded — reduces visual clarity score.")
                    if clip_scores.get("energy", 0) > 0.5:
                        insights.append(f"(+) CLIP detects high energy/excitement ({int(clip_scores['energy']*100)}%) — great for CTR.")

                if vlm_results['visual_clarity_score'] < 5:
                    suggestions.append("Low contrast image. Increase contrast so key elements pop.")
                if vlm_results['emotional_impact_score'] < 5:
                    suggestions.append("Image feels flat. Try adding faces, warm colors, or dynamic composition.")
                if vlm_results['text_density_score'] >= 8:
                    suggestions.append("Heavy text in image. Platforms may flag ads with >20% text coverage.")

            except Exception as e:
                print(f"[AdLens] VLM error: {e}")
                insights.append("Failed to process image.")

        # PHASE 1: Estimate CTR
        if vlm_score is not None:
            ctr_estimate = estimate_ctr(
                score=vlm_score, face_count=face_count_val,
                clip_scores=vlm_results.get("clip_scores", {}),
                has_cta=copy_result['has_cta'] if copy_result else False,
                brightness=brightness_val
            )

        # Engineered features
        text_d = vlm_results["text_density_score"]    if vlm_results else 5
        emo    = vlm_results["emotional_impact_score"] if vlm_results else 6
        apm, phf, vsb = get_engineered_features(ad_type, audience, platform, time_val, text_d, emo)

        # PHASE 4: XGBoost / LightGBM / GBR prediction with extended features
        ctr_raw = ctr_estimate / 100 if ctr_estimate else 0.03
        features = np.array([[
            headline_length, ad_type, color, audience, platform,
            time_val, day, apm, phf, vsb,
            ctr_raw, brightness_val, emo
        ]])
        features_scaled = scaler.transform(features)
        ml_score = int(np.clip(model.predict(features_scaled)[0], 5, 100))

        # SHAP
        shap_insights = get_shap_insights(features_scaled, features[0])
        insights.extend(shap_insights)

        # OCR insights
        if vlm_results and vlm_results.get("ocr_text"):
            wc = len(vlm_results["ocr_text"].split())
            if wc > 20:
                suggestions.append(f"OCR found {wc} words in image — text overload, risk of platform restriction.")
                insights.append(f"(-) OCR: {wc} words in image — exceeds recommended coverage.")
            elif wc > 0:
                insights.append(f"(+) OCR: {wc} words detected in image — within safe range.")

        # Spam
        if copy_result and copy_result.get("is_spam_like"):
            suggestions.append("Copy reads like keyword-stuffed hype. Rewrite with a specific value proposition.")
            insights.append("(-) Spam signal: high power-word density penalizes persuasion score.")

        # Composite score
        if vlm_score is not None:
            score = int(np.clip(0.6 * ml_score + 0.4 * vlm_score, 0, 100))
        else:
            score = ml_score
        if copy_result:
            copy_delta = (copy_result['copy_score'] - 50) * 0.1
            score = int(np.clip(score + copy_delta, 0, 100))

        # Audience/platform rules (same as before)
        if apm == 1:
            insights.append("(+) Audience–platform pairing matches high-synergy combos.")
        else:
            suggestions.append("Reconsider platform — your audience performs better elsewhere.")
            insights.append("(-) Audience–platform mismatch: lower engagement predicted.")
        if phf == 1:
            insights.append("(+) Posting during peak hours (18:00–21:00) — highest CTR window.")
        else:
            suggestions.append(f"Reschedule to 18:00–21:00 for better reach. Hour {time_val}:00 is off-peak.")
            insights.append(f"(-) Off-peak posting at {time_val}:00 — lower CTR expected.")
        if vsb > 3:
            insights.append(f"(+) Visual sentiment boost: {vsb} — strong emotional signal.")
        elif vsb < 0:
            suggestions.append("Reduce text density — text overload suppresses emotional recall.")
            insights.append(f"(-) Visual sentiment boost: {vsb} — text outweighs emotional impact.")
        if platform == 1 and ad_type != 2:
            suggestions.append("Consider turning this into a Reel for Instagram.")
            insights.append("(+) Video outperforms static on Instagram (+15% potential).")
        if audience == 2:
            if platform == 1:
                suggestions.append("Run on LinkedIn/Facebook instead for B2B targeting.")
                insights.append("(-) Instagram has lower B2B engagement (-10%).")
                score = int(np.clip(score - 10, 0, 100))
            if day in [6, 7]:
                suggestions.append("B2B audiences disengage on weekends. Move to Tue–Thu.")
                insights.append("(-) Weekend B2B posting drops reach by ~15%.")
                score = int(np.clip(score - 10, 0, 100))
        elif audience in [1, 3]:
            if day in [6, 7]:
                insights.append("(+) Weekends boost engagement for General/Student audiences (+10%).")
                score = int(np.clip(score + 8, 0, 100))
        if time_val < 7 or (13 < time_val < 17):
            suggestions.append("Reschedule to evening peak times (18:00–21:00).")
            insights.append(f"(-) Posting at {time_val}:00 reduces algorithmic reach (-8%).")
            score = int(np.clip(score - 5, 0, 100))

        score = min(100, max(0, score))

        if score >= 80:
            feedback_text = "🔥 Excellent Ad Potential"
            color_hex, color_glow = "#ccff00", "rgba(204,255,0,0.5)"
        elif score >= 60:
            feedback_text = "👍 Good, But Could Improve"
            color_hex, color_glow = "#facc15", "rgba(250,204,21,0.5)"
        else:
            feedback_text = "⚠️ Needs Optimization"
            color_hex, color_glow = "#f87171", "rgba(248,113,113,0.5)"

        # Competitor
        competitor = None
        your_metrics = competitor_metrics = {}
        if 'competitor_image' in request.files and request.files['competitor_image'].filename != '':
            try:
                competitor = score_image_for_compare(request.files['competitor_image'])
                your_metrics = {
                    "Color Score":  vlm_results["color_score"]           if vlm_results else 5,
                    "Clarity":      vlm_results["visual_clarity_score"]  if vlm_results else 5,
                    "Text Density": vlm_results["text_density_score"]    if vlm_results else 5,
                    "Emotion":      vlm_results["emotional_impact_score"] if vlm_results else 5,
                }
                competitor_metrics = {
                    "Color Score":  competitor["color_score"],
                    "Clarity":      competitor["clarity_score"],
                    "Text Density": competitor["text_density_score"],
                    "Emotion":      competitor["emotional_score"],
                }
            except Exception as e:
                print(f"[AdLens] Competitor error: {e}")

        # AI rewrite
        rewrite = None
        copy_issues = [s for s in suggestions if any(
            kw in s.lower() for kw in ['cta','urgency','negative','long','short','simplify','spam','hype']
        )]
        if ad_copy:
            rewrite = ai_rewrite_copy(ad_copy, platform, audience, copy_issues)

        # Save — PHASE 1: extended columns
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        analysis_id = save_analysis(
            ts=ts, ad_copy=ad_copy, ad_type=ad_type, audience=audience,
            platform=platform, score=score, ml_score=ml_score,
            vlm_score=vlm_score,
            copy_score=copy_result['copy_score'] if copy_result else None,
            feedback_text=feedback_text, insights=insights, suggestions=suggestions,
            ctr_estimate=ctr_estimate,
            face_count=face_count_val if vlm_results else None,
            text_area_pct=vlm_results.get("text_area_pct") if vlm_results else None,
            brightness=brightness_val if vlm_results else None,
            clip_luxury=clip_l, clip_trust=clip_t, clip_energy=clip_e
        )

        feedback_stats = get_feedback_stats()

        return render_template_string(
            HTML,
            score=score, ml_score=ml_score, vlm_score=vlm_score,
            feedback=feedback_text, color_hex=color_hex, color_glow=color_glow,
            suggestions=suggestions, insights=insights,
            copy_result=copy_result, vlm_results=vlm_results,
            ad_type=ad_type, color=color, audience=audience,
            platform=platform, day=day,
            train_size=TRAIN_SIZE, model_mae=MODEL_MAE, model_r2=MODEL_R2,
            rewrite=rewrite, ad_copy=ad_copy,
            anthropic_key_set=bool(ANTHROPIC_API_KEY),
            platform_label=PLATFORM_LABELS.get(platform, "Social"),
            audience_label=AUDIENCE_LABELS.get(audience, "General"),
            competitor=competitor,
            your_metrics=your_metrics,
            competitor_metrics=competitor_metrics,
            # New vars
            analysis_id=analysis_id,
            ctr_estimate=ctr_estimate,
            regressor_name=REGRESSOR_NAME,
            clip_available=CLIP_AVAILABLE,
            feedback_stats=feedback_stats,
        )

    return render_template_string(
        HTML,
        regressor_name=REGRESSOR_NAME,
        clip_available=CLIP_AVAILABLE,
    )

if __name__ == "__main__":
    app.run(debug=True)
