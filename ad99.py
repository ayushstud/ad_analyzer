import re
import os
import json
import sqlite3
import datetime
import numpy as np
import openpyxl
import torch
from PIL import Image, ImageStat
from transformers import CLIPProcessor, CLIPModel
from flask import Flask, request, render_template_string
import xgboost as xgb
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, r2_score

# ──  SHAP for feature explainability ──────────────
try:
    import shap
    SHAP_AVAILABLE = True
except ImportError:
    SHAP_AVAILABLE = False
    print("[AdLens] shap not installed — run: pip install shap")


try:
    import easyocr
    _ocr_reader = easyocr.Reader(['en'], gpu=False)
    OCR_BACKEND = "easyocr"
except ImportError:
    try:
        import pytesseract
        OCR_BACKEND = "pytesseract"
    except ImportError:
        OCR_BACKEND = None
        print("[AdLens] No OCR backend found — run: pip install easyocr  (or pytesseract)")

app = Flask(__name__)

# ─────────────────────────────────────────────────────────
#
# ─────────────────────────────────────────────────────────
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")  # export ANTHROPIC_API_KEY=sk-ant-...

# ─────────────────────────────────────────────────────────
# SQLITE HISTORY DB  (auto-created next to adlens.py)
# ─────────────────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "adlens_history.db")

def init_db():
    con = sqlite3.connect(DB_PATH)
    con.execute("""
        CREATE TABLE IF NOT EXISTS analyses (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            ts          TEXT    NOT NULL,
            ad_copy     TEXT,
            ad_type     INTEGER,
            audience    INTEGER,
            platform    INTEGER,
            score       INTEGER,
            ml_score    INTEGER,
            vlm_score   INTEGER,
            copy_score  INTEGER,
            feedback    TEXT,
            insights    TEXT,
            suggestions TEXT
        )
    """)
    con.commit()
    con.close()

def save_analysis(ts, ad_copy, ad_type, audience, platform, score,
                  ml_score, vlm_score, copy_score, feedback, insights, suggestions):
    con = sqlite3.connect(DB_PATH)
    con.execute("""
        INSERT INTO analyses
            (ts, ad_copy, ad_type, audience, platform, score, ml_score,
             vlm_score, copy_score, feedback, insights, suggestions)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (ts, ad_copy, ad_type, audience, platform, score, ml_score,
          vlm_score, copy_score, feedback,
          json.dumps(insights), json.dumps(suggestions)))
    con.commit()
    con.close()

def load_history(limit=20):
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    rows = con.execute(
        "SELECT * FROM analyses ORDER BY id DESC LIMIT ?", (limit,)
    ).fetchall()
    con.close()
    return [dict(r) for r in rows]

init_db()

# ─────────────────────────────────────────────────────────
#  LOAD DATASET WITH PROPER TRAIN/TEST SPLIT
# ─────────────────────────────────────────────────────────
DATASET_PATH = "adlens_hybrid_dataset.xlsx"

def load_and_train():
    wb = openpyxl.load_workbook(DATASET_PATH)
    ws = wb["Hybrid Dataset"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    X, y = [], []
    for r in range(2, ws.max_row + 1):
        row = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        try:
            feats = [
                float(row["Headline Len"]                 or 8),   #  — real from dataset
                float(row["Ad Type"]                      or 1),
                float(row["Color Score"]                  or 5),
                float(row["Audience"]                     or 3),
                float(row["Platform"]                     or 1),
                float(row["Hour"]                         or 18),
                float(row["Day"]                          or 3),
                float(row["Audience\u00d7Platform Match"] or 0),
                float(row["Peak Hour Flag"]               or 0),
                float(row["Visual Sentiment Boost"]       or 0),
            ]
            target = float(row["Final Score"] or 50)
            X.append(feats)
            y.append(target)
        except (TypeError, ValueError):
            continue

    X, y = np.array(X), np.array(y)

    # FIX 2 — proper 80/20 train-test split
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    scaler = StandardScaler()
    X_train_s = scaler.fit_transform(X_train)
    X_test_s  = scaler.transform(X_test)

    clf = xgb.XGBRegressor(
        n_estimators=200,
        max_depth=4,
        learning_rate=0.08,
        subsample=0.8,
        colsample_bytree=0.8,
        random_state=42,
        tree_method="hist",
        verbosity=0,
    )
    clf.fit(X_train_s, y_train)

    preds = clf.predict(X_test_s)
    mae   = round(mean_absolute_error(y_test, preds), 2)
    r2    = round(r2_score(y_test, preds), 3)

    # ── SHAP explainer (TreeExplainer works natively with XGBoost) ────────
    explainer = None
    if SHAP_AVAILABLE:
        try:
            explainer = shap.TreeExplainer(clf)
        except Exception as e:
            print(f"[AdLens] SHAP init failed: {e}")

    print(f"Trained on {len(X_train)} samples | Tested on {len(X_test)} | MAE={mae} | R²={r2}")
    return clf, scaler, explainer, mae, r2, len(X_train), len(X_test)

print("Loading hybrid dataset...")
model, scaler, SHAP_EXPLAINER, MODEL_MAE, MODEL_R2, TRAIN_SIZE, TEST_SIZE = load_and_train()

# ─────────────────────────────────────────────────────────
# ENGINEERED FEATURE HELPERS
# ─────────────────────────────────────────────────────────
HIGH_SYNERGY = {(4,2),(1,1),(5,1),(1,4),(3,3),(2,4),(4,4)}

def get_engineered_features(ad_type, audience, platform, hour, text_density, emotional_impact):
    apm = 1 if (platform, audience) in HIGH_SYNERGY else 0
    phf = 1 if 18 <= hour <= 21 else 0
    vsb = round(emotional_impact * 1.2 - text_density * 0.5, 2)
    return apm, phf, vsb

# ─────────────────────────────────────────────────────────
# SHAP EXPLAINABILITY HELPER
# ─────────────────────────────────────────────────────────
FEATURE_NAMES = [
    "Headline Length", "Ad Type", "Color Score", "Audience", "Platform",
    "Hour", "Day", "Audience×Platform Match", "Peak Hour Flag", "Visual Sentiment Boost"
]

def get_shap_insights(features_scaled, features_raw):
    """
    Returns a list of human-readable insight strings derived from SHAP values.
    Falls back gracefully if SHAP is unavailable.
    """
    if not SHAP_AVAILABLE or SHAP_EXPLAINER is None:
        return []

    try:
        shap_values = SHAP_EXPLAINER.shap_values(features_scaled)[0]  # shape: (n_features,)
        # Pair feature names with their SHAP contributions
        contributions = list(zip(FEATURE_NAMES, shap_values))
        # Sort by absolute impact, descending
        contributions.sort(key=lambda x: abs(x[1]), reverse=True)

        insights = []
        for name, val in contributions[:5]:  # top-5 most influential features
            direction = "(+)" if val > 0 else "(-)"
            impact    = abs(round(val, 1))
            if impact < 0.5:
                continue
            insights.append(
                f"🧠 SHAP {direction} <b>{name}</b> contributed {'+' if val>0 else '-'}{impact} pts to this prediction."
            )
        return insights
    except Exception as e:
        print(f"[AdLens] SHAP inference error: {e}")
        return []

# ─────────────────────────────────────────────────────────
#  — AD COPY TEXT ANALYZER (pure Python, no deps)
# ─────────────────────────────────────────────────────────
POSITIVE_WORDS = {'amazing','best','free','new','save','win','exclusive','limited','offer',
                  'guaranteed','proven','trusted','top','premium','instant','easy','fast',
                  'boost','grow','improve','powerful','smart','innovative','leading','award',
                  'special','unique','perfect','great','excellent','outstanding','incredible'}
NEGATIVE_WORDS = {'bad','worst','fail','lose','costly','expensive','difficult','hard',
                  'slow','weak','outdated','broken','cheap','risky','uncertain','avoid','worst'}
CTA_WORDS      = {'buy','shop','get','try','start','join','sign','click','learn','discover',
                  'order','subscribe','download','register','book','claim','grab','explore',
                  'call','visit','watch','read','find','see','check','use','apply'}
URGENCY_WORDS  = {'now','today','limited','hurry','last','expires','urgent','soon','only',
                  'left','deadline','ending','final','instant','immediately','quickly'}

def analyze_ad_copy(text):
    if not text or not text.strip():
        return None

    text      = text.strip()
    words     = text.split()
    word_count= len(words)
    text_lower= text.lower()
    word_set  = set(re.findall(r'\b\w+\b', text_lower))

    pos_hits     = word_set & POSITIVE_WORDS
    neg_hits     = word_set & NEGATIVE_WORDS
    cta_hits     = word_set & CTA_WORDS
    urgency_hits = word_set & URGENCY_WORDS

    # Sentiment: ratio of positive vs negative hits, scaled
    raw_sentiment    = (len(pos_hits) - len(neg_hits)) / max(word_count, 1)
    sentiment_score  = round(min(1.0, max(-1.0, raw_sentiment * 10)), 3)

    has_cta          = len(cta_hits) > 0
    has_urgency      = len(urgency_hits) > 0
    has_question     = '?' in text
    exclamation_count= text.count('!')

    # Readability proxy: penalize long average word length
    avg_word_len = sum(len(w) for w in words) / max(word_count, 1)
    readability  = round(min(10, max(1, 10 - avg_word_len + 2)), 1)

    # Copy score
    copy_score = 50
    copy_score += sentiment_score * 20
    copy_score += 10 if has_cta     else -10
    copy_score += 8  if has_urgency else 0
    copy_score += 5  if has_question else 0
    copy_score += min(5, exclamation_count * 2)
    copy_score += (readability - 5) * 2
    if word_count < 3:  copy_score -= 15   # too short
    if word_count > 25: copy_score -= 10   # too long
    # Spam / keyword-stuffing detection
    # Flag if >40% of unique words are positive-power-words (spammy hype)
    spam_ratio = len(pos_hits) / max(len(word_set), 1)
    is_spam_like = spam_ratio > 0.4 and word_count < 15
    if is_spam_like:
        copy_score -= 20  # heavy penalty for keyword stuffing

    copy_score = int(min(100, max(0, copy_score)))

    # Human-readable label
    if sentiment_score > 0.3:   sentiment_label = "Positive"
    elif sentiment_score < -0.3:sentiment_label = "Negative"
    else:                        sentiment_label = "Neutral"

    return {
        'word_count':       word_count,
        'sentiment_score':  sentiment_score,
        'sentiment_label':  sentiment_label,
        'has_cta':          has_cta,
        'cta_words_found':  list(cta_hits)[:3],
        'has_urgency':      has_urgency,
        'urgency_words':    list(urgency_hits)[:3],
        'readability':      readability,
        'has_question':     has_question,
        'copy_score':       copy_score,
        'is_spam_like':     is_spam_like,
        'spam_ratio':       round(spam_ratio, 2),
    }

# ─────────────────────────────────────────────────────────
# CLIP MODEL
# ─────────────────────────────────────────────────────────
device     = "cuda" if torch.cuda.is_available() else "cpu"
clip_model = CLIPModel.from_pretrained("openai/clip-vit-base-patch32").to(device)
clip_proc  = CLIPProcessor.from_pretrained("openai/clip-vit-base-patch32")

# ── Zero-shot label banks for CLIP similarity scoring ─────
_TEXT_DENSITY_LABELS = [
    "an advertisement with heavy text and many words",
    "an advertisement with a logo and minimal text",
    "a clean image with no text at all",
]
_EMOTION_LABELS = [
    "a joyful, happy, vibrant advertisement with smiling people",
    "a neutral, professional, calm advertisement",
    "a dark, dull, boring, or empty advertisement",
]
_CLARITY_LABELS = [
    "a sharp, high-contrast, visually clear image",
    "a blurry, low-contrast, unclear image",
]
_DESCRIPTION_LABELS = [
    "a colorful product advertisement",
    "a technology or gadget advertisement",
    "a food or beverage advertisement",
    "a fashion or lifestyle advertisement",
    "a travel or outdoor advertisement",
    "a business or professional services advertisement",
]

def _clip_scores(image: Image.Image, labels: list) -> list:
    """Return softmax similarity scores for image vs. each text label."""
    inputs = clip_proc(
        text=labels, images=image, return_tensors="pt", padding=True
    ).to(device)
    with torch.no_grad():
        logits = clip_model(**inputs).logits_per_image[0]  # (n_labels,)
    probs = logits.softmax(dim=-1).cpu().tolist()
    return probs

def analyze_image(image_file):
    image = Image.open(image_file).convert('RGB')

    # ── Pixel-level brightness / contrast ─────────────────
    stat       = ImageStat.Stat(image.convert("L"))
    brightness = stat.mean[0]
    contrast   = stat.stddev[0]
    color_score          = min(10, max(1, int((brightness / 255) * 10)))
    visual_clarity_score = min(10, max(1, int((contrast  / 128) * 10)))

    # ── CLIP: text-density ────────────────────────────────
    td_probs = _clip_scores(image, _TEXT_DENSITY_LABELS)
    # heavy-text label wins → high density score
    text_density_score = round(1 + td_probs[0] * 9)
    text_density_score = min(10, max(1, int(text_density_score)))

    # ── CLIP: emotional impact ────────────────────────────
    em_probs = _clip_scores(image, _EMOTION_LABELS)
    if em_probs[0] >= 0.5:           # clearly positive/vibrant
        emotional_impact_score = 9
    elif em_probs[2] >= 0.4:         # dark/dull
        emotional_impact_score = 3
    else:                            # neutral
        emotional_impact_score = 6

    # ── CLIP: best-match description (replaces BLIP caption) ─
    desc_probs  = _clip_scores(image, _DESCRIPTION_LABELS)
    best_idx    = int(torch.tensor(desc_probs).argmax())
    caption     = _DESCRIPTION_LABELS[best_idx]   # human-readable best match

    return {
        "caption":               caption,
        "color_score":           color_score,
        "visual_clarity_score":  visual_clarity_score,
        "text_density_score":    text_density_score,
        "emotional_impact_score":emotional_impact_score,
        "ocr_text":              _run_ocr(image),
    }

def _run_ocr(pil_image):
    """
    Extract text embedded inside the ad image.
    Returns the detected string, or None if OCR is unavailable.
    """
    if OCR_BACKEND is None:
        return None
    try:
        if OCR_BACKEND == "easyocr":
            import numpy as _np
            results = _ocr_reader.readtext(_np.array(pil_image))
            return " ".join(r[1] for r in results).strip() or None
        else:  # pytesseract
            import pytesseract
            text = pytesseract.image_to_string(pil_image).strip()
            return text or None
    except Exception as e:
        print(f"[AdLens] OCR error: {e}")
        return None

# ─────────────────────────────────────────────────────────
#   ( API via urllib — no extra package)
# ─────────────────────────────────────────────────────────
import urllib.request

PLATFORM_LABELS = {1: "Instagram", 2: "YouTube", 3: "Facebook"}
AUDIENCE_LABELS = {1: "Gen Z / Students", 2: "Professionals (B2B)", 3: "General / Broad"}

def ai_rewrite_copy(original_copy, platform_id, audience_id, issues):
    """
    Returns rewritten string, or None if API key is missing / call fails.
    """
    if not ANTHROPIC_API_KEY or not original_copy:
        return None

    platform = PLATFORM_LABELS.get(platform_id, "social media")
    audience = AUDIENCE_LABELS.get(audience_id, "general audience")
    issues_text = "; ".join(issues[:4]) if issues else "improve overall persuasion"

    prompt = (
        f"You are an expert advertising copywriter. "
        f"Rewrite the following ad copy for {platform}, targeting {audience}.\n\n"
        f"Original copy: \"{original_copy}\"\n\n"
        f"Issues to fix: {issues_text}\n\n"
        f"Rules:\n"
        f"- Maximum 12 words\n"
        f"- Include one clear CTA\n"
        f"- Match the platform tone (Instagram = casual/visual, YouTube = direct, Facebook = benefits-led)\n"
        f"- No hype words like 'amazing', 'best ever', 'incredible'\n"
        f"- Return ONLY the rewritten copy, nothing else, no quotes."
    )

    payload = json.dumps({
        "model": "claude-sonnet-4-20250514",
        "max_tokens": 120,
        "messages": [{"role": "user", "content": prompt}]
    }).encode()

    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={
            "Content-Type":      "application/json",
            "x-api-key":         ANTHROPIC_API_KEY,
            "anthropic-version": "2023-06-01",
        },
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
            return data["content"][0]["text"].strip()
    except Exception as e:
        print(f"[AdLens] AI rewrite error: {e}")
        return None


# ─────────────────────────────────────────────────────────
# COMPETITOR SCORING HELPER
# ─────────────────────────────────────────────────────────
def score_image_for_compare(image_file):
    """
    Runs the full image + feature pipeline on a competitor image.
    Returns a dict with all sub-scores so we can display a side-by-side.
    """
    vlm = analyze_image(image_file)
    vlm_score = int(np.mean([
        vlm["color_score"],
        vlm["visual_clarity_score"],
        vlm["text_density_score"],
        vlm["emotional_impact_score"],
    ]) * 10)
    return {
        "vlm_score":           vlm_score,
        "color_score":         vlm["color_score"],
        "clarity_score":       vlm["visual_clarity_score"],
        "text_density_score":  vlm["text_density_score"],
        "emotional_score":     vlm["emotional_impact_score"],
        "caption":             vlm["caption"],
        "ocr_text":            vlm.get("ocr_text"),
    }


# ─────────────────────────────────────────────────────────
HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>AdLens - Analyze your ad</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        :root {
            --bg-primary: #0a0a0c; --bg-secondary: #131417;
            --accent-primary: #ccff00; --accent-glow: rgba(204, 255, 0, 0.3);
            --success: #ccff00; --warning: #facc15; --danger: #f87171;
            --text-main: #ffffff; --text-muted: #a1a1aa;
            --glass-border: rgba(255, 255, 255, 0.08); --btn-text: #000;
        }
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            background-color: var(--bg-primary);
            background-image: linear-gradient(rgba(255,255,255,0.03) 1px,transparent 1px),linear-gradient(90deg,rgba(255,255,255,0.03) 1px,transparent 1px);
            background-size: 50px 50px; color: var(--text-main); font-family: 'Outfit', sans-serif;
            min-height: 100vh; display: flex; flex-direction: column; padding: 0 2rem 3rem 2rem; overflow-x: hidden;
        }
        .container { max-width: 900px; width: 100%; margin: 0 auto; display: flex; flex-direction: column; }
        nav { display: flex; justify-content: space-between; align-items: center; padding: 2rem 0; margin-bottom: 3rem; }
        .logo { font-size: 1.6rem; font-weight: 800; display: flex; align-items: center; gap: 0.5rem; letter-spacing: -0.02em; }
        .logo-icon { background: var(--accent-primary); color: #000; padding: 0.2rem; border-radius: 8px; font-size: 1.2rem; display: flex; align-items: center; justify-content: center; width: 32px; height: 32px; }
        .badge { padding: 0.4rem 1.2rem; border: 1px solid rgba(255,255,255,0.15); background: rgba(255,255,255,0.03); border-radius: 20px; font-size: 0.75rem; color: var(--text-muted); font-weight: 600; letter-spacing: 0.05em; }
        header { text-align: center; margin-bottom: 3.5rem; }
        header h1 { font-size: 5rem; font-weight: 900; line-height: 1.05; margin-bottom: 1.5rem; letter-spacing: -0.03em; text-transform: capitalize; }
        header h1 .highlight { color: var(--accent-primary); display: block; }
        header p { color: var(--text-muted); font-size: 1.15rem; max-width: 650px; margin: 0 auto; line-height: 1.6; font-weight: 300; }
        @media (max-width: 768px) { header h1 { font-size: 3.5rem; } }
        .dashboard { width: 100%; }
        .form-panel { background: rgba(255,255,255,0.01); border: 1px solid transparent; border-radius: 24px; padding: 0; backdrop-filter: blur(10px); }
        .type-selector { display: flex; justify-content: center; gap: 0.8rem; margin-bottom: 2.5rem; flex-wrap: wrap; }
        .radio-pill input { display: none; }
        .radio-pill .pill-content { background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.1); padding: 0.7rem 1.2rem; border-radius: 12px; color: var(--text-muted); cursor: pointer; display: flex; align-items: center; gap: 0.5rem; transition: all 0.2s; font-weight: 500; font-size: 0.95rem; }
        .radio-pill:hover .pill-content { background: rgba(255,255,255,0.08); color: white; }
        .radio-pill input:checked + .pill-content { background: var(--accent-primary); color: var(--btn-text); border-color: var(--accent-primary); font-weight: 600; }
        .upload-area { background: #111113; border: 1px dashed rgba(255,255,255,0.2); border-radius: 20px; padding: 4.5rem 2rem; text-align: center; cursor: pointer; transition: all 0.3s; margin-bottom: 1.5rem; position: relative; overflow: hidden; display: flex; flex-direction: column; align-items: center; justify-content: center; }
        .upload-area:hover { border-color: rgba(255,255,255,0.4); background: #161619; }
        .upload-icon { background: #3b82f6; color: white; width: 54px; height: 54px; border-radius: 14px; display: flex; align-items: center; justify-content: center; font-size: 1.8rem; margin-bottom: 1.5rem; box-shadow: 0 4px 20px rgba(59,130,246,0.4); }
        .upload-title { font-size: 1.6rem; font-weight: 700; margin-bottom: 0.5rem; color: white; letter-spacing: -0.02em; }
        .upload-subtitle { color: var(--text-muted); font-size: 0.95rem; }
        #imageInput { position: absolute; top: 0; left: 0; width: 100%; height: 100%; opacity: 0; cursor: pointer; z-index: 10; }
        #imagePreview { display: none; max-width: 100%; max-height: 400px; border-radius: 12px; box-shadow: 0 10px 40px rgba(0,0,0,0.5); object-fit: contain; }
        .upload-content-wrapper { display: flex; flex-direction: column; align-items: center; z-index: 5; pointer-events: none; }
        /* copy input — matches existing form style exactly */
        .copy-group { margin-bottom: 1.5rem; }
        .copy-group label { display: block; margin-bottom: 0.6rem; color: var(--text-muted); font-size: 0.85rem; font-weight: 500; text-transform: uppercase; letter-spacing: 0.05em; }
        .copy-group textarea { width: 100%; padding: 0.9rem 1rem; background: rgba(0,0,0,0.4); border: 1px solid rgba(255,255,255,0.1); border-radius: 10px; color: white; font-family: 'Outfit', sans-serif; font-size: 0.95rem; transition: all 0.3s ease; resize: vertical; min-height: 80px; }
        .copy-group textarea:focus { outline: none; border-color: var(--accent-primary); }
        .settings-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 1.5rem; margin-bottom: 2rem; padding: 2rem; border: 1px dashed rgba(255,255,255,0.1); border-radius: 16px; background: rgba(0,0,0,0.2); }
        .form-group label { display: block; margin-bottom: 0.6rem; color: var(--text-muted); font-size: 0.85rem; font-weight: 500; text-transform: uppercase; letter-spacing: 0.05em; }
        input, select { width: 100%; padding: 0.9rem 1rem; background: rgba(0,0,0,0.4); border: 1px solid rgba(255,255,255,0.1); border-radius: 10px; color: white; font-family: 'Outfit', sans-serif; font-size: 0.95rem; transition: all 0.3s ease; }
        input:focus, select:focus { outline: none; border-color: var(--accent-primary); }
        .btn-submit { width: 100%; padding: 1.25rem; background: linear-gradient(180deg, #788d37 0%, #516027 100%); color: #fff; border: 1px solid rgba(255,255,255,0.05); border-radius: 16px; font-size: 1.15rem; font-weight: 700; font-family: 'Outfit', sans-serif; cursor: pointer; transition: all 0.3s ease; box-shadow: inset 0 1px 0 rgba(255,255,255,0.15); display: flex; align-items: center; justify-content: center; gap: 0.8rem; }
        .btn-submit:hover { background: linear-gradient(180deg, #879f3e 0%, #5b6d2c 100%); transform: translateY(-1px); }
        .advanced-settings-toggle { text-align: center; color: var(--text-muted); font-size: 0.9rem; cursor: pointer; margin-bottom: 1.5rem; display: flex; align-items: center; justify-content: center; gap: 0.5rem; transition: color 0.2s; padding: 0.5rem; }
        .advanced-settings-toggle:hover { color: #fff; }
        .results-wrapper { margin-top: 4rem; padding-top: 4rem; border-top: 1px solid rgba(255,255,255,0.05); }
        .results-panel { display: flex; flex-direction: column; gap: 1.5rem; }
        .score-container { display: flex; align-items: center; justify-content: center; flex-direction: column; padding: 4rem; background: rgba(255,255,255,0.01); border-radius: 24px; border: 1px solid var(--glass-border); position: relative; }
        .score-circle { width: 200px; height: 200px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 5rem; font-weight: 800; background: conic-gradient(var(--score-color) var(--score-degrees), rgba(255,255,255,0.03) 0); position: relative; box-shadow: 0 0 50px var(--score-glow); animation: fadeInScale 0.8s ease-out forwards; }
        .score-circle::before { content: ""; position: absolute; inset: 12px; background: var(--bg-primary); border-radius: 50%; }
        .score-value { position: relative; z-index: 10; color: #fff; letter-spacing: -0.05em; }
        .feedback-badge { margin-top: 2.5rem; padding: 0.8rem 2.5rem; border-radius: 30px; font-weight: 700; font-size: 1.25rem; background: rgba(0,0,0,0.4); border: 1px solid; animation: slideUp 0.5s ease-out 0.3s both; }
        .metrics-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 1.5rem; }
        .metric-card { background: rgba(255,255,255,0.02); border: 1px solid var(--glass-border); padding: 2.5rem; border-radius: 20px; animation: slideUp 0.5s ease-out both; }
        .list-items { list-style: none; margin-top: 1.5rem; }
        .list-items li { padding: 1.2rem 0; border-bottom: 1px solid rgba(255,255,255,0.03); display: flex; align-items: flex-start; gap: 1rem; font-size: 1.1rem; color: #d4d4d8; line-height: 1.6; }
        .list-items li:last-child { border-bottom: none; }
        .icon-good { color: var(--success); font-size: 1.4rem; }
        .icon-bad  { color: var(--danger);  font-size: 1.4rem; }
        .icon-warn { color: var(--warning); font-size: 1.4rem; }
        /* copy score pill — new, minimal */
        .copy-pill { display:inline-flex; align-items:center; gap:0.5rem; margin-top:0.8rem; padding:0.4rem 1rem; border-radius:20px; font-size:0.85rem; font-weight:600; border:1px solid rgba(255,255,255,0.1); background:rgba(255,255,255,0.04); color:var(--text-muted); }
        .copy-pill.good { border-color:rgba(204,255,0,0.3); color:var(--accent-primary); }
        .copy-pill.warn { border-color:rgba(250,204,21,0.3); color:var(--warning); }
        .copy-pill.bad  { border-color:rgba(248,113,113,0.3); color:var(--danger); }
        @keyframes fadeInScale { from { opacity: 0; transform: scale(0.8); } to { opacity: 1; transform: scale(1); } }
        @keyframes slideUp     { from { opacity: 0; transform: translateY(20px); } to { opacity: 1; transform: translateY(0); } }
        /* ── History table ── */
        .history-table { width:100%; border-collapse:collapse; margin-top:1rem; }
        .history-table th { text-align:left; padding:0.75rem 1rem; font-size:0.78rem; text-transform:uppercase; letter-spacing:0.06em; color:var(--text-muted); border-bottom:1px solid rgba(255,255,255,0.06); font-weight:600; }
        .history-table td { padding:0.9rem 1rem; font-size:0.9rem; color:#d4d4d8; border-bottom:1px solid rgba(255,255,255,0.04); vertical-align:top; }
        .history-table tr:hover td { background:rgba(255,255,255,0.02); }
        .score-chip { display:inline-block; padding:0.25rem 0.75rem; border-radius:20px; font-weight:700; font-size:0.85rem; }
        .score-chip.high { background:rgba(204,255,0,0.15); color:var(--accent-primary); }
        .score-chip.mid  { background:rgba(250,204,21,0.15); color:var(--warning); }
        .score-chip.low  { background:rgba(248,113,113,0.15); color:var(--danger); }
        /* ── Competitor panel ── */
        .compare-grid { display:grid; grid-template-columns:1fr 1fr; gap:1.5rem; }
        .compare-col  { background:rgba(255,255,255,0.02); border:1px solid var(--glass-border); border-radius:16px; padding:1.5rem; }
        .compare-col h4 { font-size:0.9rem; text-transform:uppercase; letter-spacing:0.06em; color:var(--text-muted); margin-bottom:1rem; font-weight:600; }
        .compare-bar-wrap { margin-bottom:0.6rem; }
        .compare-bar-label { display:flex; justify-content:space-between; font-size:0.82rem; color:var(--text-muted); margin-bottom:0.3rem; }
        .compare-bar-bg { background:rgba(255,255,255,0.06); border-radius:4px; height:6px; }
        .compare-bar-fill { height:6px; border-radius:4px; transition:width 0.6s ease; }
        .bar-you  { background:var(--accent-primary); }
        .bar-them { background:#6366f1; }
        /* ── AI Rewrite card ── */
        .rewrite-card { background:rgba(99,102,241,0.08); border:1px solid rgba(99,102,241,0.25); border-radius:16px; padding:1.5rem; margin-top:1rem; }
        .rewrite-copy { font-size:1.3rem; font-weight:700; color:#fff; line-height:1.4; margin:0.75rem 0; }
        .rewrite-note { font-size:0.82rem; color:var(--text-muted); }
    </style>
</head>
<body>
<div class="container">
    <nav>
        <div class="logo">
            <span class="logo-icon">🔍</span>
            <span>Ad<span style="color: var(--accent-primary);">Lens</span></span>
        </div>
        <div class="badge">AI-POWERED</div>
        <a href="/history" style="font-size:0.85rem;color:var(--text-muted);text-decoration:none;padding:0.4rem 1rem;border:1px solid rgba(255,255,255,0.1);border-radius:20px;transition:color 0.2s;" onmouseover="this.style.color='#fff'" onmouseout="this.style.color=''">📋 History</a>
    </nav>
    <header>
        <h1>
            <div>Analyze your ad.</div>
            <div class="highlight">Improve it instantly.</div>
        </h1>
        <p>Upload any ad — poster, banner, or micro-site screenshot — and get an AI-powered audit with scores, suggestions, and actionable improvements.</p>
    </header>
    <div class="dashboard">
        <div class="form-panel">
            <form id="adForm" method="POST" enctype="multipart/form-data">
                <div class="type-selector">
                    <label class="radio-pill"><input type="radio" name="ad_type" value="1" {% if request.method == 'GET' or ad_type == 1 %}checked{% endif %}><span class="pill-content">📱 Poster</span></label>
                    <label class="radio-pill"><input type="radio" name="ad_type" value="3" {% if request.method == 'POST' and ad_type == 3 %}checked{% endif %}><span class="pill-content">📐 Banner</span></label>
                    <label class="radio-pill"><input type="radio" name="ad_type" value="1" {% if request.method == 'POST' and ad_type == 1 and request.form.get('ad_type_raw') == 'Micro Website' %}checked{% endif %}><span class="pill-content">🌐 Micro Website</span></label>
                    <label class="radio-pill"><input type="radio" name="ad_type" value="2" {% if request.method == 'POST' and ad_type == 2 %}checked{% endif %}><span class="pill-content">📱 Social Ad</span></label>
                    <label class="radio-pill"><input type="radio" name="ad_type" value="1" {% if request.method == 'POST' and ad_type == 1 and request.form.get('ad_type_raw') == 'Email Ad' %}checked{% endif %}><span class="pill-content">✉️ Email Ad</span></label>
                </div>

                <!-- FIX 3: Ad copy input -->
                <div class="copy-group">
                    <label>✍️ Ad Headline / Copy (optional)</label>
                    <textarea name="ad_copy" placeholder="e.g. Get 50% off today only — limited time offer!">{{ request.form.get('ad_copy', '') }}</textarea>
                </div>

                <div class="upload-area" id="dropZone">
                    <input type="file" id="imageInput" name="ad_image" accept="image/png, image/jpeg, image/webp" onchange="previewImage(event)">
                    <div class="upload-content-wrapper" id="uploadPlaceholder">
                        <div class="upload-icon">⬆️</div>
                        <div class="upload-title">Drop your ad here</div>
                        <div class="upload-subtitle">PNG, JPG, WEBP — up to 10MB</div>
                    </div>
                    <img id="imagePreview" src="" alt="Ad Preview">
                </div>

                <div class="advanced-settings-toggle" onclick="toggleSettings()">
                    <span>⚙️ Advanced Settings</span> <span id="toggleIcon">▼</span>
                </div>
                <div class="settings-grid" id="advancedSettings" style="display: none;">
                    <div class="form-group">
                        <label>Color Style</label>
                        <select name="color">
                            <option value="3" {% if request.form.get('color') == '3' %}selected{% endif %}>Dark & Moody</option>
                            <option value="6" {% if not request.form.get('color') or request.form.get('color') == '6' %}selected{% endif %}>Balanced/Neutral</option>
                            <option value="9" {% if request.form.get('color') == '9' %}selected{% endif %}>Bright & Vibrant</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Target Audience</label>
                        <select name="audience">
                            <option value="1" {% if request.form.get('audience') == '1' %}selected{% endif %}>Students / Gen Z</option>
                            <option value="2" {% if request.form.get('audience') == '2' %}selected{% endif %}>Professionals (B2B)</option>
                            <option value="3" {% if not request.form.get('audience') or request.form.get('audience') == '3' %}selected{% endif %}>General / Broad</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Platform</label>
                        <select name="platform">
                            <option value="1" {% if not request.form.get('platform') or request.form.get('platform') == '1' %}selected{% endif %}>Instagram</option>
                            <option value="2" {% if request.form.get('platform') == '2' %}selected{% endif %}>YouTube</option>
                            <option value="3" {% if request.form.get('platform') == '3' %}selected{% endif %}>Facebook</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Day of Week</label>
                        <select name="day">
                            <option value="1" {% if request.form.get('day') == '1' %}selected{% endif %}>Monday</option>
                            <option value="2" {% if request.form.get('day') == '2' %}selected{% endif %}>Tuesday</option>
                            <option value="3" {% if not request.form.get('day') or request.form.get('day') == '3' %}selected{% endif %}>Wednesday</option>
                            <option value="4" {% if request.form.get('day') == '4' %}selected{% endif %}>Thursday</option>
                            <option value="5" {% if request.form.get('day') == '5' %}selected{% endif %}>Friday</option>
                            <option value="6" {% if request.form.get('day') == '6' %}selected{% endif %}>Saturday</option>
                            <option value="7" {% if request.form.get('day') == '7' %}selected{% endif %}>Sunday</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Posting Hour (0-23)</label>
                        <input type="number" name="time" min="0" max="23" value="{{ request.form.get('time', '18') }}" required>
                    </div>
                </div>
                <button type="submit" class="btn-submit"><span>🔬</span> Analyze My Ad</button>

                <!-- Competitor benchmarking (optional) -->
                <details style="margin-top:1.5rem;">
                    <summary style="cursor:pointer;color:var(--text-muted);font-size:0.9rem;font-weight:600;padding:0.5rem 0;list-style:none;display:flex;align-items:center;gap:0.5rem;">
                        ⚡ Compare against a competitor ad (optional)
                    </summary>
                    <div style="margin-top:1rem;padding:1rem;background:rgba(99,102,241,0.06);border:1px dashed rgba(99,102,241,0.3);border-radius:12px;">
                        <label style="display:block;margin-bottom:0.5rem;font-size:0.82rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-muted);font-weight:600;">Competitor Ad Image</label>
                        <input type="file" name="competitor_image" accept="image/png,image/jpeg,image/webp"
                            style="font-size:0.85rem;color:var(--text-muted);cursor:pointer;">
                        <p style="font-size:0.78rem;color:var(--text-muted);margin-top:0.5rem;">Upload a competitor's ad to get a side-by-side visual score breakdown.</p>
                    </div>
                </details>
            </form>
        </div>

        <script>
            function previewImage(event) {
                const reader = new FileReader();
                reader.onload = function(){
                    const output = document.getElementById('imagePreview');
                    const placeholder = document.getElementById('uploadPlaceholder');
                    const dropZone = document.getElementById('dropZone');
                    output.src = reader.result;
                    output.style.display = 'block';
                    placeholder.style.display = 'none';
                    dropZone.style.padding = '1rem';
                };
                if(event.target.files[0]) { reader.readAsDataURL(event.target.files[0]); }
            }
            function toggleSettings() {
                const settings = document.getElementById('advancedSettings');
                const icon = document.getElementById('toggleIcon');
                if(settings.style.display === 'none') { settings.style.display = 'grid'; icon.innerHTML = '▲'; }
                else { settings.style.display = 'none'; icon.innerHTML = '▼'; }
            }
        </script>

        {% if score is defined %}
        <div class="results-wrapper" id="results">
            <h2 style="text-align: center; margin-bottom: 2.5rem; font-size: 2.5rem; font-weight: 800; text-transform: capitalize;">Analysis Results</h2>
            <div class="results-panel">
                <div class="score-container" style="--score-degrees: {{ score * 3.6 }}deg; --score-color: {{ color_hex }}; --score-glow: {{ color_glow }};">
                    <h3 style="margin-bottom: 2rem; position: relative; z-index: 10; font-size: 1.1rem; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.1em; font-weight: 600;">Final Score</h3>
                    <div class="score-circle"><span class="score-value">{{ score }}</span></div>
                    <div class="feedback-badge" style="color: {{ color_hex }}; border-color: {{ color_glow }};">{{ feedback }}</div>
                </div>

                <!-- Score breakdown grid -->
                <div class="metrics-grid">
                    <div class="metric-card" style="text-align: center;">
                        <span style="display: block; font-size: 1rem; color: var(--text-muted); margin-bottom: 1rem; text-transform: uppercase; letter-spacing: 0.05em; font-weight: 600;">🤖 AI Algorithm Score</span>
                        <span style="font-size: 3.5rem; font-weight: 800; color: #d4d4d8;">{{ ml_score }}</span>
                        <div style="font-size:0.78rem;color:var(--text-muted);margin-top:0.5rem;">XGBoost · {{ train_size }} training samples</div>
                        <div style="font-size:0.75rem;color:var(--text-muted);margin-top:0.2rem;">MAE {{ model_mae }} · R² {{ model_r2 }}</div>
                    </div>
                    {% if vlm_score is defined and vlm_score %}
                    <div class="metric-card" style="text-align: center;">
                        <span style="display: block; font-size: 1rem; color: var(--text-muted); margin-bottom: 1rem; text-transform: uppercase; letter-spacing: 0.05em; font-weight: 600;">👁️ Visual Impact Score</span>
                        <span style="font-size: 3.5rem; font-weight: 800; color: var(--accent-primary);">{{ vlm_score }}</span>
                        <div style="font-size:0.78rem;color:var(--text-muted);margin-top:0.5rem;">CLIP · Pitt Image Ads signals</div>
                    </div>
                    {% else %}
                    <div class="metric-card" style="text-align: center; opacity: 0.4;">
                        <span style="display: block; font-size: 1rem; color: var(--text-muted); margin-bottom: 1rem; text-transform: uppercase; letter-spacing: 0.05em; font-weight: 600;">👁️ Visual Impact Score</span>
                        <span style="font-size: 1.5rem; font-weight: 600; color: var(--text-muted);">No image</span>
                        <div style="font-size:0.78rem;color:var(--text-muted);margin-top:0.5rem;">Upload an ad image to enable</div>
                    </div>
                    {% endif %}
                </div>

                {% if copy_result %}
                <!-- Copy analysis card -->
                <div class="metric-card">
                    <h3 style="color:#fff;font-size:1.4rem;font-weight:700;margin-bottom:1rem;display:flex;align-items:center;gap:0.75rem;">✍️ Ad Copy Analysis</h3>
                    <div style="display:flex;flex-wrap:wrap;gap:0.6rem;margin-bottom:1rem;">
                        <span class="copy-pill {% if copy_result.copy_score >= 70 %}good{% elif copy_result.copy_score >= 45 %}warn{% else %}bad{% endif %}">
                            Copy Score: {{ copy_result.copy_score }}/100
                        </span>
                        <span class="copy-pill {% if copy_result.sentiment_label == 'Positive' %}good{% elif copy_result.sentiment_label == 'Negative' %}bad{% else %}warn{% endif %}">
                            {{ copy_result.sentiment_label }} Tone
                        </span>
                        {% if copy_result.has_cta %}<span class="copy-pill good">✓ CTA Detected</span>{% else %}<span class="copy-pill bad">✗ No CTA</span>{% endif %}
                        {% if copy_result.has_urgency %}<span class="copy-pill good">⚡ Urgency</span>{% endif %}
                        {% if copy_result.has_question %}<span class="copy-pill warn">? Engagement Hook</span>{% endif %}
                        {% if copy_result.is_spam_like %}<span class="copy-pill bad">🚨 Spam-like Copy</span>{% endif %}
                        <span class="copy-pill">{{ copy_result.word_count }} words</span>
                        <span class="copy-pill">Readability {{ copy_result.readability }}/10</span>
                    </div>
                    {% if copy_result.is_spam_like %}
                    <div style="background:rgba(248,113,113,0.1);border:1px solid rgba(248,113,113,0.3);border-radius:10px;padding:0.75rem 1rem;font-size:0.88rem;color:#f87171;margin-top:0.5rem;">
                        ⚠️ Keyword stuffing detected ({{ (copy_result.spam_ratio * 100)|int }}% power-word density). Rewrite with a specific value proposition instead of stacking hype words.
                    </div>
                    {% endif %}
                </div>
                {% endif %}

                {% if vlm_results and vlm_results.ocr_text %}
                <!-- OCR results card -->
                <div class="metric-card">
                    <h3 style="color:#fff;font-size:1.4rem;font-weight:700;margin-bottom:1rem;display:flex;align-items:center;gap:0.75rem;">🔤 OCR — Text Detected in Image</h3>
                    <p style="color:var(--text-muted);font-size:0.9rem;margin-bottom:0.75rem;">Text extracted directly from your ad image:</p>
                    <div style="background:rgba(0,0,0,0.4);border:1px solid rgba(255,255,255,0.1);border-radius:10px;padding:0.75rem 1rem;font-size:0.9rem;color:#d4d4d8;font-family:monospace;white-space:pre-wrap;">{{ vlm_results.ocr_text }}</div>
                </div>
                {% endif %}

                <div class="metrics-grid" style="grid-template-columns: 1fr;">
                    <div class="metric-card">
                        <h3 style="color: #fff; font-size: 1.4rem; font-weight: 700; margin-bottom: 1.5rem; display: flex; align-items: center; gap: 0.75rem;">💡 Optimization Suggestions</h3>
                        <ul class="list-items">
                            {% for s in suggestions %}<li><span class="icon-warn">⚡</span> <span>{{ s }}</span></li>{% endfor %}
                            {% if not suggestions %}<li><span class="icon-good">✓</span> <span>Your ad looks perfect! No major optimizations needed.</span></li>{% endif %}
                        </ul>
                    </div>
                    <div class="metric-card">
                        <h3 style="color: #fff; font-size: 1.4rem; font-weight: 700; margin-bottom: 1.5rem; display: flex; align-items: center; gap: 0.75rem;">🔍 Explainable Insights</h3>
                        <ul class="list-items">
                            {% for i in insights %}
                            <li>
                                {% if "(+" in i %}<span class="icon-good">📈</span>
                                {% elif "(-" in i %}<span class="icon-bad">📉</span>
                                {% else %}<span>👁️</span>{% endif %}
                                <span>{{ i }}</span>
                            </li>
                            {% endfor %}
                        </ul>
                    </div>
                </div>
            </div>

                <!-- AI Rewrite card -->
                {% if rewrite %}
                <div class="metric-card rewrite-card" style="margin-top:0;">
                    <h3 style="color:#a5b4fc;font-size:1.1rem;font-weight:700;display:flex;align-items:center;gap:0.6rem;">✨ AI-Rewritten Copy</h3>
                    <p class="rewrite-note" style="margin-top:0.4rem;">Claude rewrote your copy based on the issues found above:</p>
                    <div class="rewrite-copy">"{{ rewrite }}"</div>
                    <p class="rewrite-note">Targeting <b>{{ platform_label }}</b> · <b>{{ audience_label }}</b></p>
                </div>
                {% elif ad_copy and not anthropic_key_set %}
                <div class="metric-card" style="opacity:0.5;margin-top:0;">
                    <p style="font-size:0.9rem;color:var(--text-muted);">✨ <b>AI Copy Rewrite</b> — Set <code>ANTHROPIC_API_KEY</code> environment variable to enable Claude-powered rewrites.</p>
                </div>
                {% endif %}

                <!-- Competitor benchmarking card -->
                {% if competitor %}
                <div class="metric-card" style="margin-top:0;">
                    <h3 style="color:#fff;font-size:1.4rem;font-weight:700;margin-bottom:1.5rem;display:flex;align-items:center;gap:0.75rem;">⚔️ Competitor Benchmarking</h3>
                    <div class="compare-grid">
                        <div class="compare-col">
                            <h4>🟢 Your Ad</h4>
                            {% for metric, val in your_metrics.items() %}
                            <div class="compare-bar-wrap">
                                <div class="compare-bar-label"><span>{{ metric }}</span><span>{{ val }}/10</span></div>
                                <div class="compare-bar-bg"><div class="compare-bar-fill bar-you" style="width:{{ val * 10 }}%"></div></div>
                            </div>
                            {% endfor %}
                            <div style="margin-top:1rem;font-size:1.5rem;font-weight:800;color:var(--accent-primary);">{{ vlm_score or '—' }}</div>
                            <div style="font-size:0.8rem;color:var(--text-muted);">Visual Score</div>
                        </div>
                        <div class="compare-col">
                            <h4>🟣 Competitor</h4>
                            {% for metric, val in competitor_metrics.items() %}
                            <div class="compare-bar-wrap">
                                <div class="compare-bar-label"><span>{{ metric }}</span><span>{{ val }}/10</span></div>
                                <div class="compare-bar-bg"><div class="compare-bar-fill bar-them" style="width:{{ val * 10 }}%"></div></div>
                            </div>
                            {% endfor %}
                            <div style="margin-top:1rem;font-size:1.5rem;font-weight:800;color:#6366f1;">{{ competitor.vlm_score }}</div>
                            <div style="font-size:0.8rem;color:var(--text-muted);">Visual Score</div>
                        </div>
                    </div>
                    <div style="margin-top:1.25rem;padding:1rem;background:rgba(0,0,0,0.3);border-radius:10px;font-size:0.9rem;color:#d4d4d8;">
                        {% if (vlm_score or 0) > competitor.vlm_score %}
                        🏆 <b>Your ad scores higher visually</b> by {{ (vlm_score or 0) - competitor.vlm_score }} pts. Focus on copy and timing to maximise the lead.
                        {% elif (vlm_score or 0) < competitor.vlm_score %}
                        📉 <b>Competitor's visual score is higher</b> by {{ competitor.vlm_score - (vlm_score or 0) }} pts. Improve clarity and emotional impact of your image.
                        {% else %}
                        ⚖️ <b>Visual scores are tied.</b> Differentiate through copy, CTA, and audience targeting.
                        {% endif %}
                        {% if competitor.caption %}
                        <br><br>CLIP best-match for competitor image: <i>"{{ competitor.caption }}"</i>
                        {% endif %}
                    </div>
                </div>
                {% endif %}

            <script>
                window.onload = function() {
                    setTimeout(() => { document.getElementById('results').scrollIntoView({ behavior: 'smooth', block: 'start' }); }, 100);
                }
            </script>
        </div>
        {% endif %}
    </div>
</div>
</body>
</html>
"""

HISTORY_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>AdLens — History</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        :root { --bg-primary:#0a0a0c; --accent-primary:#ccff00; --text-main:#ffffff; --text-muted:#a1a1aa; --glass-border:rgba(255,255,255,0.08); }
        * { box-sizing:border-box; margin:0; padding:0; }
        body { background-color:var(--bg-primary); background-image:linear-gradient(rgba(255,255,255,0.03) 1px,transparent 1px),linear-gradient(90deg,rgba(255,255,255,0.03) 1px,transparent 1px); background-size:50px 50px; color:var(--text-main); font-family:'Outfit',sans-serif; min-height:100vh; padding:0 2rem 3rem; }
        .container { max-width:1000px; margin:0 auto; }
        nav { display:flex; justify-content:space-between; align-items:center; padding:2rem 0; margin-bottom:2rem; }
        .logo { font-size:1.4rem; font-weight:800; }
        h1 { font-size:2.5rem; font-weight:800; margin-bottom:0.5rem; }
        .sub { color:var(--text-muted); margin-bottom:2rem; }
        .card { background:rgba(255,255,255,0.02); border:1px solid var(--glass-border); border-radius:20px; padding:2rem; overflow-x:auto; }
        table { width:100%; border-collapse:collapse; }
        th { text-align:left; padding:0.75rem 1rem; font-size:0.78rem; text-transform:uppercase; letter-spacing:0.06em; color:var(--text-muted); border-bottom:1px solid rgba(255,255,255,0.06); }
        td { padding:0.9rem 1rem; font-size:0.9rem; color:#d4d4d8; border-bottom:1px solid rgba(255,255,255,0.04); vertical-align:top; max-width:220px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
        tr:hover td { background:rgba(255,255,255,0.02); }
        .chip { display:inline-block; padding:0.25rem 0.7rem; border-radius:20px; font-weight:700; font-size:0.82rem; }
        .high { background:rgba(204,255,0,0.15); color:#ccff00; }
        .mid  { background:rgba(250,204,21,0.15); color:#facc15; }
        .low  { background:rgba(248,113,113,0.15); color:#f87171; }
        .btn-back { display:inline-block; padding:0.5rem 1.25rem; border:1px solid rgba(255,255,255,0.12); border-radius:20px; color:var(--text-muted); text-decoration:none; font-size:0.85rem; font-weight:600; transition:color 0.2s; }
        .btn-back:hover { color:#fff; }
        .empty { text-align:center; padding:4rem; color:var(--text-muted); }
    </style>
</head>
<body>
<div class="container">
    <nav>
        <div class="logo">Ad<span style="color:var(--accent-primary);">Lens</span></div>
        <a href="/" class="btn-back">← Back to Analyzer</a>
    </nav>
    <h1>📋 Analysis History</h1>
    <p class="sub">Last {{ rows|length }} analyses — stored locally in SQLite.</p>
    <div class="card">
    {% if rows %}
    <table>
        <thead>
            <tr>
                <th>#</th><th>Date &amp; Time</th><th>Ad Copy</th>
                <th>Score</th><th>ML</th><th>Visual</th><th>Copy</th>
                <th>Platform</th><th>Audience</th><th>Feedback</th>
            </tr>
        </thead>
        <tbody>
        {% for r in rows %}
        <tr>
            <td style="color:var(--text-muted);">{{ r.id }}</td>
            <td style="white-space:nowrap;color:var(--text-muted);">{{ r.ts }}</td>
            <td title="{{ r.ad_copy }}">{{ (r.ad_copy or '—')[:40] }}{% if r.ad_copy and r.ad_copy|length > 40 %}…{% endif %}</td>
            <td>
                <span class="chip {% if r.score >= 80 %}high{% elif r.score >= 60 %}mid{% else %}low{% endif %}">
                    {{ r.score }}
                </span>
            </td>
            <td>{{ r.ml_score or '—' }}</td>
            <td>{{ r.vlm_score or '—' }}</td>
            <td>{{ r.copy_score or '—' }}</td>
            <td>{{ {1:'Instagram',2:'YouTube',3:'Facebook'}.get(r.platform,'?') }}</td>
            <td>{{ {1:'Gen Z',2:'B2B',3:'General'}.get(r.audience,'?') }}</td>
            <td style="color:var(--text-muted);font-size:0.82rem;">{{ r.feedback }}</td>
        </tr>
        {% endfor %}
        </tbody>
    </table>
    {% else %}
    <div class="empty">No analyses yet. Run your first ad through the analyzer!</div>
    {% endif %}
    </div>
</div>
</body>
</html>
"""

# ─────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────
@app.route("/history")
def history():
    rows = load_history(50)
    return render_template_string(HISTORY_HTML, rows=rows)
@app.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        ad_type  = int(request.form["ad_type"])
        color    = int(request.form["color"])
        audience = int(request.form["audience"])
        platform = int(request.form["platform"])
        day      = int(request.form.get("day", 1))
        time     = int(request.form.get("time", 18))
        ad_copy  = request.form.get("ad_copy", "").strip()

        suggestions, insights = [], []
        vlm_score, vlm_results = None, None

        # FIX 1 — dynamic headline length from actual copy
        headline_length = len(ad_copy.split()) if ad_copy else 8

        # FIX 3 — analyze ad copy text
        copy_result = analyze_ad_copy(ad_copy) if ad_copy else None
        if copy_result:
            if not copy_result['has_cta']:
                suggestions.append("Add a clear Call-to-Action (e.g. 'Shop Now', 'Get Started') — ads with CTAs get 2× higher CTR.")
            if not copy_result['has_urgency']:
                suggestions.append("Add urgency words ('today only', 'limited offer') to drive faster conversions.")
            if copy_result['sentiment_label'] == 'Negative':
                suggestions.append("Your copy has negative language. Reframe around benefits and positive outcomes.")
            if copy_result['word_count'] > 25:
                suggestions.append("Your headline is too long. Cut to under 10 words for better readability on mobile.")
            if copy_result['word_count'] < 3:
                suggestions.append("Your copy is too short. Add more context or a value proposition.")
            if copy_result['readability'] < 5:
                suggestions.append("Simplify your language — use shorter, common words for broader audience reach.")
            if copy_result['copy_score'] >= 70:
                insights.append(f"(+) Ad copy score {copy_result['copy_score']}/100 — strong headline with good persuasive signals.")
            elif copy_result['copy_score'] >= 45:
                insights.append(f"👁️ Ad copy score {copy_result['copy_score']}/100 — decent but missing some key persuasion elements.")
            else:
                insights.append(f"(-) Ad copy score {copy_result['copy_score']}/100 — weak persuasive structure. Revise using CTA + positive framing.")

        # VLM branch (CLIP image analysis)
        if 'ad_image' in request.files and request.files['ad_image'].filename != '':
            try:
                vlm_results = analyze_image(request.files['ad_image'])
                vlm_score   = int(np.mean([
                    vlm_results["color_score"],
                    vlm_results["visual_clarity_score"],
                    vlm_results["text_density_score"],
                    vlm_results["emotional_impact_score"]
                ]) * 10)
                insights.append(f"CLIP best-match: '{vlm_results['caption'].capitalize()}'")
                insights.append(f"Visual Clarity Score: {vlm_results['visual_clarity_score']}/10")
                insights.append(f"Emotional Impact Score: {vlm_results['emotional_impact_score']}/10")
                if vlm_results['visual_clarity_score'] < 5:
                    suggestions.append("Upload a higher-contrast/clearer image to make main elements pop.")
                if vlm_results['emotional_impact_score'] < 5:
                    suggestions.append("Image feels a bit plain. Try adding human faces or vibrant colors.")
                if vlm_results['text_density_score'] >= 8:
                    suggestions.append("Image contains heavy text. Platforms may flag ads with >20% text coverage.")
            except Exception as e:
                print(f"VLM error: {e}")
                insights.append("Failed to process image through CLIP.")

        # Engineered features (hybrid dataset signals)
        text_d = vlm_results["text_density_score"]    if vlm_results else 5
        emo    = vlm_results["emotional_impact_score"] if vlm_results else 6
        apm, phf, vsb = get_engineered_features(ad_type, audience, platform, time, text_d, emo)

        # FIX 1+2 — ML branch with real headline_length + trained on 80% split
        features = np.array([[headline_length, ad_type, color, audience, platform, time, day, apm, phf, vsb]])
        features_scaled = scaler.transform(features)
        # ── FIX: deterministic — no random noise ──────────────────
        ml_score = int(np.clip(model.predict(features_scaled)[0], 5, 100))

        # ── SHAP explainability insights ──────────────────────────
        shap_insights = get_shap_insights(features_scaled, features[0])
        insights.extend(shap_insights)

        # ── OCR insight (if image was uploaded) ───────────────────
        if vlm_results and vlm_results.get("ocr_text"):
            ocr_text = vlm_results["ocr_text"]
            word_count_ocr = len(ocr_text.split())
            if word_count_ocr > 20:
                suggestions.append(
                    f"OCR detected {word_count_ocr} words inside your image — that's text-heavy. "
                    "Platforms flag ads with >20% text coverage. Simplify the image copy."
                )
                insights.append(f"(-) OCR read {word_count_ocr} words in the image — text overload risk detected.")
            elif word_count_ocr > 0:
                insights.append(f"(+) OCR detected image text ({word_count_ocr} words) — within acceptable range.")

        # ── Spam-like copy warning ────────────────────────────────
        if copy_result and copy_result.get("is_spam_like"):
            suggestions.append(
                "Your copy reads like keyword-stuffed hype (e.g. 'BEST AMAZING FREE LIMITED NOW!'). "
                "Ad platforms and audiences distrust this style — rewrite with specific value propositions."
            )
            insights.append("(-) Spam signal: high density of power-words detected in short copy — persuasion score penalized.")
        if vlm_score is not None:
            score = int(np.clip(0.6 * ml_score + 0.4 * vlm_score, 0, 100))
        else:
            score = ml_score

        # Boost/penalize from copy score if available
        if copy_result:
            copy_delta = (copy_result['copy_score'] - 50) * 0.1
            score = int(np.clip(score + copy_delta, 0, 100))

        # Engineered feature insights
        if apm == 1:
            insights.append("(+) Audience–platform pairing matches high-synergy combos in the hybrid dataset.")
        else:
            suggestions.append("Reconsider platform — your audience performs better elsewhere per dataset patterns.")
            insights.append("(-) Audience–platform mismatch: lower engagement predicted from hybrid dataset.")

        if phf == 1:
            insights.append("(+) Posting during peak hours (18:00–21:00) — Avazu CTR data shows highest clicks here.")
        else:
            suggestions.append(f"Reschedule to 18:00–21:00 for better reach. Hour {time}:00 is off-peak.")
            insights.append(f"(-) Off-peak posting at {time}:00 — lower CTR expected outside evening window.")

        if vsb > 3:
            insights.append(f"(+) Visual sentiment boost: {vsb} — strong emotional signal relative to text density.")
        elif vsb < 0:
            suggestions.append("Reduce text density — text overload suppresses emotional recall (Pitt Image Ads data).")
            insights.append(f"(-) Visual sentiment boost: {vsb} — text is outweighing emotional impact.")

        if platform == 1 and ad_type != 2:
            suggestions.append("Consider turning this into a video/Reel for Instagram.")
            insights.append("(+) Video outperforms static images on Instagram (+15% potential).")

        if audience == 2:
            if platform == 1:
                suggestions.append("Run this on LinkedIn/Facebook instead for targeted B2B.")
                insights.append("(-) Instagram has lower engagement for B2B audiences (-10%).")
                score = int(np.clip(score - 10, 0, 100))
            if day in [6, 7]:
                suggestions.append("B2B audiences disengage on weekends. Reschedule to Tue–Thu.")
                insights.append("(-) Weekend B2B posting drops reach by ~15%.")
                score = int(np.clip(score - 10, 0, 100))
        elif audience in [1, 3]:
            if day in [6, 7]:
                insights.append("(+) Weekends boost engagement for General/Student audiences (+10%).")
                score = int(np.clip(score + 8, 0, 100))

        if time < 7 or (time > 13 and time < 17):
            suggestions.append("Reschedule to evening peak times (18:00–21:00).")
            insights.append(f"(-) Posting at {time}:00 works against algorithmic reach (-8%).")
            score = int(np.clip(score - 5, 0, 100))

        score = min(100, max(0, score))

        if score >= 80:
            feedback = "🔥 Excellent Ad Potential"
            color_hex, color_glow = "#ccff00", "rgba(204, 255, 0, 0.5)"
        elif score >= 60:
            feedback = "👍 Good, But Could Improve"
            color_hex, color_glow = "#facc15", "rgba(250, 204, 21, 0.5)"
        else:
            feedback = "⚠️ Needs Optimization"
            color_hex, color_glow = "#f87171", "rgba(248, 113, 113, 0.5)"

        # ── Competitor benchmarking ───────────────────────────────
        competitor = None
        your_metrics = competitor_metrics = {}
        if 'competitor_image' in request.files and request.files['competitor_image'].filename != '':
            try:
                competitor = score_image_for_compare(request.files['competitor_image'])
                METRIC_KEYS = {
                    "Color Score":    "color_score",
                    "Clarity":        "clarity_score",
                    "Text Density":   "text_density_score",
                    "Emotion":        "emotional_score",
                }
                your_metrics = {
                    label: (vlm_results[key] if vlm_results else 5)
                    for label, key in [
                        ("Color Score",  "color_score"),
                        ("Clarity",      "visual_clarity_score"),
                        ("Text Density", "text_density_score"),
                        ("Emotion",      "emotional_impact_score"),
                    ]
                }
                competitor_metrics = {
                    "Color Score":  competitor["color_score"],
                    "Clarity":      competitor["clarity_score"],
                    "Text Density": competitor["text_density_score"],
                    "Emotion":      competitor["emotional_score"],
                }
            except Exception as e:
                print(f"[AdLens] Competitor scoring error: {e}")

        # ── AI rewrite (Claude API) ───────────────────────────────
        rewrite = None
        copy_issues = [s for s in suggestions if any(
            kw in s.lower() for kw in ['cta','urgency','negative','long','short','simplify','spam','hype']
        )]
        if ad_copy:
            rewrite = ai_rewrite_copy(ad_copy, platform, audience, copy_issues)

        # ── Save to SQLite history ────────────────────────────────
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        save_analysis(
            ts=ts, ad_copy=ad_copy, ad_type=ad_type, audience=audience,
            platform=platform, score=score, ml_score=ml_score,
            vlm_score=vlm_score,
            copy_score=copy_result['copy_score'] if copy_result else None,
            feedback=feedback, insights=insights, suggestions=suggestions
        )

        return render_template_string(
            HTML,
            score=score, ml_score=ml_score, vlm_score=vlm_score,
            feedback=feedback, color_hex=color_hex, color_glow=color_glow,
            suggestions=suggestions, insights=insights,
            copy_result=copy_result, vlm_results=vlm_results,
            ad_type=ad_type, color=color, audience=audience, platform=platform, day=day,
            train_size=TRAIN_SIZE, model_mae=MODEL_MAE, model_r2=MODEL_R2,
            # new vars
            rewrite=rewrite,
            ad_copy=ad_copy,
            anthropic_key_set=bool(ANTHROPIC_API_KEY),
            platform_label=PLATFORM_LABELS.get(platform, "Social"),
            audience_label=AUDIENCE_LABELS.get(audience, "General"),
            competitor=competitor,
            your_metrics=your_metrics,
            competitor_metrics=competitor_metrics,
        )

    return render_template_string(HTML)

if __name__ == "__main__":
    app.run(debug=True)
